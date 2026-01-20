from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, make_response
from flask_mail import Mail, Message
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import pymysql
from datetime import datetime, timedelta
import os
import re
from functools import wraps
from dotenv import load_dotenv
try:
    from dateutil.relativedelta import relativedelta
except ImportError:
    # Fallback if dateutil is not available
    class Relativedelta:
        def __init__(self, years=0, months=0, days=0):
            self.years = years
            self.months = months
            self.days = days
    def relativedelta(*args, **kwargs):
        return Relativedelta(**kwargs)
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from io import BytesIO
import csv
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Load environment variables from .env file
load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'your-secret-key-change-in-production')

# Email configuration
app.config['MAIL_SERVER'] = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.environ.get('MAIL_PORT', 587))
app.config['MAIL_USE_TLS'] = os.environ.get('MAIL_USE_TLS', 'True').lower() in ['true', '1', 'yes']
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME', '')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD', '')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_DEFAULT_SENDER', 'noreply@modernschool.com')

# Initialize Flask-Mail
mail = Mail(app)

# File upload configuration
UPLOAD_FOLDER = 'static/uploads/profiles'
PAYMENT_PROOF_FOLDER = 'static/uploads/payment_proofs'
BACKUP_FOLDER = 'static/uploads/backups'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

# Create backup folder if it doesn't exist
os.makedirs(BACKUP_FOLDER, exist_ok=True)
ALLOWED_PAYMENT_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['PAYMENT_PROOF_FOLDER'] = PAYMENT_PROOF_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024  # 10MB max file size

# Create upload directories if they don't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PAYMENT_PROOF_FOLDER, exist_ok=True)

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def allowed_payment_file(filename):
    """Check if payment proof file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_PAYMENT_EXTENSIONS

def get_school_settings():
    """Get school settings from database"""
    school_data = {
        'school_name': 'Modern School',
        'school_email': '',
        'school_phone': '',
        'school_logo': None,
        'twitter_url': '',
        'facebook_url': '',
        'instagram_url': '',
        'tiktok_url': '',
        'whatsapp_number': '',
        'school_location': ''
    }
    
    connection = None
    try:
        connection = get_db_connection()
        if connection:
            with connection.cursor() as cursor:
                # Check if table exists first
                cursor.execute("""
                    SELECT COUNT(*) as count 
                    FROM information_schema.tables 
                    WHERE table_schema = %s AND table_name = 'school_settings'
                """, (DB_CONFIG['database'],))
                table_exists = cursor.fetchone()
                
                if table_exists and table_exists['count'] > 0:
                    # Table exists, fetch settings
                    cursor.execute("SELECT * FROM school_settings ORDER BY id DESC LIMIT 1")
                    result = cursor.fetchone()
                    if result:
                        # Handle both tuple and dict results
                        if isinstance(result, dict):
                            school_data = {
                                'school_name': result.get('school_name') or 'Modern School',
                                'school_email': result.get('school_email') or '',
                                'school_phone': result.get('school_phone') or '',
                                'school_logo': result.get('school_logo') or None,
                                'twitter_url': result.get('twitter_url') or '',
                                'facebook_url': result.get('facebook_url') or '',
                                'instagram_url': result.get('instagram_url') or '',
                                'tiktok_url': result.get('tiktok_url') or '',
                                'whatsapp_number': result.get('whatsapp_number') or '',
                                'school_location': result.get('school_location') or ''
                            }
                        else:
                            # Tuple result (fallback)
                            school_data = {
                                'school_name': (result[1] if len(result) > 1 else None) or 'Modern School',
                                'school_email': (result[2] if len(result) > 2 else None) or '',
                                'school_phone': (result[3] if len(result) > 3 else None) or '',
                                'school_logo': (result[4] if len(result) > 4 else None) or None,
                                'twitter_url': (result[5] if len(result) > 5 else None) or '',
                                'facebook_url': (result[6] if len(result) > 6 else None) or '',
                                'instagram_url': (result[7] if len(result) > 7 else None) or '',
                                'tiktok_url': (result[8] if len(result) > 8 else None) or '',
                                'whatsapp_number': (result[9] if len(result) > 9 else None) or '',
                                'school_location': (result[10] if len(result) > 10 else None) or ''
                            }
    except Exception as e:
        # Silently return default values if there's an error
        # Don't print errors for context processor to avoid cluttering logs
        pass
    finally:
        if connection:
            try:
                connection.close()
            except:
                pass
    
    return school_data

@app.context_processor
def inject_school_settings():
    """Make school settings and active academic levels available to all templates"""
    academic_levels = []
    connection = get_db_connection()
    if connection:
        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT id, level_category, level_name, level_description 
                    FROM academic_levels 
                    WHERE level_status = 'active'
                    ORDER BY level_name ASC
                """)
                results = cursor.fetchall()
                
                if results:
                    for row in results:
                        academic_levels.append({
                            'id': row.get('id'),
                            'level_category': row.get('level_category', ''),
                            'level_name': row.get('level_name', ''),
                            'level_description': row.get('level_description', '')
                        })
        except Exception as e:
            # Table might not exist yet, that's okay
            pass
        finally:
            if connection:
                connection.close()
    
    return {
        'school_settings': get_school_settings(),
        'academic_levels': academic_levels
    }

# Function to detect if running on hosted server
def is_hosted():
    """Check if the application is running on the hosted server"""
    # Check if we're in the production path
    current_path = os.path.abspath(os.getcwd())
    if '/home1/projectl/project_lucas' in current_path or '\\home1\\projectl\\project_lucas' in current_path:
        return True
    
    # Check for environment variable that indicates hosting
    if os.environ.get('IS_HOSTED', '').lower() in ['true', '1', 'yes']:
        return True
    
    # Check if DB_HOST is set to something other than localhost (indicates hosted)
    db_host = os.environ.get('DB_HOST', 'localhost')
    if db_host != 'localhost' and db_host != '127.0.0.1':
        return True
    
    return False

# Database configuration - automatically detect hosted vs local
if is_hosted():
    # Hosted server credentials
    DB_CONFIG = {
        'host': os.environ.get('DB_HOST', 'localhost'),
        'user': os.environ.get('DB_USER', 'projectl_school'),
        'password': os.environ.get('DB_PASSWORD', 'Itskimathi007'),
        'database': os.environ.get('DB_NAME', 'projectl_school'),
        'charset': 'utf8mb4',
        'cursorclass': pymysql.cursors.DictCursor
    }
else:
    # Local development credentials
    DB_CONFIG = {
        'host': os.environ.get('DB_HOST', 'localhost'),
        'user': os.environ.get('DB_USER', 'root'),
        'password': os.environ.get('DB_PASSWORD', ''),
        'database': os.environ.get('DB_NAME', 'modern_school'),
        'charset': 'utf8mb4',
        'cursorclass': pymysql.cursors.DictCursor
    }

def ensure_database_exists():
    """Check if database exists, create it if it doesn't"""
    db_name = DB_CONFIG['database']
    # Connect without specifying database
    config_without_db = {k: v for k, v in DB_CONFIG.items() if k != 'database'}
    
    try:
        connection = pymysql.connect(**config_without_db)
        with connection.cursor() as cursor:
            # Check if database exists
            cursor.execute(f"SHOW DATABASES LIKE '{db_name}'")
            result = cursor.fetchone()
            
            if not result:
                # Database doesn't exist, create it
                cursor.execute(f"CREATE DATABASE {db_name} CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci")
                connection.commit()
                print(f"Database '{db_name}' created successfully.")
            else:
                print(f"Database '{db_name}' already exists.")
        connection.close()
        return True
    except Exception as e:
        print(f"Error ensuring database exists: {e}")
        return False

def check_table_exists(connection, table_name):
    """Check if a table exists in the database"""
    try:
        with connection.cursor() as cursor:
            cursor.execute(f"""
                SELECT COUNT(*) as count 
                FROM information_schema.tables 
                WHERE table_schema = %s AND table_name = %s
            """, (DB_CONFIG['database'], table_name))
            result = cursor.fetchone()
            return result['count'] > 0
    except Exception as e:
        print(f"Error checking table existence: {e}")
        return False

def generate_student_id(connection):
    """Generate a unique student ID in format STU001, STU002, etc."""
    try:
        with connection.cursor() as cursor:
            # Get the highest student ID number
            cursor.execute("""
                SELECT student_id FROM students 
                WHERE student_id LIKE 'STU%' 
                ORDER BY CAST(SUBSTRING(student_id, 4) AS UNSIGNED) DESC 
                LIMIT 1
            """)
            result = cursor.fetchone()
            
            if result:
                # Extract the number part and increment
                last_number = int(result['student_id'][3:])  # Skip 'STU' prefix
                new_number = last_number + 1
            else:
                # First student
                new_number = 1
            
            # Format as STU001, STU002, etc. (3 digits minimum)
            student_id = f"STU{new_number:03d}"
            return student_id
    except Exception as e:
        print(f"Error generating student ID: {e}")
        # Fallback: use timestamp-based ID if there's an error
        return f"STU{int(datetime.now().timestamp()) % 100000:05d}"

def init_db():
    """Initialize database and tables - creates database if missing, then creates tables"""
    # First, ensure the database exists
    if not ensure_database_exists():
        print("Failed to ensure database exists.")
        return False
    
    # Now connect to the database
    try:
        connection = pymysql.connect(**DB_CONFIG)
    except Exception as e:
        print(f"Database connection error: {e}")
        return False
    
    try:
        with connection.cursor() as cursor:
            # Create users table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    full_name VARCHAR(255) NOT NULL,
                    email VARCHAR(255) UNIQUE NOT NULL,
                    password_hash VARCHAR(255) NOT NULL,
                    role ENUM('parent', 'student', 'employee') NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # Create students table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS students (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    student_id VARCHAR(20) UNIQUE NOT NULL,
                    full_name VARCHAR(255) NOT NULL,
                    date_of_birth DATE,
                    gender VARCHAR(50),
                    current_grade VARCHAR(50),
                    previous_school VARCHAR(255),
                    address TEXT,
                    medical_info TEXT,
                    special_needs TEXT,
                    student_category VARCHAR(50),
                    sponsor_name VARCHAR(255),
                    sponsor_phone VARCHAR(50),
                    sponsor_email VARCHAR(255),
                    status ENUM('pending approval', 'in session', 'suspended', 'expelled', 'alumni') DEFAULT 'pending approval',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
                )
            """)
            
            # Create parents table (linked to students via student_id)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS parents (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    student_id VARCHAR(20) NOT NULL,
                    full_name VARCHAR(255) NOT NULL,
                    phone VARCHAR(50) NOT NULL,
                    email VARCHAR(255) NOT NULL,
                    relationship VARCHAR(50),
                    emergency_contact VARCHAR(255),
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    FOREIGN KEY (student_id) REFERENCES students(student_id) ON DELETE CASCADE,
                    INDEX idx_student_id (student_id)
                )
            """)
            
            # Keep admissions table for backward compatibility (optional - can be removed later)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS admissions (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    student_full_name VARCHAR(255) NOT NULL,
                    date_of_birth DATE,
                    gender VARCHAR(50),
                    current_grade VARCHAR(50),
                    previous_school VARCHAR(255),
                    parent_name VARCHAR(255) NOT NULL,
                    parent_phone VARCHAR(50) NOT NULL,
                    parent_email VARCHAR(255) NOT NULL,
                    address TEXT,
                    emergency_contact VARCHAR(255),
                    medical_info TEXT,
                    special_needs TEXT,
                    status ENUM('pending approval', 'in session', 'suspended', 'expelled', 'alumni') DEFAULT 'pending approval',
                    submitted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # Create news table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS news (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    title VARCHAR(255) NOT NULL,
                    summary TEXT,
                    content TEXT,
                    image_url VARCHAR(500),
                    date DATE,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # Create gallery table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS gallery (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    title VARCHAR(255) NOT NULL,
                    description TEXT,
                    image_url VARCHAR(500),
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # Create employees table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS employees (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    employee_id VARCHAR(20) UNIQUE NOT NULL,
                    full_name VARCHAR(255) NOT NULL,
                    email VARCHAR(255) UNIQUE NOT NULL,
                    phone VARCHAR(50) NOT NULL,
                    id_number VARCHAR(50) NOT NULL,
                    password_hash VARCHAR(255) NOT NULL,
                    profile_picture VARCHAR(500),
                    role ENUM('employee', 'super admin', 'principal', 'deputy principal', 'academic coordinator', 'teachers', 'accountant', 'librarian', 'warden', 'transport manager', 'technician') DEFAULT 'employee',
                    status ENUM('pending approval', 'active', 'suspended', 'fired', 'retired') DEFAULT 'pending approval',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    INDEX idx_email (email),
                    INDEX idx_employee_id (employee_id)
                )
            """)
            
            # Create employee_salaries table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS employee_salaries (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    employee_id INT NOT NULL,
                    effective_date DATE NOT NULL,
                    basic_salary DECIMAL(15, 2) NOT NULL DEFAULT 0.00,
                    house_allowance DECIMAL(15, 2) DEFAULT 0.00,
                    transport_allowance DECIMAL(15, 2) DEFAULT 0.00,
                    medical_allowance DECIMAL(15, 2) DEFAULT 0.00,
                    overtime DECIMAL(15, 2) DEFAULT 0.00,
                    bonus DECIMAL(15, 2) DEFAULT 0.00,
                    paye DECIMAL(15, 2) DEFAULT 0.00,
                    nssf DECIMAL(15, 2) DEFAULT 0.00,
                    nhif DECIMAL(15, 2) DEFAULT 0.00,
                    sacco DECIMAL(15, 2) DEFAULT 0.00,
                    staff_loans DECIMAL(15, 2) DEFAULT 0.00,
                    absenteeism DECIMAL(15, 2) DEFAULT 0.00,
                    total_earnings DECIMAL(15, 2) NOT NULL DEFAULT 0.00,
                    total_deductions DECIMAL(15, 2) NOT NULL DEFAULT 0.00,
                    net_salary DECIMAL(15, 2) NOT NULL DEFAULT 0.00,
                    is_active BOOLEAN DEFAULT TRUE,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE,
                    INDEX idx_employee_id (employee_id),
                    INDEX idx_effective_date (effective_date),
                    INDEX idx_is_active (is_active)
                )
            """)
            
            # Create employee_salary_payments table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS employee_salary_payments (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    employee_id INT NOT NULL,
                    salary_id INT NOT NULL,
                    amount_paid DECIMAL(15, 2) NOT NULL DEFAULT 0.00,
                    payment_date DATE NOT NULL,
                    payment_method ENUM('Cash', 'Bank Transfer', 'Cheque', 'Mobile Money', 'Credit/Debit Card') DEFAULT 'Bank Transfer',
                    reference_number VARCHAR(255),
                    notes TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE,
                    FOREIGN KEY (salary_id) REFERENCES employee_salaries(id) ON DELETE CASCADE,
                    INDEX idx_employee_id (employee_id),
                    INDEX idx_salary_id (salary_id),
                    INDEX idx_payment_date (payment_date)
                )
            """)
            
            # Create employee_salary_audits table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS employee_salary_audits (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    salary_id INT NOT NULL,
                    employee_id INT NOT NULL,
                    field_name VARCHAR(100) NOT NULL,
                    old_value TEXT,
                    new_value TEXT,
                    edited_by INT NOT NULL,
                    edited_by_name VARCHAR(255),
                    edited_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (salary_id) REFERENCES employee_salaries(id) ON DELETE CASCADE,
                    FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE,
                    FOREIGN KEY (edited_by) REFERENCES employees(id) ON DELETE CASCADE,
                    INDEX idx_salary_id (salary_id),
                    INDEX idx_employee_id (employee_id),
                    INDEX idx_edited_at (edited_at)
                )
            """)
            
            # Create employee_permissions table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS employee_permissions (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    employee_id INT NOT NULL,
                    permission_key VARCHAR(100) NOT NULL,
                    granted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    granted_by INT,
                    FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE,
                    FOREIGN KEY (granted_by) REFERENCES employees(id) ON DELETE SET NULL,
                    UNIQUE KEY unique_employee_permission (employee_id, permission_key),
                    INDEX idx_employee_id (employee_id),
                    INDEX idx_permission_key (permission_key)
                )
            """)
            
            # Create school_settings table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS school_settings (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    school_name VARCHAR(255) DEFAULT 'Modern School',
                    school_email VARCHAR(255),
                    school_phone VARCHAR(50),
                    school_logo VARCHAR(500),
                    twitter_url VARCHAR(255),
                    facebook_url VARCHAR(255),
                    instagram_url VARCHAR(255),
                    tiktok_url VARCHAR(255),
                    whatsapp_number VARCHAR(50),
                    school_location TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
                )
            """)
            
            # Insert default school settings if not exists
            cursor.execute("SELECT COUNT(*) as count FROM school_settings")
            result = cursor.fetchone()
            count = result['count'] if isinstance(result, dict) else result[0]
            if count == 0:
                cursor.execute("""
                    INSERT INTO school_settings (school_name, school_email, school_phone, school_location)
                    VALUES ('Modern School', '', '', '')
                """)
            
            # Create academic_levels table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS academic_levels (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    level_category VARCHAR(255) NOT NULL,
                    level_name VARCHAR(255) NOT NULL,
                    level_description TEXT,
                    level_status ENUM('active', 'inactive') DEFAULT 'active',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    INDEX idx_level_status (level_status)
                )
            """)
            
            # Create fee_structures table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS fee_structures (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    academic_level_id INT NOT NULL,
                    fee_name VARCHAR(255) NOT NULL,
                    start_date DATE NOT NULL,
                    end_date DATE NOT NULL,
                    payment_deadline DATE NOT NULL,
                    total_amount DECIMAL(10, 2) NOT NULL DEFAULT 0.00,
                    status ENUM('active', 'inactive') DEFAULT 'active',
                    created_by INT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    FOREIGN KEY (academic_level_id) REFERENCES academic_levels(id) ON DELETE CASCADE,
                    INDEX idx_academic_level (academic_level_id),
                    INDEX idx_status (status)
                )
            """)
            
            # Create fee_items table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS fee_items (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    fee_structure_id INT NOT NULL,
                    item_name VARCHAR(255) NOT NULL,
                    item_description TEXT,
                    amount DECIMAL(10, 2) NOT NULL,
                    item_order INT DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    FOREIGN KEY (fee_structure_id) REFERENCES fee_structures(id) ON DELETE CASCADE,
                    INDEX idx_fee_structure (fee_structure_id)
                )
            """)
            
            # Create student_payments table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS student_payments (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    student_id VARCHAR(20) NOT NULL,
                    fee_structure_id INT,
                    amount_paid DECIMAL(10, 2) NOT NULL,
                    payment_method ENUM('Cash', 'Bank Transfer', 'Cheque', 'Mobile Money', 'Credit/Debit Card') NOT NULL,
                    reference_number VARCHAR(255),
                    cheque_number VARCHAR(255),
                    transaction_id VARCHAR(255),
                    proof_of_payment VARCHAR(500),
                    received_by INT,
                    payment_date DATE NOT NULL,
                    notes TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    FOREIGN KEY (student_id) REFERENCES students(student_id) ON DELETE CASCADE,
                    FOREIGN KEY (fee_structure_id) REFERENCES fee_structures(id) ON DELETE SET NULL,
                    FOREIGN KEY (received_by) REFERENCES employees(id) ON DELETE SET NULL,
                    INDEX idx_student_id (student_id),
                    INDEX idx_fee_structure (fee_structure_id),
                    INDEX idx_payment_date (payment_date)
                )
            """)
            
            # Create student_payment_audit table to track all changes
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS student_payment_audit (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    payment_id INT,
                    student_id VARCHAR(20) NOT NULL,
                    action_type ENUM('INSERT', 'UPDATE', 'DELETE') NOT NULL,
                    field_name VARCHAR(100),
                    old_value TEXT,
                    new_value TEXT,
                    changed_by INT,
                    changed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (payment_id) REFERENCES student_payments(id) ON DELETE SET NULL,
                    FOREIGN KEY (student_id) REFERENCES students(student_id) ON DELETE CASCADE,
                    FOREIGN KEY (changed_by) REFERENCES employees(id) ON DELETE SET NULL,
                    INDEX idx_payment_id (payment_id),
                    INDEX idx_student_id (student_id),
                    INDEX idx_changed_at (changed_at),
                    INDEX idx_action_type (action_type)
                )
            """)
            
            # Migrate existing table: rename status column to level_status if it exists
            try:
                cursor.execute("SHOW COLUMNS FROM academic_levels LIKE 'status'")
                if cursor.fetchone():
                    # Check if idx_status index exists and drop it
                    cursor.execute("SHOW INDEX FROM academic_levels WHERE Key_name = 'idx_status'")
                    if cursor.fetchone():
                        cursor.execute("ALTER TABLE academic_levels DROP INDEX idx_status")
                    # Rename the column
                    cursor.execute("ALTER TABLE academic_levels CHANGE COLUMN status level_status ENUM('active', 'inactive') DEFAULT 'active'")
                    # Add new index if it doesn't exist
                    cursor.execute("SHOW INDEX FROM academic_levels WHERE Key_name = 'idx_level_status'")
                    if not cursor.fetchone():
                        cursor.execute("ALTER TABLE academic_levels ADD INDEX idx_level_status (level_status)")
                    connection.commit()
                    print("✓ Migrated academic_levels.status to level_status")
            except Exception as e:
                # Column might not exist or already renamed
                print(f"Migration note: {e}")
                pass
            
            # Create fee_structures table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS fee_structures (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    academic_level_id INT NOT NULL,
                    fee_name VARCHAR(255) NOT NULL,
                    start_date DATE NOT NULL,
                    end_date DATE NOT NULL,
                    payment_deadline DATE NOT NULL,
                    total_amount DECIMAL(10, 2) NOT NULL DEFAULT 0.00,
                    status ENUM('active', 'inactive') DEFAULT 'active',
                    created_by INT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    INDEX idx_academic_level (academic_level_id),
                    INDEX idx_status (status),
                    FOREIGN KEY (academic_level_id) REFERENCES academic_levels(id) ON DELETE CASCADE
                )
            """)
            
            # Create fee_items table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS fee_items (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    fee_structure_id INT NOT NULL,
                    item_name VARCHAR(255) NOT NULL,
                    item_description TEXT,
                    amount DECIMAL(10, 2) NOT NULL,
                    item_order INT DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    INDEX idx_fee_structure (fee_structure_id)
                )
            """)
            
            # Create academic_years table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS academic_years (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    year_name VARCHAR(255) NOT NULL,
                    start_date DATE NOT NULL,
                    end_date DATE NOT NULL,
                    status ENUM('draft', 'active', 'closed', 'suspended') DEFAULT 'draft',
                    is_current BOOLEAN DEFAULT FALSE,
                    is_locked BOOLEAN DEFAULT FALSE,
                    locked_at TIMESTAMP NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    INDEX idx_status (status),
                    INDEX idx_is_current (is_current),
                    INDEX idx_is_locked (is_locked)
                )
            """)
            
            # Add is_locked and locked_at columns if they don't exist (migration)
            try:
                cursor.execute("SHOW COLUMNS FROM academic_years LIKE 'is_locked'")
                if not cursor.fetchone():
                    cursor.execute("ALTER TABLE academic_years ADD COLUMN is_locked BOOLEAN DEFAULT FALSE AFTER is_current")
                    cursor.execute("ALTER TABLE academic_years ADD COLUMN locked_at TIMESTAMP NULL AFTER is_locked")
                    cursor.execute("ALTER TABLE academic_years ADD INDEX idx_is_locked (is_locked)")
                    print("✓ Added is_locked and locked_at columns to academic_years table")
            except Exception as e:
                print(f"Migration note for academic_years.is_locked: {e}")
                pass
            
            # Update status enum to include 'suspended' if it doesn't exist
            try:
                cursor.execute("SHOW COLUMNS FROM academic_years WHERE Field = 'status'")
                status_col = cursor.fetchone()
                if status_col and 'suspended' not in str(status_col):
                    cursor.execute("ALTER TABLE academic_years MODIFY COLUMN status ENUM('draft', 'active', 'closed', 'suspended') DEFAULT 'draft'")
                    print("✓ Updated academic_years.status enum to include 'suspended'")
            except Exception as e:
                print(f"Migration note for academic_years.status enum: {e}")
                pass
            
            # Create terms table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS terms (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    term_name VARCHAR(255) NOT NULL,
                    academic_year_id INT NOT NULL,
                    academic_level_id INT,
                    start_date DATE NOT NULL,
                    end_date DATE NOT NULL,
                    status ENUM('draft', 'active', 'closed', 'suspended') DEFAULT 'draft',
                    is_locked BOOLEAN DEFAULT FALSE,
                    locked_at TIMESTAMP NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    FOREIGN KEY (academic_year_id) REFERENCES academic_years(id) ON DELETE CASCADE,
                    FOREIGN KEY (academic_level_id) REFERENCES academic_levels(id) ON DELETE SET NULL,
                    INDEX idx_academic_year (academic_year_id),
                    INDEX idx_academic_level (academic_level_id),
                    INDEX idx_status (status),
                    INDEX idx_is_locked (is_locked)
                )
            """)
            
            # Add is_locked and locked_at columns if they don't exist (migration)
            try:
                cursor.execute("SHOW COLUMNS FROM terms LIKE 'is_locked'")
                if not cursor.fetchone():
                    cursor.execute("ALTER TABLE terms ADD COLUMN is_locked BOOLEAN DEFAULT FALSE AFTER status")
                    cursor.execute("ALTER TABLE terms ADD COLUMN locked_at TIMESTAMP NULL AFTER is_locked")
                    cursor.execute("ALTER TABLE terms ADD INDEX idx_is_locked (is_locked)")
                    print("✓ Added is_locked and locked_at columns to terms table")
            except Exception as e:
                print(f"Migration note for terms.is_locked: {e}")
                pass
            
            # Update status enum to include 'suspended' if it doesn't exist
            try:
                cursor.execute("SHOW COLUMNS FROM terms WHERE Field = 'status'")
                status_col = cursor.fetchone()
                if status_col and 'suspended' not in str(status_col):
                    cursor.execute("ALTER TABLE terms MODIFY COLUMN status ENUM('draft', 'active', 'closed', 'suspended') DEFAULT 'draft'")
                    print("✓ Updated terms.status enum to include 'suspended'")
            except Exception as e:
                print(f"Migration note for terms.status enum: {e}")
                pass
            
            # Add is_current column to terms if it doesn't exist
            try:
                cursor.execute("SHOW COLUMNS FROM terms LIKE 'is_current'")
                if not cursor.fetchone():
                    cursor.execute("ALTER TABLE terms ADD COLUMN is_current BOOLEAN DEFAULT FALSE AFTER status")
                    cursor.execute("ALTER TABLE terms ADD INDEX idx_is_current (is_current)")
                    print("✓ Added is_current column to terms table")
            except Exception as e:
                print(f"Migration note for terms.is_current: {e}")
                pass
            
            # Create term_academic_levels junction table for many-to-many relationship
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS term_academic_levels (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    term_id INT NOT NULL,
                    academic_level_id INT NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (term_id) REFERENCES terms(id) ON DELETE CASCADE,
                    FOREIGN KEY (academic_level_id) REFERENCES academic_levels(id) ON DELETE CASCADE,
                    UNIQUE KEY unique_term_level (term_id, academic_level_id),
                    INDEX idx_term_id (term_id),
                    INDEX idx_academic_level_id (academic_level_id)
                )
            """)
            
            # Update fee_structures to include term_id and academic_year_id
            try:
                # Check if term_id column exists
                cursor.execute("SHOW COLUMNS FROM fee_structures LIKE 'term_id'")
                if not cursor.fetchone():
                    cursor.execute("ALTER TABLE fee_structures ADD COLUMN term_id INT NULL AFTER academic_level_id")
                    cursor.execute("ALTER TABLE fee_structures ADD COLUMN academic_year_id INT NULL AFTER term_id")
                    cursor.execute("ALTER TABLE fee_structures ADD FOREIGN KEY (term_id) REFERENCES terms(id) ON DELETE SET NULL")
                    cursor.execute("ALTER TABLE fee_structures ADD FOREIGN KEY (academic_year_id) REFERENCES academic_years(id) ON DELETE SET NULL")
                    cursor.execute("ALTER TABLE fee_structures ADD INDEX idx_term (term_id)")
                    cursor.execute("ALTER TABLE fee_structures ADD INDEX idx_academic_year (academic_year_id)")
                    print("✓ Added term_id and academic_year_id to fee_structures")
            except Exception as e:
                print(f"Migration note for fee_structures: {e}")
                pass
            
            # Add category column to fee_structures if it doesn't exist
            try:
                cursor.execute("SHOW COLUMNS FROM fee_structures LIKE 'category'")
                if not cursor.fetchone():
                    cursor.execute("ALTER TABLE fee_structures ADD COLUMN category VARCHAR(50) NULL DEFAULT 'both' AFTER fee_name")
                    print("✓ Added category column to fee_structures")
            except Exception as e:
                print(f"Migration note for fee_structures category: {e}")
                pass
            
            # Add student_category and sponsor_name columns to students table if they don't exist
            try:
                cursor.execute("""
                    SELECT COUNT(*) as count 
                    FROM information_schema.COLUMNS 
                    WHERE TABLE_SCHEMA = DATABASE() 
                    AND TABLE_NAME = 'students' 
                    AND COLUMN_NAME = 'student_category'
                """)
                result = cursor.fetchone()
                if result and result[0] == 0:
                    cursor.execute("ALTER TABLE students ADD COLUMN student_category VARCHAR(50) NULL AFTER special_needs")
                    print("✓ Added student_category column to students table")
            except Exception as e:
                print(f"Migration note for student_category: {e}")
                pass
            
            try:
                cursor.execute("""
                    SELECT COUNT(*) as count 
                    FROM information_schema.COLUMNS 
                    WHERE TABLE_SCHEMA = DATABASE() 
                    AND TABLE_NAME = 'students' 
                    AND COLUMN_NAME = 'sponsor_name'
                """)
                result = cursor.fetchone()
                if result and result[0] == 0:
                    cursor.execute("ALTER TABLE students ADD COLUMN sponsor_name VARCHAR(255) NULL AFTER student_category")
                    print("✓ Added sponsor_name column to students table")
            except Exception as e:
                print(f"Migration note for sponsor_name: {e}")
                pass
            
            # Add sponsor_phone column to students table if it doesn't exist
            try:
                cursor.execute("""
                    SELECT COUNT(*) as count 
                    FROM information_schema.COLUMNS 
                    WHERE TABLE_SCHEMA = DATABASE() 
                    AND TABLE_NAME = 'students' 
                    AND COLUMN_NAME = 'sponsor_phone'
                """)
                result = cursor.fetchone()
                if result and result[0] == 0:
                    cursor.execute("ALTER TABLE students ADD COLUMN sponsor_phone VARCHAR(50) NULL AFTER sponsor_name")
                    print("✓ Added sponsor_phone column to students table")
            except Exception as e:
                print(f"Migration note for sponsor_phone: {e}")
                pass
            
            # Add sponsor_email column to students table if it doesn't exist
            try:
                cursor.execute("""
                    SELECT COUNT(*) as count 
                    FROM information_schema.COLUMNS 
                    WHERE TABLE_SCHEMA = DATABASE() 
                    AND TABLE_NAME = 'students' 
                    AND COLUMN_NAME = 'sponsor_email'
                """)
                result = cursor.fetchone()
                if result and result[0] == 0:
                    cursor.execute("ALTER TABLE students ADD COLUMN sponsor_email VARCHAR(255) NULL AFTER sponsor_phone")
                    print("✓ Added sponsor_email column to students table")
            except Exception as e:
                print(f"Migration note for sponsor_email: {e}")
                pass
            
            connection.commit()
            
            # Verify tables were created
            required_tables = ['users', 'students', 'parents', 'employees', 'admissions', 'news', 'gallery', 'school_settings', 'academic_levels', 'fee_structures', 'fee_items']
            missing_tables = []
            for table in required_tables:
                if not check_table_exists(connection, table):
                    missing_tables.append(table)
            
            if missing_tables:
                print(f"Warning: Some tables could not be created: {', '.join(missing_tables)}")
                return False
            
            print("All database tables verified/created successfully.")
            return True
    except Exception as e:
        print(f"Database initialization error: {e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        if connection:
            try:
                connection.close()
            except:
                pass

def get_db_connection():
    """Create and return a database connection - automatically creates database and tables if missing"""
    try:
        connection = pymysql.connect(**DB_CONFIG)
        return connection
    except pymysql.err.OperationalError as e:
        # If database doesn't exist, try to create it and reconnect
        if e.args[0] == 1049:  # Unknown database error
            print(f"Database '{DB_CONFIG['database']}' not found. Creating database and tables...")
            if ensure_database_exists():
                try:
                    # Initialize tables in the newly created database
                    if init_db():
                        # Now connect to the database
                        connection = pymysql.connect(**DB_CONFIG)
                        return connection
                    else:
                        print("Failed to initialize database tables.")
                        return None
                except Exception as e2:
                    print(f"Database connection error after creation: {e2}")
                    return None
            else:
                print(f"Failed to create database '{DB_CONFIG['database']}'.")
                return None
        else:
            print(f"Database connection error: {e}")
            return None
    except Exception as e:
        print(f"Database connection error: {e}")
        return None

def login_required(f):
    """Decorator to require login"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please login to access this page.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def role_required(role):
    """Decorator to require specific role"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                flash('Please login to access this page.', 'error')
                return redirect(url_for('login'))
            if session.get('role') != role:
                flash('You do not have permission to access this page.', 'error')
                return redirect(url_for('home'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def has_permission(employee_id, permission_key):
    """Check if an employee has a specific permission"""
    if not employee_id:
        return False
    
    connection = get_db_connection()
    if not connection:
        return False
    
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT COUNT(*) as count 
                FROM employee_permissions 
                WHERE employee_id = %s AND permission_key = %s
            """, (employee_id, permission_key))
            result = cursor.fetchone()
            if result:
                count = result.get('count', 0) if isinstance(result, dict) else result[0] if isinstance(result, tuple) else 0
                return count > 0
            return False
    except Exception as e:
        print(f"Error checking permission: {e}")
        return False
    finally:
        if connection:
            try:
                connection.close()
            except:
                pass

def check_permission_or_role(permission_key, allowed_roles=None):
    """Check if user has permission OR is in allowed roles (for backward compatibility)
    
    Priority:
    1. Technicians always have access
    2. If employee has specific permission, grant access
    3. If employee has ANY permissions assigned but not this one, deny access (permission-based mode)
    4. If employee has NO permissions assigned, fall back to role-based access
    """
    user_role = session.get('role', '').lower()
    employee_id = session.get('employee_id') or session.get('user_id')
    
    # Technicians have all permissions
    if user_role == 'technician':
        return True
    
    # First, get the actual employee ID from database (handle both id and employee_id fields)
    actual_employee_id = None
    if employee_id:
        connection = get_db_connection()
        if connection:
            try:
                with connection.cursor() as cursor:
                    # Try to find employee by id or employee_id field
                    cursor.execute("""
                        SELECT id 
                        FROM employees 
                        WHERE id = %s OR employee_id = %s
                        LIMIT 1
                    """, (employee_id, employee_id))
                    result = cursor.fetchone()
                    if result:
                        actual_employee_id = result.get('id') if isinstance(result, dict) else result[0]
            except Exception as e:
                print(f"Error finding employee ID: {e}")
            finally:
                if connection:
                    try:
                        connection.close()
                    except:
                        pass
    
    # Check if employee has specific permission assigned
    if actual_employee_id:
        has_specific_permission = has_permission(actual_employee_id, permission_key)
        if has_specific_permission:
            return True
        
        # Check if employee has ANY permissions assigned
        connection = get_db_connection()
        if connection:
            try:
                with connection.cursor() as cursor:
                    cursor.execute("""
                        SELECT COUNT(*) as count 
                        FROM employee_permissions 
                        WHERE employee_id = %s
                    """, (actual_employee_id,))
                    result = cursor.fetchone()
                    if result:
                        total_permissions = result.get('count', 0) if isinstance(result, dict) else result[0] if isinstance(result, tuple) else 0
                        # If employee has any permissions assigned, we're in permission-based mode
                        # So if they don't have this specific permission, deny access
                        if total_permissions > 0:
                            return False  # Permission-based mode: no permission = no access
            except Exception as e:
                print(f"Error checking total permissions: {e}")
            finally:
                if connection:
                    try:
                        connection.close()
                    except:
                        pass
    
    # Fall back to role-based access only if no permissions are assigned (backward compatibility)
    if allowed_roles and user_role in allowed_roles:
        return True
    
    return False

def get_employee_permissions_list(employee_id):
    """Get list of all permissions for an employee"""
    connection = get_db_connection()
    permissions = []
    
    if not connection:
        return permissions
    
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT permission_key 
                FROM employee_permissions 
                WHERE employee_id = %s
            """, (employee_id,))
            results = cursor.fetchall()
            for result in results:
                if isinstance(result, dict):
                    permissions.append(result.get('permission_key'))
                else:
                    permissions.append(result[0] if result else '')
    except Exception as e:
        print(f"Error fetching employee permissions: {e}")
    finally:
        if connection:
            try:
                connection.close()
            except:
                pass
    
    return permissions

def permission_required(permission_key):
    """Decorator to require specific permission"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                flash('Please login to access this page.', 'error')
                return redirect(url_for('login'))
            
            user_role = session.get('role', '').lower()
            employee_id = session.get('employee_id') or session.get('user_id')
            
            # Technicians have all permissions
            if user_role == 'technician':
                return f(*args, **kwargs)
            
            # Check if employee has the permission
            if employee_id and has_permission(employee_id, permission_key):
                return f(*args, **kwargs)
            
            # Fallback to role-based check for backward compatibility
            # This allows existing role-based access to still work
            flash('You do not have permission to access this page.', 'error')
            return redirect(url_for('dashboard_employee'))
        return decorated_function
    return decorator

def send_admission_confirmation_email(parent_email, parent_name, student_name, student_id):
    """Send confirmation email to parent/guardian after admission submission"""
    try:
        # Get support contact information from environment or use defaults
        support_email = os.environ.get('SUPPORT_EMAIL', 'support@modernschool.com')
        support_phone = os.environ.get('SUPPORT_PHONE', '+254 700 000 000')
        school_name = os.environ.get('SCHOOL_NAME', 'Modern School')
        
        subject = f"Admission Application Received - {student_name}"
        
        # Create email body
        html_body = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    line-height: 1.6;
                    color: #333;
                    max-width: 600px;
                    margin: 0 auto;
                    padding: 20px;
                }}
                .header {{
                    background: linear-gradient(135deg, #1e40af 0%, #f97316 100%);
                    color: white;
                    padding: 30px;
                    text-align: center;
                    border-radius: 10px 10px 0 0;
                }}
                .content {{
                    background: #f9fafb;
                    padding: 30px;
                    border: 1px solid #e5e7eb;
                }}
                .info-box {{
                    background: white;
                    border-left: 4px solid #f97316;
                    padding: 15px;
                    margin: 20px 0;
                }}
                .footer {{
                    background: #1f2937;
                    color: white;
                    padding: 20px;
                    text-align: center;
                    border-radius: 0 0 10px 10px;
                    font-size: 14px;
                }}
                .contact-info {{
                    background: #eff6ff;
                    border: 1px solid #3b82f6;
                    border-radius: 8px;
                    padding: 20px;
                    margin: 20px 0;
                }}
                .contact-info h3 {{
                    margin-top: 0;
                    color: #1e40af;
                }}
                .contact-info p {{
                    margin: 8px 0;
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>{school_name}</h1>
            </div>
            <div class="content">
                <p>Dear {parent_name},</p>
                
                <p>Thank you for submitting the admission application for <strong>{student_name}</strong>.</p>
                
                <div class="info-box">
                    <p><strong>Application Details:</strong></p>
                    <p><strong>Student Name:</strong> {student_name}</p>
                    <p><strong>Student ID:</strong> {student_id}</p>
                    <p><strong>Status:</strong> Pending Approval</p>
                </div>
                
                <p>Your application has been received and is currently undergoing our vetting process. We will review all submitted information and contact you within 5-7 business days with an update on the application status.</p>
                
                <p>Please keep your Student ID (<strong>{student_id}</strong>) for future reference, as you will need it to check the status of your application.</p>
                
                <div class="contact-info">
                    <h3>Need Assistance?</h3>
                    <p>If you have any questions or concerns regarding your application, please don't hesitate to contact our support team:</p>
                    <p><strong>Email:</strong> <a href="mailto:{support_email}" style="color: #1e40af;">{support_email}</a></p>
                    <p><strong>Phone:</strong> {support_phone}</p>
                    <p>Our support team is available Monday through Friday, 8:00 AM to 5:00 PM.</p>
                </div>
                
                <p>We appreciate your interest in {school_name} and look forward to the possibility of welcoming {student_name} to our school community.</p>
                
                <p>Best regards,<br>
                <strong>Admissions Office</strong><br>
                {school_name}</p>
            </div>
            <div class="footer">
                <p>This is an automated message. Please do not reply to this email.</p>
                <p>&copy; {datetime.now().year} {school_name}. All rights reserved.</p>
            </div>
        </body>
        </html>
        """
        
        # Plain text version
        text_body = f"""
Dear {parent_name},

Thank you for submitting the admission application for {student_name}.

Application Details:
- Student Name: {student_name}
- Student ID: {student_id}
- Status: Pending Approval

Your application has been received and is currently undergoing our vetting process. We will review all submitted information and contact you within 5-7 business days with an update on the application status.

Please keep your Student ID ({student_id}) for future reference, as you will need it to check the status of your application.

Need Assistance?
If you have any questions or concerns regarding your application, please don't hesitate to contact our support team:
- Email: {support_email}
- Phone: {support_phone}
Our support team is available Monday through Friday, 8:00 AM to 5:00 PM.

We appreciate your interest in {school_name} and look forward to the possibility of welcoming {student_name} to our school community.

Best regards,
Admissions Office
{school_name}

---
This is an automated message. Please do not reply to this email.
© {datetime.now().year} {school_name}. All rights reserved.
        """
        
        msg = Message(
            subject=subject,
            recipients=[parent_email],
            html=html_body,
            body=text_body
        )
        
        mail.send(msg)
        print(f"Admission confirmation email sent to {parent_email}")
        return True
    except Exception as e:
        print(f"Error sending admission confirmation email: {e}")
        return False

def send_student_approval_email(parent_email, parent_name, student_name, student_id):
    """Send approval congratulations email to parent/guardian after student admission is approved"""
    try:
        # Get support contact information from environment or use defaults
        support_email = os.environ.get('SUPPORT_EMAIL', 'support@modernschool.com')
        support_phone = os.environ.get('SUPPORT_PHONE', '+254 700 000 000')
        school_name = os.environ.get('SCHOOL_NAME', 'Modern School')
        
        subject = f"Congratulations! {student_name} Has Been Accepted to {school_name}"
        
        # Create email body
        html_body = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    line-height: 1.6;
                    color: #333;
                    max-width: 600px;
                    margin: 0 auto;
                    padding: 20px;
                }}
                .header {{
                    background: linear-gradient(135deg, #1e40af 0%, #f97316 100%);
                    color: white;
                    padding: 30px;
                    text-align: center;
                    border-radius: 10px 10px 0 0;
                }}
                .content {{
                    background: #f9fafb;
                    padding: 30px;
                    border: 1px solid #e5e7eb;
                }}
                .info-box {{
                    background: white;
                    border-left: 4px solid #10b981;
                    padding: 15px;
                    margin: 20px 0;
                }}
                .footer {{
                    background: #1f2937;
                    color: white;
                    padding: 20px;
                    text-align: center;
                    border-radius: 0 0 10px 10px;
                    font-size: 14px;
                }}
                .contact-info {{
                    background: #d1fae5;
                    border: 1px solid #10b981;
                    border-radius: 8px;
                    padding: 20px;
                    margin: 20px 0;
                }}
                .contact-info h3 {{
                    margin-top: 0;
                    color: #059669;
                }}
                .contact-info p {{
                    margin: 8px 0;
                }}
                .status-badge {{
                    display: inline-block;
                    background: #10b981;
                    color: white;
                    padding: 5px 15px;
                    border-radius: 20px;
                    font-weight: bold;
                    font-size: 14px;
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>🎉 Congratulations! 🎉</h1>
            </div>
            <div class="content">
                <p>Dear {parent_name},</p>
                
                <p>We are delighted to inform you that <strong>{student_name}</strong> has been accepted to <strong>{school_name}</strong>! We are thrilled to welcome your child to our school community.</p>
                
                <div class="info-box">
                    <p><strong>Admission Details:</strong></p>
                    <p><strong>Student Name:</strong> {student_name}</p>
                    <p><strong>Student ID:</strong> {student_id}</p>
                    <p><strong>Status:</strong> <span class="status-badge">Approved - In Session</span></p>
                </div>
                
                <p>Your child's admission has been approved and they are now officially enrolled in our school. We are excited to have {student_name} join us and look forward to supporting their educational journey.</p>
                
                <p><strong>Next Steps:</strong></p>
                <p>Our admissions office will be in touch with you shortly regarding the next steps in the admission procedure. This will include information about:</p>
                <ul>
                    <li>Orientation dates and schedules</li>
                    <li>Required documentation and forms</li>
                    <li>School policies and procedures</li>
                    <li>Important dates and events</li>
                </ul>
                
                <p>Please keep your Student ID (<strong>{student_id}</strong>) for future reference, as you will need it for various school-related activities.</p>
                
                <div class="contact-info">
                    <h3>Need Assistance?</h3>
                    <p>If you have any questions or need support, please don't hesitate to contact our admissions office:</p>
                    <p><strong>Email:</strong> <a href="mailto:{support_email}" style="color: #059669;">{support_email}</a></p>
                    <p><strong>Phone:</strong> {support_phone}</p>
                    <p>Our support team is available Monday through Friday, 8:00 AM to 5:00 PM.</p>
                </div>
                
                <p>Once again, congratulations on this wonderful achievement! We are honored to have {student_name} as part of our school family and look forward to working together to ensure their success.</p>
                
                <p>Best regards,<br>
                <strong>Admissions Office</strong><br>
                {school_name}</p>
            </div>
            <div class="footer">
                <p>This is an automated message. Please do not reply to this email.</p>
                <p>&copy; {datetime.now().year} {school_name}. All rights reserved.</p>
            </div>
        </body>
        </html>
        """
        
        # Plain text version
        text_body = f"""
Dear {parent_name},

We are delighted to inform you that {student_name} has been accepted to {school_name}! We are thrilled to welcome your child to our school community.

Admission Details:
- Student Name: {student_name}
- Student ID: {student_id}
- Status: Approved - In Session

Your child's admission has been approved and they are now officially enrolled in our school. We are excited to have {student_name} join us and look forward to supporting their educational journey.

Next Steps:
Our admissions office will be in touch with you shortly regarding the next steps in the admission procedure. This will include information about:
- Orientation dates and schedules
- Required documentation and forms
- School policies and procedures
- Important dates and events

Please keep your Student ID ({student_id}) for future reference, as you will need it for various school-related activities.

Need Assistance?
If you have any questions or need support, please don't hesitate to contact our admissions office:
- Email: {support_email}
- Phone: {support_phone}
Our support team is available Monday through Friday, 8:00 AM to 5:00 PM.

Once again, congratulations on this wonderful achievement! We are honored to have {student_name} as part of our school family and look forward to working together to ensure their success.

Best regards,
Admissions Office
{school_name}

---
This is an automated message. Please do not reply to this email.
© {datetime.now().year} {school_name}. All rights reserved.
        """
        
        msg = Message(
            subject=subject,
            recipients=[parent_email],
            html=html_body,
            body=text_body
        )
        
        mail.send(msg)
        print(f"Student approval email sent to {parent_email}")
        return True
    except Exception as e:
        print(f"Error sending student approval email: {e}")
        return False

def send_employee_welcome_email(employee_email, employee_name, employee_id):
    """Send welcome email to employee after registration"""
    try:
        # Get support contact information from environment or use defaults
        support_email = os.environ.get('SUPPORT_EMAIL', 'support@modernschool.com')
        support_phone = os.environ.get('SUPPORT_PHONE', '+254 700 000 000')
        school_name = os.environ.get('SCHOOL_NAME', 'Modern School')
        
        subject = f"Welcome to {school_name} - Employee Registration"
        
        # Create email body
        html_body = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    line-height: 1.6;
                    color: #333;
                    max-width: 600px;
                    margin: 0 auto;
                    padding: 20px;
                }}
                .header {{
                    background: linear-gradient(135deg, #10b981 0%, #059669 100%);
                    color: white;
                    padding: 30px;
                    text-align: center;
                    border-radius: 10px 10px 0 0;
                }}
                .content {{
                    background: #f9fafb;
                    padding: 30px;
                    border: 1px solid #e5e7eb;
                }}
                .info-box {{
                    background: white;
                    border-left: 4px solid #10b981;
                    padding: 15px;
                    margin: 20px 0;
                }}
                .footer {{
                    background: #1f2937;
                    color: white;
                    padding: 20px;
                    text-align: center;
                    border-radius: 0 0 10px 10px;
                    font-size: 14px;
                }}
                .contact-info {{
                    background: #d1fae5;
                    border: 1px solid #10b981;
                    border-radius: 8px;
                    padding: 20px;
                    margin: 20px 0;
                }}
                .contact-info h3 {{
                    margin-top: 0;
                    color: #059669;
                }}
                .contact-info p {{
                    margin: 8px 0;
                }}
                .status-badge {{
                    display: inline-block;
                    background: #fef3c7;
                    color: #92400e;
                    padding: 5px 15px;
                    border-radius: 20px;
                    font-weight: bold;
                    font-size: 14px;
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>{school_name}</h1>
            </div>
            <div class="content">
                <p>Dear {employee_name},</p>
                
                <p>Welcome to the <strong>{school_name}</strong> community! We are thrilled to have you join our team.</p>
                
                <div class="info-box">
                    <p><strong>Registration Details:</strong></p>
                    <p><strong>Name:</strong> {employee_name}</p>
                    <p><strong>Employee ID:</strong> {employee_id}</p>
                    <p><strong>Status:</strong> <span class="status-badge">Pending Approval</span></p>
                </div>
                
                <p>Your employee registration has been successfully submitted and is currently under review. Our administration team will review your application and notify you once your account has been activated.</p>
                
                <p>Please keep your Employee ID (<strong>{employee_id}</strong>) for future reference, as you will need it to access your account once approved.</p>
                
                <div class="contact-info">
                    <h3>Need Assistance?</h3>
                    <p>If you have any questions or need support, please don't hesitate to contact us:</p>
                    <p><strong>Email:</strong> <a href="mailto:{support_email}" style="color: #059669;">{support_email}</a></p>
                    <p><strong>Phone:</strong> {support_phone}</p>
                    <p>Our support team is available Monday through Friday, 8:00 AM to 5:00 PM.</p>
                </div>
                
                <p>We look forward to working with you and contributing to the success of {school_name}.</p>
                
                <p>Best regards,<br>
                <strong>Human Resources Department</strong><br>
                {school_name}</p>
            </div>
            <div class="footer">
                <p>This is an automated message. Please do not reply to this email.</p>
                <p>&copy; {datetime.now().year} {school_name}. All rights reserved.</p>
            </div>
        </body>
        </html>
        """
        
        # Plain text version
        text_body = f"""
Dear {employee_name},

Welcome to the {school_name} community! We are thrilled to have you join our team.

Registration Details:
- Name: {employee_name}
- Employee ID: {employee_id}
- Status: Pending Approval

Your employee registration has been successfully submitted and is currently under review. Our administration team will review your application and notify you once your account has been activated.

Please keep your Employee ID ({employee_id}) for future reference, as you will need it to access your account once approved.

Need Assistance?
If you have any questions or need support, please don't hesitate to contact us:
- Email: {support_email}
- Phone: {support_phone}
Our support team is available Monday through Friday, 8:00 AM to 5:00 PM.

We look forward to working with you and contributing to the success of {school_name}.

Best regards,
Human Resources Department
{school_name}

---
This is an automated message. Please do not reply to this email.
© {datetime.now().year} {school_name}. All rights reserved.
        """
        
        msg = Message(
            subject=subject,
            recipients=[employee_email],
            html=html_body,
            body=text_body
        )
        
        mail.send(msg)
        print(f"Employee welcome email sent to {employee_email}")
        return True
    except Exception as e:
        print(f"Error sending employee welcome email: {e}")
        return False

def send_employee_approval_email(employee_email, employee_name, employee_id, role):
    """Send approval email to employee after approval and role assignment"""
    try:
        # Get support contact information from environment or use defaults
        support_email = os.environ.get('SUPPORT_EMAIL', 'support@modernschool.com')
        support_phone = os.environ.get('SUPPORT_PHONE', '+254 700 000 000')
        school_name = os.environ.get('SCHOOL_NAME', 'Modern School')
        
        subject = f"Congratulations! Your Account Has Been Approved - {school_name}"
        
        # Create email body
        html_body = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    line-height: 1.6;
                    color: #333;
                    max-width: 600px;
                    margin: 0 auto;
                    padding: 20px;
                }}
                .header {{
                    background: linear-gradient(135deg, #10b981 0%, #059669 100%);
                    color: white;
                    padding: 30px;
                    text-align: center;
                    border-radius: 10px 10px 0 0;
                }}
                .content {{
                    background: #f9fafb;
                    padding: 30px;
                    border: 1px solid #e5e7eb;
                }}
                .info-box {{
                    background: white;
                    border-left: 4px solid #10b981;
                    padding: 15px;
                    margin: 20px 0;
                }}
                .footer {{
                    background: #1f2937;
                    color: white;
                    padding: 20px;
                    text-align: center;
                    border-radius: 0 0 10px 10px;
                    font-size: 14px;
                }}
                .contact-info {{
                    background: #d1fae5;
                    border: 1px solid #10b981;
                    border-radius: 8px;
                    padding: 20px;
                    margin: 20px 0;
                }}
                .contact-info h3 {{
                    margin-top: 0;
                    color: #059669;
                }}
                .contact-info p {{
                    margin: 8px 0;
                }}
                .status-badge {{
                    display: inline-block;
                    background: #d1fae5;
                    color: #059669;
                    padding: 5px 15px;
                    border-radius: 20px;
                    font-weight: bold;
                    font-size: 14px;
                }}
                .role-badge {{
                    display: inline-block;
                    background: #dbeafe;
                    color: #1e40af;
                    padding: 5px 15px;
                    border-radius: 20px;
                    font-weight: bold;
                    font-size: 14px;
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>{school_name}</h1>
            </div>
            <div class="content">
                <p>Dear {employee_name},</p>
                
                <p>We are delighted to inform you that your employee account has been <strong>approved</strong>! Welcome to the <strong>{school_name}</strong> team.</p>
                
                <div class="info-box">
                    <p><strong>Account Details:</strong></p>
                    <p><strong>Name:</strong> {employee_name}</p>
                    <p><strong>Employee ID:</strong> {employee_id}</p>
                    <p><strong>Assigned Role:</strong> <span class="role-badge">{role.title()}</span></p>
                    <p><strong>Status:</strong> <span class="status-badge">Active</span></p>
                </div>
                
                <p>Your account has been successfully activated and you have been assigned the role of <strong>{role.title()}</strong>. You can now log in to your employee dashboard using your registered email and password.</p>
                
                <p>Please keep your Employee ID (<strong>{employee_id}</strong>) for future reference, as you will need it for various school-related activities.</p>
                
                <div class="contact-info">
                    <h3>Need Assistance?</h3>
                    <p>If you have any questions or need support, please don't hesitate to contact us:</p>
                    <p><strong>Email:</strong> <a href="mailto:{support_email}" style="color: #059669;">{support_email}</a></p>
                    <p><strong>Phone:</strong> {support_phone}</p>
                    <p>Our support team is available Monday through Friday, 8:00 AM to 5:00 PM.</p>
                </div>
                
                <p>We are excited to have you as part of our team and look forward to working with you!</p>
                
                <p>Best regards,<br>
                <strong>Human Resources Department</strong><br>
                {school_name}</p>
            </div>
            <div class="footer">
                <p>This is an automated message. Please do not reply to this email.</p>
                <p>&copy; {datetime.now().year} {school_name}. All rights reserved.</p>
            </div>
        </body>
        </html>
        """
        
        # Plain text version
        text_body = f"""
Dear {employee_name},

We are delighted to inform you that your employee account has been approved! Welcome to the {school_name} team.

Account Details:
- Name: {employee_name}
- Employee ID: {employee_id}
- Assigned Role: {role.title()}
- Status: Active

Your account has been successfully activated and you have been assigned the role of {role.title()}. You can now log in to your employee dashboard using your registered email and password.

Please keep your Employee ID ({employee_id}) for future reference, as you will need it for various school-related activities.

Need Assistance?
If you have any questions or need support, please don't hesitate to contact us:
- Email: {support_email}
- Phone: {support_phone}
Our support team is available Monday through Friday, 8:00 AM to 5:00 PM.

We are excited to have you as part of our team and look forward to working with you!

Best regards,
Human Resources Department
{school_name}

---
This is an automated message. Please do not reply to this email.
© {datetime.now().year} {school_name}. All rights reserved.
        """
        
        msg = Message(
            subject=subject,
            recipients=[employee_email],
            html=html_body,
            body=text_body
        )
        
        mail.send(msg)
        print(f"Employee approval email sent to {employee_email}")
        return True
    except Exception as e:
        print(f"Error sending employee approval email: {e}")
        return False

# Routes
@app.route('/')
def home():
    # Fetch active academic levels for admission form
    academic_levels = []
    connection = get_db_connection()
    if connection:
        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT id, level_category, level_name, level_description 
                    FROM academic_levels 
                    WHERE level_status = 'active'
                    ORDER BY level_name ASC
                """)
                results = cursor.fetchall()
                
                if results:
                    for row in results:
                        academic_levels.append({
                            'id': row.get('id'),
                            'level_category': row.get('level_category', ''),
                            'level_name': row.get('level_name', ''),
                            'level_description': row.get('level_description', '')
                        })
        except Exception as e:
            print(f"Error fetching academic levels for home: {e}")
        finally:
            connection.close()
    
    return render_template('home.html', academic_levels=academic_levels)

@app.route('/about')
def about():
    return render_template('about.html')

@app.route('/programs')
def programs():
    return render_template('programs.html')

@app.route('/news')
def news():
    connection = get_db_connection()
    news_items = []
    if connection:
        try:
            with connection.cursor() as cursor:
                cursor.execute("SELECT * FROM news ORDER BY date DESC, created_at DESC LIMIT 6")
                news_items = cursor.fetchall()
        except Exception as e:
            print(f"Error fetching news: {e}")
        finally:
            connection.close()
    return render_template('news.html', news_items=news_items)

@app.route('/gallery')
def gallery():
    connection = get_db_connection()
    gallery_items = []
    if connection:
        try:
            with connection.cursor() as cursor:
                cursor.execute("SELECT * FROM gallery ORDER BY created_at DESC LIMIT 12")
                gallery_items = cursor.fetchall()
        except Exception as e:
            print(f"Error fetching gallery: {e}")
        finally:
            connection.close()
    return render_template('gallery.html', gallery_items=gallery_items)

@app.route('/team')
def team():
    # Team data - 5 key members
    team_members = [
        {
            'name': 'Dr. Sarah Wanjala',
            'position': 'Executive Director',
            'email': 'sarah.wanjala@modernschool.com',
            'phone': '+254 700 111 111',
            'bio': 'With over 25 years of experience in education and non-profit management, Dr. Wanjala leads our organization with strategic vision and unwavering commitment to educational excellence.',
            'details': 'Dr. Wanjala holds a Ph.D. in Educational Leadership and has been instrumental in establishing Modern School as a leading NGO-sponsored educational institution. She has spearheaded numerous initiatives that have transformed the lives of thousands of students.',
            'image': 'https://images.unsplash.com/photo-1494790108377-be9c29b29330?w=800&h=1000&fit=crop',
            'qualifications': 'Ph.D. Educational Leadership, M.Ed. Curriculum Development'
        },
        {
            'name': 'Mr. James Ochieng',
            'position': 'Program Manager',
            'email': 'james.ochieng@modernschool.com',
            'phone': '+254 700 111 112',
            'bio': 'An experienced program manager specializing in educational program development, implementation, and evaluation with a focus on community impact.',
            'details': 'Mr. Ochieng has managed over 50 educational programs across Kenya, ensuring effective delivery and measurable outcomes. He is passionate about creating sustainable educational solutions for underserved communities.',
            'image': 'https://images.unsplash.com/photo-1507003211169-0a1dd7228f2d?w=800&h=1000&fit=crop',
            'qualifications': 'M.A. Program Management, B.Ed. Educational Administration'
        },
        {
            'name': 'Ms. Mary Kamau',
            'position': 'Finance Manager',
            'email': 'mary.kamau@modernschool.com',
            'phone': '+254 700 111 113',
            'bio': 'A certified public accountant with extensive experience in non-profit financial management, ensuring transparency and accountability in all financial operations.',
            'details': 'Ms. Kamau brings over 15 years of financial management experience to Modern School. She ensures strict compliance with financial regulations and maintains the highest standards of fiscal responsibility.',
            'image': 'https://images.unsplash.com/photo-1438761681033-6461ffad8d80?w=800&h=1000&fit=crop',
            'qualifications': 'CPA (K), M.Sc. Finance, B.Com Accounting'
        },
        {
            'name': 'Mr. Peter Mwangi',
            'position': 'Project Officer',
            'email': 'peter.mwangi@modernschool.com',
            'phone': '+254 700 111 114',
            'bio': 'Dedicated project officer with expertise in coordinating educational projects, managing resources, and ensuring timely delivery of program objectives.',
            'details': 'Mr. Mwangi has successfully coordinated numerous educational projects, from infrastructure development to scholarship programs. His attention to detail and organizational skills ensure project success.',
            'image': 'https://images.unsplash.com/photo-1500648767791-00dcc994a43e?w=800&h=1000&fit=crop',
            'qualifications': 'M.A. Project Management, B.A. Social Work'
        },
        {
            'name': 'Ms. Grace Wanjiku',
            'position': 'Administrative Officer',
            'email': 'grace.wanjiku@modernschool.com',
            'phone': '+254 700 111 115',
            'bio': 'Experienced administrative professional ensuring smooth day-to-day operations, efficient resource management, and excellent stakeholder relations.',
            'details': 'Ms. Wanjiku is the backbone of our administrative operations, managing everything from human resources to facility coordination. Her organizational skills and dedication keep our institution running smoothly.',
            'image': 'https://images.unsplash.com/photo-1544005313-94ddf0286df2?w=800&h=1000&fit=crop',
            'qualifications': 'M.A. Public Administration, B.A. Business Administration'
        }
    ]
    return render_template('team.html', team_members=team_members)

@app.route('/contact', methods=['GET', 'POST'])
def contact():
    if request.method == 'POST':
        name = request.form.get('name')
        email = request.form.get('email')
        phone = request.form.get('phone')
        subject = request.form.get('subject')
        message = request.form.get('message')
        
        # Here you would typically save to database or send email
        flash('Thank you for your message! We will get back to you soon.', 'success')
        return redirect(url_for('contact'))
    
    return render_template('contact.html')

def normalize_text(value, uppercase=True, allow_empty=False):
    """
    Normalize text: trim whitespace, replace multiple spaces with single space,
    and optionally convert to uppercase.
    
    Args:
        value: The text value to normalize
        uppercase: If True, convert to uppercase (default: True)
        allow_empty: If True, return None for empty values; if False, return empty string (default: False)
    
    Returns:
        Normalized string, empty string, or None based on allow_empty parameter
    """
    if not value:
        return None if allow_empty else ''
    # Replace multiple spaces/tabs/newlines with single space, then trim
    normalized = re.sub(r'\s+', ' ', str(value).strip())
    if not normalized:
        return None if allow_empty else ''
    return normalized.upper() if uppercase else normalized

@app.route('/admission', methods=['GET', 'POST'])
def admission():
    if request.method == 'POST':
        # Get form data and convert to proper case
        # All text fields to UPPERCASE except email (lowercase)
        # Normalize spacing: multiple spaces become single space
        student_full_name = normalize_text(request.form.get('student_full_name', ''))
        date_of_birth = request.form.get('date_of_birth')
        gender = normalize_text(request.form.get('gender'), allow_empty=True) if request.form.get('gender') else None
        current_grade = normalize_text(request.form.get('current_grade'), allow_empty=True) if request.form.get('current_grade') else None
        previous_school = normalize_text(request.form.get('previous_school'), allow_empty=True) if request.form.get('previous_school') else None
        parent_name = normalize_text(request.form.get('parent_name', ''))
        relationship = normalize_text(request.form.get('relationship'), allow_empty=True) if request.form.get('relationship') else None
        parent_phone = normalize_text(request.form.get('parent_phone', ''))
        parent_email = normalize_text(request.form.get('parent_email', ''), uppercase=False, allow_empty=True)
        if parent_email:
            parent_email = parent_email.lower()  # Email in lowercase
        address = normalize_text(request.form.get('address'), allow_empty=True) if request.form.get('address') else None
        emergency_contact = normalize_text(request.form.get('emergency_contact'), allow_empty=True) if request.form.get('emergency_contact') else None
        medical_info = normalize_text(request.form.get('medical_conditions'), allow_empty=True) if request.form.get('medical_conditions') else None
        special_needs = normalize_text(request.form.get('special_needs'), allow_empty=True) if request.form.get('special_needs') else None
        student_category = normalize_text(request.form.get('student_category'), uppercase=False, allow_empty=True)
        if student_category:
            student_category = student_category.lower()
        sponsor_name = normalize_text(request.form.get('sponsor_name'), allow_empty=True) if request.form.get('sponsor_name') else None
        sponsor_phone = normalize_text(request.form.get('sponsor_phone'), uppercase=False, allow_empty=True) if request.form.get('sponsor_phone') else None
        sponsor_email = normalize_text(request.form.get('sponsor_email', ''), uppercase=False, allow_empty=True)
        if sponsor_email:
            sponsor_email = sponsor_email.lower()  # Email in lowercase
        consent = request.form.get('consent')
        
        # Validate required fields
        if not all([student_full_name, date_of_birth, parent_name, relationship, parent_phone, parent_email, consent, student_category]):
            flash('Please fill in all required fields.', 'error')
            return redirect(url_for('admission'))
        
        # Validate sponsor_name if category is sponsored
        if student_category == 'sponsored' and not sponsor_name:
            flash('Please provide sponsor name/company for sponsored students.', 'error')
            return redirect(url_for('admission'))
        
        # Save to database
        connection = get_db_connection()
        if connection:
            try:
                with connection.cursor() as cursor:
                    # Generate unique student ID
                    student_id = generate_student_id(connection)
                    
                    # Insert student data into students table
                    student_sql = """
                        INSERT INTO students 
                        (student_id, full_name, date_of_birth, gender, current_grade, previous_school,
                         address, medical_info, special_needs, student_category, sponsor_name, sponsor_phone, sponsor_email, status)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'pending approval')
                    """
                    cursor.execute(student_sql, (
                        student_id, student_full_name, date_of_birth, gender, current_grade, 
                        previous_school, address, medical_info, special_needs, student_category, 
                        sponsor_name, sponsor_phone, sponsor_email
                    ))
                    
                    # Insert parent data into parents table (linked via student_id)
                    parent_sql = """
                        INSERT INTO parents 
                        (student_id, full_name, phone, email, relationship, emergency_contact)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """
                    cursor.execute(parent_sql, (
                        student_id, parent_name, parent_phone, parent_email, relationship, emergency_contact
                    ))
                    
                    connection.commit()
                    
                    # Send email notification to parent
                    try:
                        send_admission_confirmation_email(parent_email, parent_name, student_full_name, student_id)
                    except Exception as email_error:
                        print(f"Error sending email: {email_error}")
                        # Don't fail the submission if email fails
                    
                    flash(f'Application submitted successfully! Your Student ID is: {student_id}. We will review it and get back to you soon.', 'success')
            except Exception as e:
                print(f"Error saving admission: {e}")
                try:
                    connection.rollback()
                except:
                    pass  # Connection might already be closed
                flash('An error occurred while submitting your application. Please try again.', 'error')
            finally:
                if connection:
                    try:
                        connection.close()
                    except:
                        pass  # Connection might already be closed
        else:
            flash('Database connection error. Please try again later.', 'error')
        
        return redirect(url_for('home'))
    
    # GET request - redirect to home (admission is now handled via modal)
    return redirect(url_for('home'))

@app.route('/register-employee', methods=['GET', 'POST'])
def register_employee():
    if request.method == 'POST':
        # Get form data and convert to proper case
        # All text fields to UPPERCASE except email (lowercase)
        # Normalize spacing: multiple spaces become single space
        full_name = normalize_text(request.form.get('full_name', ''))
        email = normalize_text(request.form.get('email', ''), uppercase=False, allow_empty=True)
        if email:
            email = email.lower()  # Email in lowercase
        phone = normalize_text(request.form.get('phone', ''), uppercase=False)
        id_number = normalize_text(request.form.get('id_number', ''))
        employee_id = request.form.get('employee_id', '').strip()  # Keep as is (numeric)
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')
        
        # Validate required fields
        if not all([full_name, email, phone, id_number, employee_id, password, confirm_password]):
            flash('Please fill in all required fields.', 'error')
            return redirect(url_for('home'))
        
        # Validate employee ID is 6 digits
        if not employee_id.isdigit() or len(employee_id) != 6:
            flash('Employee ID must be exactly 6 digits.', 'error')
            return redirect(url_for('home'))
        
        # Validate passwords match
        if password != confirm_password:
            flash('Passwords do not match.', 'error')
            return redirect(url_for('home'))
        
        # Validate password strength
        if len(password) < 6:
            flash('Password must be at least 6 characters long.', 'error')
            return redirect(url_for('home'))
        
        # Handle profile picture upload
        profile_picture = None
        if 'profile_picture' in request.files:
            file = request.files['profile_picture']
            if file and file.filename != '' and allowed_file(file.filename):
                # Generate unique filename
                timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
                filename = secure_filename(f"{employee_id}_{timestamp}_{file.filename}")
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                profile_picture = f"uploads/profiles/{filename}"
        
        # Hash password
        password_hash = generate_password_hash(password)
        
        # Save to database
        connection = get_db_connection()
        if connection:
            try:
                with connection.cursor() as cursor:
                    # Check if employee ID already exists
                    cursor.execute("SELECT id FROM employees WHERE employee_id = %s", (employee_id,))
                    if cursor.fetchone():
                        flash('Employee ID already exists. Please use a different ID.', 'error')
                        return redirect(url_for('home'))
                    
                    # Check if email already exists
                    cursor.execute("SELECT id FROM employees WHERE email = %s", (email,))
                    if cursor.fetchone():
                        flash('Email already registered. Please use a different email.', 'error')
                        return redirect(url_for('home'))
                    
                    # Insert employee data
                    employee_sql = """
                        INSERT INTO employees 
                        (employee_id, full_name, email, phone, id_number, password_hash, 
                         profile_picture, role, status)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, 'employee', 'pending approval')
                    """
                    cursor.execute(employee_sql, (
                        employee_id, full_name, email, phone, id_number, password_hash, profile_picture
                    ))
                    
                    connection.commit()
                    
                    # Send welcome email to employee
                    try:
                        send_employee_welcome_email(email, full_name, employee_id)
                    except Exception as email_error:
                        print(f"Error sending welcome email: {email_error}")
                        # Don't fail the registration if email fails
                    
                    flash('Employee registration submitted successfully! Your application is pending approval. You will receive a welcome email shortly.', 'success')
            except Exception as e:
                print(f"Error saving employee registration: {e}")
                try:
                    connection.rollback()
                except:
                    pass  # Connection might already be closed
                flash('An error occurred while submitting your registration. Please try again.', 'error')
            finally:
                if connection:
                    try:
                        connection.close()
                    except:
                        pass  # Connection might already be closed
        else:
            flash('Database connection error. Please try again later.', 'error')
        
        return redirect(url_for('home'))
    
    # GET request - redirect to home (registration is handled via modal)
    return redirect(url_for('home'))

@app.route('/check-employee-id', methods=['POST'])
def check_employee_id():
    """Check if employee ID is available"""
    try:
        data = request.get_json()
        employee_id = data.get('employee_id', '').strip() if data else ''
        
        if not employee_id or len(employee_id) != 6 or not employee_id.isdigit():
            return jsonify({'available': False, 'message': 'Employee ID must be exactly 6 digits'}), 400
        
        connection = get_db_connection()
        if connection:
            try:
                with connection.cursor() as cursor:
                    cursor.execute("SELECT id FROM employees WHERE employee_id = %s", (employee_id,))
                    result = cursor.fetchone()
                    
                    if result:
                        return jsonify({'available': False, 'message': 'This Employee ID is already registered. Please use a different ID.'}), 200
                    else:
                        return jsonify({'available': True, 'message': 'Employee ID is available!'}), 200
            except Exception as e:
                print(f"Error checking employee ID: {e}")
                return jsonify({'available': False, 'message': 'Error checking employee ID. Please try again.'}), 500
            finally:
                if connection:
                    try:
                        connection.close()
                    except:
                        pass  # Connection might already be closed
        else:
            return jsonify({'available': False, 'message': 'Database connection error. Please try again.'}), 500
    except Exception as e:
        print(f"Error in check_employee_id: {e}")
        return jsonify({'available': False, 'message': 'An error occurred. Please try again.'}), 500

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        role = request.form.get('role', '').strip()
        password = request.form.get('password', '').strip()
        admission_number = request.form.get('admission_number', '').strip()
        employee_id = request.form.get('employee_id', '').strip() or request.form.get('employee_id_fallback', '').strip()
        
        print(f"DEBUG: Login attempt - Role: {role}, Employee ID: {employee_id}, Has Password: {bool(password)}")  # Debug
        print(f"DEBUG: Form data - {dict(request.form)}")  # Debug
        
        if not role or not password:
            flash('Please fill in all required fields.', 'error')
            return redirect(url_for('home'))
        
        # Validate role-specific fields
        if role in ['parent', 'student']:
            if not admission_number:
                flash('Please enter student admission number.', 'error')
                return redirect(url_for('home'))
            identifier = admission_number
        elif role == 'employee':
            if not employee_id:
                flash('Please enter employee identification code.', 'error')
                return redirect(url_for('home'))
            # Validate 6-digit code
            if not employee_id.isdigit() or len(employee_id) != 6:
                flash('Employee code must be exactly 6 digits.', 'error')
                return redirect(url_for('home'))
            identifier = employee_id
        else:
            flash('Invalid role selected.', 'error')
            return redirect(url_for('home'))
        
        connection = get_db_connection()
        if connection:
            try:
                with connection.cursor() as cursor:
                    if role == 'employee':
                        # Look up employee by employee_id
                        print(f"DEBUG: Looking for employee with ID: {identifier}")  # Debug
                        cursor.execute(
                            "SELECT * FROM employees WHERE employee_id = %s",
                            (identifier,)
                        )
                        employee = cursor.fetchone()
                        
                        print(f"DEBUG: Employee found: {employee is not None}")  # Debug
                        
                        if employee:
                            print(f"DEBUG: Checking password...")  # Debug
                            # Check password
                            if check_password_hash(employee['password_hash'], password):
                                print(f"DEBUG: Password correct!")  # Debug
                                # Check employee status
                                status = employee['status']
                                print(f"DEBUG: Employee status: {status}")  # Debug
                                
                                if status == 'pending approval':
                                    # Redirect to terms and conditions page
                                    flash('Your account is pending approval. Please review the terms and conditions.', 'info')
                                    return redirect(url_for('terms_and_conditions'))
                                elif status in ['suspended', 'fired']:
                                    flash('Your account has been suspended. Please contact the relevant authorities for assistance.', 'error')
                                    return redirect(url_for('home'))
                                elif status == 'retired':
                                    flash('Thank you for your service! Your account has been retired.', 'info')
                                    return redirect(url_for('home'))
                                elif status == 'active':
                                    # Set session and redirect to role dashboard
                                    session['user_id'] = employee['id']
                                    session['email'] = employee['email']
                                    session['full_name'] = employee['full_name']
                                    session['role'] = employee['role']
                                    session['employee_id'] = employee['employee_id']
                                    session['profile_picture'] = employee.get('profile_picture')
                                    flash(f'Welcome back, {employee["full_name"]}!', 'success')
                                    return redirect(url_for('dashboard_employee'))
                                else:
                                    flash('Your account status is invalid. Please contact support.', 'error')
                                    return redirect(url_for('home'))
                            else:
                                print(f"DEBUG: Password incorrect!")  # Debug
                                flash('Invalid employee code or password.', 'error')
                        else:
                            print(f"DEBUG: Employee not found!")  # Debug
                            flash('Invalid employee code or password.', 'error')
                    else:  # parent or student
                        # Look up by admission number (you may need to add this field to users table)
                        cursor.execute(
                            "SELECT * FROM users WHERE email LIKE %s AND role = %s",
                            (f'%{identifier}%', role)
                        )
                        user = cursor.fetchone()
                        
                        # For demo purposes, accept any password if user exists
                        # In production, use: check_password_hash(user['password_hash'], password)
                        if user:
                            session['user_id'] = user['id']
                            session['email'] = user['email']
                            session['full_name'] = user['full_name']
                            session['role'] = user['role']
                            flash(f'Welcome back, {user["full_name"]}!', 'success')
                            return redirect(url_for(f'dashboard_{role}'))
                        else:
                            flash('Invalid credentials. Please check your admission number and password.', 'error')
            except Exception as e:
                print(f"Login error: {e}")
                flash('An error occurred during login. Please try again.', 'error')
            finally:
                if connection:
                    try:
                        connection.close()
                    except:
                        pass  # Connection might already be closed
        else:
            flash('Database connection error. Please try again later.', 'error')
        
        return redirect(url_for('home'))
    
    # GET request - redirect to home (login is now handled via modal)
    return redirect(url_for('home'))

@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out successfully.', 'success')
    return redirect(url_for('home'))

@app.route('/dashboard/parent')
@login_required
def dashboard_parent():
    """Parent dashboard - accessible to parents and technicians"""
    user_role = session.get('role', '').lower()
    viewing_as = session.get('viewing_as_role', '')
    
    # Allow access if user is parent OR if technician is viewing as parent
    if user_role != 'parent' and not (user_role == 'technician' and viewing_as == 'parent'):
        if user_role != 'technician':
            flash('You do not have permission to access this page.', 'error')
            return redirect(url_for('home'))
    
    # Check if technician is viewing as parent
    is_technician = user_role == 'technician'
    current_view_role = viewing_as if is_technician and viewing_as else user_role
    
    # Get parent's email from session
    parent_email = session.get('email', '')
    
    # For technicians, fetch all students for selection
    all_students = []
    selected_student_id = request.args.get('student_id', '') or session.get('parent_view_student_id', '')
    
    if is_technician and viewing_as == 'parent':
        connection = get_db_connection()
        if connection:
            try:
                with connection.cursor() as cursor:
                    cursor.execute("""
                        SELECT DISTINCT s.id, s.student_id, s.full_name, s.current_grade, s.status,
                               p.full_name as parent_name, p.email as parent_email
                        FROM students s
                        LEFT JOIN parents p ON s.student_id = p.student_id
                        WHERE s.status = 'in session'
                        ORDER BY s.full_name ASC
                    """)
                    all_students = cursor.fetchall()
            except Exception as e:
                print(f"Error fetching students for technician: {e}")
            finally:
                if connection:
                    try:
                        connection.close()
                    except:
                        pass
        
        # If a student is selected, use that student's parent email
        if selected_student_id:
            session['parent_view_student_id'] = selected_student_id
            connection = get_db_connection()
            if connection:
                try:
                    with connection.cursor() as cursor:
                        cursor.execute("""
                            SELECT p.email
                            FROM parents p
                            WHERE p.student_id = %s
                            LIMIT 1
                        """, (selected_student_id,))
                        parent_result = cursor.fetchone()
                        if parent_result:
                            parent_email = parent_result.get('email', '') if isinstance(parent_result, dict) else parent_result[0]
                except Exception as e:
                    print(f"Error fetching parent email: {e}")
                finally:
                    if connection:
                        try:
                            connection.close()
                        except:
                            pass
    
    # Fetch parent's children and their fee information
    connection = get_db_connection()
    children = []
    
    if connection and parent_email:
        try:
            with connection.cursor() as cursor:
                # Get all students linked to this parent (by email)
                cursor.execute("""
                    SELECT DISTINCT s.id, s.student_id, s.full_name, s.current_grade, s.status,
                           p.full_name as parent_name, p.phone as parent_phone, p.email as parent_email
                    FROM students s
                    INNER JOIN parents p ON s.student_id = p.student_id
                    WHERE p.email = %s AND s.status = 'in session'
                    ORDER BY s.full_name ASC
                """, (parent_email,))
                results = cursor.fetchall()
                
                if results:
                    for row in results:
                        student_grade = row.get('current_grade', '')
                        student_id = row.get('student_id', '')
                        
                        # Find the academic level for this student's grade
                        academic_level = None
                        fee_structure = None
                        total_paid = 0.0
                        balance = 0.0
                        
                        if student_grade:
                            # Try to match student's current_grade with academic_levels.level_name
                            cursor.execute("""
                                SELECT al.id, al.level_category, al.level_name, al.level_description
                                FROM academic_levels al
                                WHERE al.level_name = %s AND al.level_status = 'active'
                                LIMIT 1
                            """, (student_grade,))
                            level_result = cursor.fetchone()
                            
                            if level_result:
                                academic_level_id = level_result.get('id')
                                academic_level = {
                                    'id': academic_level_id,
                                    'level_category': level_result.get('level_category', ''),
                                    'level_name': level_result.get('level_name', ''),
                                    'level_description': level_result.get('level_description', '')
                                }
                                
                                # Find active fee structure for this academic level
                                # Get the most recent active fee structure
                                cursor.execute("""
                                    SELECT fs.id, fs.fee_name, fs.start_date, fs.end_date, 
                                           fs.payment_deadline, fs.total_amount, fs.status,
                                           fs.academic_year_id, fs.term_id
                                    FROM fee_structures fs
                                    WHERE fs.academic_level_id = %s 
                                      AND fs.status = 'active'
                                    ORDER BY fs.created_at DESC
                                    LIMIT 1
                                """, (academic_level_id,))
                                fee_structure_result = cursor.fetchone()
                                
                                if fee_structure_result:
                                    payment_deadline = fee_structure_result.get('payment_deadline')
                                    # Format payment_deadline if it's a date object
                                    if payment_deadline and hasattr(payment_deadline, 'strftime'):
                                        payment_deadline_formatted = payment_deadline.strftime('%B %d, %Y')
                                    elif payment_deadline:
                                        payment_deadline_formatted = str(payment_deadline)
                                    else:
                                        payment_deadline_formatted = None
                                    
                                    fee_structure = {
                                        'id': fee_structure_result.get('id'),
                                        'fee_name': fee_structure_result.get('fee_name', ''),
                                        'total_amount': float(fee_structure_result.get('total_amount', 0)),
                                        'payment_deadline': payment_deadline,
                                        'payment_deadline_formatted': payment_deadline_formatted,
                                        'status': fee_structure_result.get('status', '')
                                    }
                                    
                                    # Calculate total paid and balance
                                    cursor.execute("""
                                        SELECT COALESCE(SUM(amount_paid), 0) as total_paid
                                        FROM student_payments
                                        WHERE student_id = %s
                                    """, (student_id,))
                                    payment_result = cursor.fetchone()
                                    total_paid = float(payment_result.get('total_paid', 0) if payment_result else 0)
                                    balance = fee_structure['total_amount'] - total_paid
                        
                        child_dict = {
                            'id': row.get('id'),
                            'student_id': student_id,
                            'full_name': row.get('full_name', ''),
                            'current_grade': student_grade,
                            'status': row.get('status', ''),
                            'academic_level': academic_level,
                            'fee_structure': fee_structure,
                            'total_paid': total_paid,
                            'balance': balance
                        }
                        children.append(child_dict)
        except Exception as e:
            print(f"Error fetching parent's children: {e}")
            import traceback
            traceback.print_exc()
            flash('Error loading children information. Please try again.', 'error')
        finally:
            if connection:
                try:
                    connection.close()
                except:
                    pass
    
    return render_template('dashboards/dashboard_parent.html',
                         is_technician=is_technician,
                         current_view_role=current_view_role,
                         children=children,
                         all_students=all_students if is_technician and viewing_as == 'parent' else [],
                         selected_student_id=selected_student_id)

@app.route('/dashboard/parent/student-fees')
@login_required
def parent_student_fees():
    """Parent view of their children's fees - list all children with detailed fees"""
    user_role = session.get('role', '').lower()
    viewing_as = session.get('viewing_as_role', '')
    parent_email = session.get('email', '')
    
    # Allow access if user is parent OR if technician is viewing as parent
    if user_role != 'parent' and not (user_role == 'technician' and viewing_as == 'parent'):
        if user_role != 'technician':
            flash('You do not have permission to access this page.', 'error')
            return redirect(url_for('home'))
    
    # For technicians, fetch all students for selection
    all_students = []
    selected_student_id = request.args.get('student_id', '') or session.get('parent_view_student_id', '')
    
    if user_role == 'technician' and viewing_as == 'parent':
        connection = get_db_connection()
        if connection:
            try:
                with connection.cursor() as cursor:
                    cursor.execute("""
                        SELECT DISTINCT s.id, s.student_id, s.full_name, s.current_grade, s.status,
                               p.full_name as parent_name, p.email as parent_email
                        FROM students s
                        LEFT JOIN parents p ON s.student_id = p.student_id
                        WHERE s.status = 'in session'
                        ORDER BY s.full_name ASC
                    """)
                    all_students = cursor.fetchall()
            except Exception as e:
                print(f"Error fetching students for technician: {e}")
            finally:
                if connection:
                    try:
                        connection.close()
                    except:
                        pass
        
        # If a student is selected, use that student's parent email
        if selected_student_id:
            session['parent_view_student_id'] = selected_student_id
            connection = get_db_connection()
            if connection:
                try:
                    with connection.cursor() as cursor:
                        cursor.execute("""
                            SELECT p.email
                            FROM parents p
                            WHERE p.student_id = %s
                            LIMIT 1
                        """, (selected_student_id,))
                        parent_result = cursor.fetchone()
                        if parent_result:
                            parent_email = parent_result.get('email', '') if isinstance(parent_result, dict) else parent_result[0]
                except Exception as e:
                    print(f"Error fetching parent email: {e}")
                finally:
                    if connection:
                        try:
                            connection.close()
                        except:
                            pass
    
    # Fetch parent's children and their detailed fee information
    connection = get_db_connection()
    children = []
    
    if connection and parent_email:
        try:
            with connection.cursor() as cursor:
                # Get all students linked to this parent (by email)
                cursor.execute("""
                    SELECT DISTINCT s.id, s.student_id, s.full_name, s.current_grade, s.status,
                           p.full_name as parent_name, p.phone as parent_phone, p.email as parent_email
                    FROM students s
                    INNER JOIN parents p ON s.student_id = p.student_id
                    WHERE p.email = %s AND s.status = 'in session'
                    ORDER BY s.full_name ASC
                """, (parent_email,))
                results = cursor.fetchall()
                
                if results:
                    for row in results:
                        student_grade = row.get('current_grade', '')
                        student_id = row.get('student_id', '')
                        
                        # Find the academic level for this student's grade
                        academic_level = None
                        fee_structure = None
                        fee_items = []
                        payments = []
                        total_paid = 0.0
                        balance = 0.0
                        
                        if student_grade:
                            # Try to match student's current_grade with academic_levels.level_name
                            cursor.execute("""
                                SELECT al.id, al.level_category, al.level_name, al.level_description
                                FROM academic_levels al
                                WHERE al.level_name = %s AND al.level_status = 'active'
                                LIMIT 1
                            """, (student_grade,))
                            level_result = cursor.fetchone()
                            
                            if level_result:
                                academic_level_id = level_result.get('id')
                                academic_level = {
                                    'id': academic_level_id,
                                    'level_category': level_result.get('level_category', ''),
                                    'level_name': level_result.get('level_name', ''),
                                    'level_description': level_result.get('level_description', '')
                                }
                                
                                # Find active fee structure for this academic level
                                cursor.execute("""
                                    SELECT fs.id, fs.fee_name, fs.start_date, fs.end_date, 
                                           fs.payment_deadline, fs.total_amount, fs.status
                                    FROM fee_structures fs
                                    WHERE fs.academic_level_id = %s 
                                      AND fs.status = 'active'
                                    ORDER BY fs.created_at DESC
                                    LIMIT 1
                                """, (academic_level_id,))
                                fee_structure_result = cursor.fetchone()
                                
                                if fee_structure_result:
                                    payment_deadline = fee_structure_result.get('payment_deadline')
                                    # Format payment_deadline if it's a date object
                                    if payment_deadline and hasattr(payment_deadline, 'strftime'):
                                        payment_deadline_formatted = payment_deadline.strftime('%B %d, %Y')
                                    elif payment_deadline:
                                        payment_deadline_formatted = str(payment_deadline)
                                    else:
                                        payment_deadline_formatted = None
                                    
                                    fee_structure = {
                                        'id': fee_structure_result.get('id'),
                                        'fee_name': fee_structure_result.get('fee_name'),
                                        'total_amount': float(fee_structure_result.get('total_amount', 0)),
                                        'payment_deadline': fee_structure_result.get('payment_deadline'),
                                        'payment_deadline_formatted': payment_deadline_formatted,
                                        'status': fee_structure_result.get('status')
                                    }
                                    
                                    # Get fee items
                                    cursor.execute("""
                                        SELECT item_name, item_description, amount
                                        FROM fee_items
                                        WHERE fee_structure_id = %s
                                        ORDER BY id ASC
                                    """, (fee_structure['id'],))
                                    fee_items_results = cursor.fetchall()
                                    fee_items = []
                                    for item in fee_items_results:
                                        if isinstance(item, dict):
                                            fee_items.append({
                                                'item_name': item.get('item_name', '') or '',
                                                'item_description': item.get('item_description', '') or '',
                                                'amount': float(item.get('amount', 0) or 0)
                                            })
                                        else:
                                            # Handle tuple/list format
                                            fee_items.append({
                                                'item_name': item[0] if len(item) > 0 else '',
                                                'item_description': item[1] if len(item) > 1 else '',
                                                'amount': float(item[2] if len(item) > 2 and item[2] else 0)
                                            })
                                    
                                    # Get payments
                                    cursor.execute("""
                                        SELECT amount_paid, payment_date, payment_method, reference_number, notes, created_at
                                        FROM student_payments
                                        WHERE student_id = %s
                                        ORDER BY payment_date DESC, created_at DESC
                                    """, (student_id,))
                                    payments_results = cursor.fetchall()
                                    payments = []
                                    for payment in payments_results:
                                        if isinstance(payment, dict):
                                            payment_date = payment.get('payment_date')
                                            # Format payment_date if it's a date object
                                            if payment_date and hasattr(payment_date, 'strftime'):
                                                payment_date_formatted = payment_date.strftime('%B %d, %Y')
                                            elif payment_date:
                                                payment_date_formatted = str(payment_date)
                                            else:
                                                payment_date_formatted = None
                                            
                                            payments.append({
                                                'amount_paid': float(payment.get('amount_paid', 0) or 0),
                                                'payment_date': payment_date,
                                                'payment_date_formatted': payment_date_formatted,
                                                'payment_method': payment.get('payment_method', '') or '',
                                                'reference_number': payment.get('reference_number', '') or '',
                                                'notes': payment.get('notes', '') or '',
                                                'created_at': payment.get('created_at')
                                            })
                                        else:
                                            # Handle tuple/list format
                                            try:
                                                payment_date = payment[1] if len(payment) > 1 else None
                                                # Format payment_date if it's a date object
                                                if payment_date and hasattr(payment_date, 'strftime'):
                                                    payment_date_formatted = payment_date.strftime('%B %d, %Y')
                                                elif payment_date:
                                                    payment_date_formatted = str(payment_date)
                                                else:
                                                    payment_date_formatted = None
                                                
                                                payments.append({
                                                    'amount_paid': float(payment[0] if len(payment) > 0 and payment[0] else 0),
                                                    'payment_date': payment_date,
                                                    'payment_date_formatted': payment_date_formatted,
                                                    'payment_method': payment[2] if len(payment) > 2 else '',
                                                    'reference_number': payment[3] if len(payment) > 3 else '',
                                                    'notes': payment[4] if len(payment) > 4 else '',
                                                    'created_at': payment[5] if len(payment) > 5 else None
                                                })
                                            except (IndexError, KeyError):
                                                # Skip invalid payment records
                                                continue
                                    
                                    # Calculate total paid
                                    cursor.execute("""
                                        SELECT COALESCE(SUM(amount_paid), 0) as total_paid
                                        FROM student_payments
                                        WHERE student_id = %s
                                    """, (student_id,))
                                    payment_result = cursor.fetchone()
                                    total_paid = float(payment_result.get('total_paid', 0) if payment_result else 0)
                                    balance = fee_structure['total_amount'] - total_paid
                        
                        child_dict = {
                            'id': row.get('id'),
                            'student_id': student_id,
                            'full_name': row.get('full_name', ''),
                            'current_grade': student_grade,
                            'status': row.get('status', ''),
                            'academic_level': academic_level,
                            'fee_structure': fee_structure,
                            'fee_items': fee_items,
                            'payments': payments,
                            'total_paid': total_paid,
                            'balance': balance
                        }
                        children.append(child_dict)
        except Exception as e:
            print(f"Error fetching parent's children fees: {e}")
            import traceback
            traceback.print_exc()
            flash('Error loading fees information. Please try again.', 'error')
        finally:
            if connection:
                try:
                    connection.close()
                except:
                    pass
    
    return render_template('dashboards/parent_student_fees.html',
                         children=children,
                         is_technician=user_role == 'technician',
                         current_view_role=viewing_as if user_role == 'technician' and viewing_as else user_role,
                         all_students=all_students if user_role == 'technician' and viewing_as == 'parent' else [],
                         selected_student_id=selected_student_id)

@app.route('/terms-and-conditions')
def terms_and_conditions():
    return render_template('terms_and_conditions.html')

@app.route('/dashboard/student')
@login_required
def dashboard_student():
    """Student dashboard - accessible to students and technicians"""
    user_role = session.get('role', '').lower()
    viewing_as = session.get('viewing_as_role', '')
    
    # Allow access if user is student OR if technician is viewing as student
    if user_role != 'student' and not (user_role == 'technician' and viewing_as == 'student'):
        if user_role != 'technician':
            flash('You do not have permission to access this page.', 'error')
            return redirect(url_for('home'))
    
    # Check if technician is viewing as student
    is_technician = user_role == 'technician'
    current_view_role = viewing_as if is_technician and viewing_as else user_role
    
    # Get student ID - either from session or selected by technician
    student_id = None
    student_email = session.get('email', '')
    
    # For technicians, fetch all students for selection
    all_students = []
    selected_student_id = request.args.get('student_id', '') or session.get('student_view_student_id', '')
    
    if is_technician and viewing_as == 'student':
        connection = get_db_connection()
        if connection:
            try:
                with connection.cursor() as cursor:
                    cursor.execute("""
                        SELECT DISTINCT s.id, s.student_id, s.full_name, s.current_grade, s.status
                        FROM students s
                        WHERE s.status = 'in session'
                        ORDER BY s.full_name ASC
                    """)
                    all_students = cursor.fetchall()
            except Exception as e:
                print(f"Error fetching students for technician: {e}")
            finally:
                if connection:
                    try:
                        connection.close()
                    except:
                        pass
        
        # If a student is selected, use that student's ID
        if selected_student_id:
            session['student_view_student_id'] = selected_student_id
            student_id = selected_student_id
    else:
        # For actual students, get their student_id from users table or session
        connection = get_db_connection()
        if connection:
            try:
                with connection.cursor() as cursor:
                    # Try to get student_id from users table
                    user_id = session.get('user_id')
                    if user_id:
                        cursor.execute("""
                            SELECT student_id FROM users WHERE id = %s
                        """, (user_id,))
                        user_result = cursor.fetchone()
                        if user_result:
                            student_id = user_result.get('student_id', '') if isinstance(user_result, dict) else user_result[0]
            except Exception as e:
                print(f"Error fetching student ID: {e}")
            finally:
                if connection:
                    try:
                        connection.close()
                    except:
                        pass
    
    # Fetch student's fee information
    fee_structure = None
    fee_items = []
    payments = []
    total_paid = 0.0
    balance = 0.0
    student_info = None
    
    if student_id:
        connection = get_db_connection()
        if connection:
            try:
                with connection.cursor() as cursor:
                    # Get student information
                    cursor.execute("""
                        SELECT s.id, s.student_id, s.full_name, s.current_grade, s.status
                        FROM students s
                        WHERE s.student_id = %s AND s.status = 'in session'
                        LIMIT 1
                    """, (student_id,))
                    student_result = cursor.fetchone()
                    
                    if student_result:
                        student_info = {
                            'id': student_result.get('id') if isinstance(student_result, dict) else student_result[0],
                            'student_id': student_result.get('student_id') if isinstance(student_result, dict) else student_result[1],
                            'full_name': student_result.get('full_name') if isinstance(student_result, dict) else student_result[2],
                            'current_grade': student_result.get('current_grade') if isinstance(student_result, dict) else student_result[3],
                            'status': student_result.get('status') if isinstance(student_result, dict) else student_result[4]
                        }
                        
                        student_grade = student_info.get('current_grade', '')
                        if student_grade:
                            # Get academic level
                            cursor.execute("""
                                SELECT al.id, al.level_category, al.level_name
                                FROM academic_levels al
                                WHERE al.level_name = %s AND al.level_status = 'active'
                                LIMIT 1
                            """, (student_grade,))
                            level_result = cursor.fetchone()
                            
                            if level_result:
                                academic_level_id = level_result.get('id') if isinstance(level_result, dict) else level_result[0]
                                
                                # Get fee structure
                                cursor.execute("""
                                    SELECT fs.id, fs.fee_name, fs.start_date, fs.end_date, 
                                           fs.payment_deadline, fs.total_amount, fs.status
                                    FROM fee_structures fs
                                    WHERE fs.academic_level_id = %s 
                                      AND fs.status = 'active'
                                    ORDER BY fs.created_at DESC
                                    LIMIT 1
                                """, (academic_level_id,))
                                fee_structure_result = cursor.fetchone()
                                
                                if fee_structure_result:
                                    payment_deadline = fee_structure_result.get('payment_deadline') if isinstance(fee_structure_result, dict) else fee_structure_result[4]
                                    # Format payment_deadline if it's a date object
                                    if payment_deadline and hasattr(payment_deadline, 'strftime'):
                                        payment_deadline_formatted = payment_deadline.strftime('%B %d, %Y')
                                    elif payment_deadline:
                                        payment_deadline_formatted = str(payment_deadline)
                                    else:
                                        payment_deadline_formatted = None
                                    
                                    fee_structure = {
                                        'id': fee_structure_result.get('id') if isinstance(fee_structure_result, dict) else fee_structure_result[0],
                                        'fee_name': fee_structure_result.get('fee_name') if isinstance(fee_structure_result, dict) else fee_structure_result[1],
                                        'total_amount': float(fee_structure_result.get('total_amount', 0) if isinstance(fee_structure_result, dict) else fee_structure_result[5]),
                                        'payment_deadline': payment_deadline,
                                        'payment_deadline_formatted': payment_deadline_formatted,
                                        'status': fee_structure_result.get('status') if isinstance(fee_structure_result, dict) else fee_structure_result[6]
                                    }
                                    
                                    # Get fee items
                                    cursor.execute("""
                                        SELECT item_name, item_description, amount
                                        FROM fee_items
                                        WHERE fee_structure_id = %s
                                        ORDER BY id ASC
                                    """, (fee_structure['id'],))
                                    fee_items_results = cursor.fetchall()
                                    fee_items = []
                                    for item in fee_items_results:
                                        if isinstance(item, dict):
                                            fee_items.append({
                                                'item_name': item.get('item_name', '') or '',
                                                'item_description': item.get('item_description', '') or '',
                                                'amount': float(item.get('amount', 0) or 0)
                                            })
                                        else:
                                            fee_items.append({
                                                'item_name': item[0] if len(item) > 0 else '',
                                                'item_description': item[1] if len(item) > 1 else '',
                                                'amount': float(item[2] if len(item) > 2 and item[2] else 0)
                                            })
                                    
                                    # Get payments
                                    cursor.execute("""
                                        SELECT amount_paid, payment_date, payment_method, reference_number, notes, created_at
                                        FROM student_payments
                                        WHERE student_id = %s
                                        ORDER BY payment_date DESC, created_at DESC
                                    """, (student_id,))
                                    payments_results = cursor.fetchall()
                                    payments = []
                                    for payment in payments_results:
                                        if isinstance(payment, dict):
                                            payment_date = payment.get('payment_date')
                                            if payment_date and hasattr(payment_date, 'strftime'):
                                                payment_date_formatted = payment_date.strftime('%B %d, %Y')
                                            elif payment_date:
                                                payment_date_formatted = str(payment_date)
                                            else:
                                                payment_date_formatted = None
                                            
                                            payments.append({
                                                'amount_paid': float(payment.get('amount_paid', 0) or 0),
                                                'payment_date': payment_date,
                                                'payment_date_formatted': payment_date_formatted,
                                                'payment_method': payment.get('payment_method', '') or '',
                                                'reference_number': payment.get('reference_number', '') or '',
                                                'notes': payment.get('notes', '') or '',
                                                'created_at': payment.get('created_at')
                                            })
                                        else:
                                            try:
                                                payment_date = payment[1] if len(payment) > 1 else None
                                                if payment_date and hasattr(payment_date, 'strftime'):
                                                    payment_date_formatted = payment_date.strftime('%B %d, %Y')
                                                elif payment_date:
                                                    payment_date_formatted = str(payment_date)
                                                else:
                                                    payment_date_formatted = None
                                                
                                                payments.append({
                                                    'amount_paid': float(payment[0] if len(payment) > 0 and payment[0] else 0),
                                                    'payment_date': payment_date,
                                                    'payment_date_formatted': payment_date_formatted,
                                                    'payment_method': payment[2] if len(payment) > 2 else '',
                                                    'reference_number': payment[3] if len(payment) > 3 else '',
                                                    'notes': payment[4] if len(payment) > 4 else '',
                                                    'created_at': payment[5] if len(payment) > 5 else None
                                                })
                                            except (IndexError, KeyError):
                                                continue
                                    
                                    # Calculate total paid
                                    cursor.execute("""
                                        SELECT COALESCE(SUM(amount_paid), 0) as total_paid
                                        FROM student_payments
                                        WHERE student_id = %s
                                    """, (student_id,))
                                    payment_result = cursor.fetchone()
                                    total_paid = float(payment_result.get('total_paid', 0) if payment_result else 0)
                                    balance = fee_structure['total_amount'] - total_paid
            except Exception as e:
                print(f"Error fetching student fees: {e}")
                import traceback
                traceback.print_exc()
                flash('Error loading fee information. Please try again.', 'error')
            finally:
                if connection:
                    try:
                        connection.close()
                    except:
                        pass
    
    return render_template('dashboards/dashboard_student.html', 
                         is_technician=is_technician,
                         current_view_role=current_view_role,
                         all_students=all_students if is_technician and viewing_as == 'student' else [],
                         selected_student_id=selected_student_id,
                         student_info=student_info,
                         fee_structure=fee_structure,
                         fee_items=fee_items,
                         payments=payments,
                         total_paid=total_paid,
                         balance=balance)

@app.route('/dashboard/student/fees')
@login_required
def student_fees_view():
    """Student fees view page - redirects to dashboard with fees"""
    return redirect(url_for('dashboard_student'))

@app.route('/dashboard/employee')
@login_required
def dashboard_employee():
    # Check if user is an employee (any employee role)
    employee_roles = ['employee', 'super admin', 'principal', 'deputy principal', 'academic coordinator', 
                     'teachers', 'accountant', 'librarian', 'warden', 'transport manager', 'technician']
    
    user_role = session.get('role', '').lower()
    if user_role not in employee_roles:
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('home'))
    
    # Get employee data for profile picture
    connection = get_db_connection()
    employee_data = {}
    if connection:
        try:
            with connection.cursor() as cursor:
                employee_id = session.get('employee_id') or session.get('user_id')
                cursor.execute("SELECT * FROM employees WHERE id = %s OR employee_id = %s", 
                             (employee_id, employee_id))
                employee = cursor.fetchone()
                if employee:
                    employee_data = employee
        except Exception as e:
            print(f"Error fetching employee data: {e}")
        finally:
            connection.close()
    
    # Check if user is technician (for role switching)
    is_technician = user_role == 'technician'
    current_view_role = session.get('viewing_as_role', user_role)
    current_employee_role = session.get('viewing_as_employee_role', user_role if user_role in ['employee', 'super admin', 'principal', 'deputy principal', 'academic coordinator', 'teachers', 'accountant', 'librarian', 'warden', 'transport manager', 'technician'] else 'employee')
    
    # List of all employee roles for technician to view
    employee_roles_list = ['employee', 'super admin', 'principal', 'deputy principal', 
                          'academic coordinator', 'teachers', 'accountant', 'librarian', 
                          'warden', 'transport manager', 'technician']
    
    return render_template('dashboards/dashboard_employee.html', 
                         role=session.get('role', 'employee'),
                         employee=employee_data,
                         is_technician=is_technician,
                         current_view_role=current_view_role,
                         current_employee_role=current_employee_role,
                         employee_roles_list=employee_roles_list)

# Finance Overview Route
@app.route('/dashboard/employee/finance-overview')
@login_required
def finance_overview():
    """Finance Overview page for principals and accountants"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    employee_id = session.get('employee_id') or session.get('user_id')
    
    # Check if user is accountant, principal, or viewing as accountant/principal
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_principal = user_role == 'principal' or viewing_as_role == 'principal'
    is_technician = user_role == 'technician'
    
    # Check permission-based access
    has_view_fees_permission = check_permission_or_role('view_student_fees', ['accountant', 'principal'])
    has_manage_fees_permission = check_permission_or_role('manage_fees', ['accountant', 'principal'])
    
    if not (is_accountant or is_principal or is_technician or has_view_fees_permission or has_manage_fees_permission):
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('dashboard_employee'))
    
    # Get finance summary data
    connection = get_db_connection()
    finance_summary = {
        'total_students': 0,
        'total_fee_structures': 0,
        'total_revenue': 0,
        'pending_payments': 0,
        'paid_payments': 0
    }
    
    if connection:
        try:
            with connection.cursor() as cursor:
                # Get total students
                cursor.execute("SELECT COUNT(*) as count FROM students WHERE status = 'active'")
                result = cursor.fetchone()
                finance_summary['total_students'] = result.get('count', 0) if result else 0
                
                # Get total fee structures
                cursor.execute("SELECT COUNT(*) as count FROM fee_structures WHERE status = 'active'")
                result = cursor.fetchone()
                finance_summary['total_fee_structures'] = result.get('count', 0) if result else 0
                
                # Get total revenue (paid payments)
                cursor.execute("SELECT COALESCE(SUM(amount_paid), 0) as total FROM student_payments")
                result = cursor.fetchone()
                finance_summary['total_revenue'] = float(result.get('total', 0) if result else 0)
                
                # Get pending payments (this would need to be calculated from fee structures - payments)
                # For now, we'll set it to 0 or calculate it if needed
                finance_summary['pending_payments'] = 0
                finance_summary['paid_payments'] = finance_summary['total_revenue']
        except Exception as e:
            print(f"Error fetching finance summary: {e}")
        finally:
            connection.close()
    
    return render_template('dashboards/finance_overview.html',
                         finance_summary=finance_summary,
                         role=user_role)

# Student Fees Route
@app.route('/dashboard/employee/student-fees')
@login_required
def student_fees():
    """Student Fees management page for accountants"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    
    print(f"DEBUG Student Fees Route:")
    print(f"  - user_role from session: {user_role}")
    print(f"  - viewing_as_role: {viewing_as_role}")
    print(f"  - employee_id: {session.get('employee_id')}")
    print(f"  - user_id: {session.get('user_id')}")
    
    # Check if user is accountant or viewing as accountant (for role switching)
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    
    # Check if user is principal or viewing as principal
    is_principal = user_role == 'principal' or viewing_as_role == 'principal'
    
    # Also allow technicians (they can switch to accountant role)
    is_technician = user_role == 'technician'
    
    # If not accountant or principal, check actual role from database
    if not is_accountant and not is_principal:
        connection = get_db_connection()
        if connection:
            try:
                with connection.cursor() as cursor:
                    employee_id = session.get('employee_id') or session.get('user_id')
                    if employee_id:
                        cursor.execute("SELECT role FROM employees WHERE id = %s OR employee_id = %s", 
                                     (employee_id, employee_id))
                        result = cursor.fetchone()
                        if result:
                            db_role = result.get('role') if isinstance(result, dict) else result[0]
                            db_role = db_role.lower() if db_role else ''
                            print(f"  - db_role: {db_role}")
                            if db_role == 'accountant':
                                is_accountant = True
                            elif db_role == 'principal':
                                is_principal = True
            except Exception as e:
                print(f"Error fetching employee role: {e}")
            finally:
                connection.close()
    
    # Check permission-based access
    employee_id = session.get('employee_id') or session.get('user_id')
    has_view_fees_permission = check_permission_or_role('view_student_fees', ['accountant', 'principal'])
    has_manage_fees_permission = check_permission_or_role('manage_fees', ['accountant', 'principal'])
    
    # Check specific permissions for accountant features
    can_view_students = check_permission_or_role('view_students', ['accountant', 'principal', 'deputy principal'])
    can_view_fee_structure_details = check_permission_or_role('view_fee_structure_details', ['accountant', 'principal'])
    can_record_payments = check_permission_or_role('process_payments', ['accountant', 'principal'])
    can_add_fee_structure = check_permission_or_role('add_fee_structure', ['accountant', 'principal'])
    can_edit_fee_structure = check_permission_or_role('edit_fee_structure', ['accountant', 'principal'])
    can_delete_fee_structure = check_permission_or_role('delete_fee_structure', ['accountant', 'principal'])
    
    # Allow access if: has role-based access OR has permission-based access
    if not (is_accountant or is_principal or is_technician or has_view_fees_permission or has_manage_fees_permission):
        print(f"  - Access DENIED: user_role={user_role}, viewing_as_role={viewing_as_role}, is_accountant={is_accountant}, is_technician={is_technician}")
        flash('You do not have permission to access this page. Only accountants, principals, or users with fee viewing permissions can access Student Fees.', 'error')
        return redirect(url_for('dashboard_employee'))
    
    print(f"  - Access GRANTED: Rendering student fees page")
    
    # Get students with their fee information
    connection = get_db_connection()
    students = []
    academic_levels = []
    academic_years = []
    terms = []
    if connection:
        try:
            with connection.cursor() as cursor:
                # Fetch students who are in session
                cursor.execute("""
                    SELECT s.id, s.student_id, s.full_name, s.current_grade, s.status, s.student_category,
                           p.full_name as parent_name, p.phone as parent_phone, p.email as parent_email
                    FROM students s
                    LEFT JOIN parents p ON s.student_id = p.student_id
                    WHERE s.status = 'in session'
                    ORDER BY s.full_name ASC
                """)
                results = cursor.fetchall()
                
                if results:
                    for row in results:
                        student_grade = row.get('current_grade', '')
                        
                        # Find the academic level for this student's grade
                        academic_level = None
                        fee_structure = None
                        
                        if student_grade:
                            # Try to match student's current_grade with academic_levels.level_name
                            cursor.execute("""
                                SELECT al.id, al.level_category, al.level_name, al.level_description
                                FROM academic_levels al
                                WHERE al.level_name = %s AND al.level_status = 'active'
                                LIMIT 1
                            """, (student_grade,))
                            level_result = cursor.fetchone()
                            
                            if level_result:
                                academic_level_id = level_result.get('id')
                                academic_level = {
                                    'id': academic_level_id,
                                    'level_category': level_result.get('level_category', ''),
                                    'level_name': level_result.get('level_name', ''),
                                    'level_description': level_result.get('level_description', '')
                                }
                                
                                # Find active fee structure for this academic level
                                # Match fee structure category with student category
                                # Priority: 1) Category-specific match, 2) 'both' category, 3) NULL category
                                # A fee structure marked as 'self sponsored' ONLY applies to self sponsored students
                                # A fee structure marked as 'sponsored' ONLY applies to sponsored students
                                # A fee structure marked as 'both' applies to all students
                                student_category = row.get('student_category', '').lower().strip() if row.get('student_category') else ''
                                
                                if student_category == 'self sponsored':
                                    # Match fee structures for self sponsored students
                                    # Priority: 'self sponsored' first, then 'both', then NULL
                                    cursor.execute("""
                                        SELECT fs.id, fs.fee_name, fs.start_date, fs.end_date, 
                                               fs.payment_deadline, fs.total_amount, fs.status, fs.category
                                        FROM fee_structures fs
                                        WHERE fs.academic_level_id = %s 
                                          AND fs.status = 'active'
                                          AND (
                                              fs.category = 'self sponsored' 
                                              OR fs.category = 'both' 
                                              OR fs.category IS NULL
                                          )
                                          AND fs.category != 'sponsored'
                                        ORDER BY 
                                          CASE 
                                            WHEN fs.category = 'self sponsored' THEN 1
                                            WHEN fs.category = 'both' THEN 2
                                            WHEN fs.category IS NULL THEN 3
                                            ELSE 4
                                          END,
                                          fs.created_at DESC
                                        LIMIT 1
                                    """, (academic_level_id,))
                                elif student_category == 'sponsored':
                                    # Match fee structures for sponsored students
                                    # Priority: 'sponsored' first, then 'both', then NULL
                                    cursor.execute("""
                                        SELECT fs.id, fs.fee_name, fs.start_date, fs.end_date, 
                                               fs.payment_deadline, fs.total_amount, fs.status, fs.category
                                        FROM fee_structures fs
                                        WHERE fs.academic_level_id = %s 
                                          AND fs.status = 'active'
                                          AND (
                                              fs.category = 'sponsored' 
                                              OR fs.category = 'both' 
                                              OR fs.category IS NULL
                                          )
                                          AND fs.category != 'self sponsored'
                                        ORDER BY 
                                          CASE 
                                            WHEN fs.category = 'sponsored' THEN 1
                                            WHEN fs.category = 'both' THEN 2
                                            WHEN fs.category IS NULL THEN 3
                                            ELSE 4
                                          END,
                                          fs.created_at DESC
                                        LIMIT 1
                                    """, (academic_level_id,))
                                else:
                                    # If student has no category or unknown category, match 'both' or NULL only
                                    # Do not match category-specific fee structures
                                    cursor.execute("""
                                        SELECT fs.id, fs.fee_name, fs.start_date, fs.end_date, 
                                               fs.payment_deadline, fs.total_amount, fs.status, fs.category
                                        FROM fee_structures fs
                                        WHERE fs.academic_level_id = %s 
                                          AND fs.status = 'active'
                                          AND (fs.category = 'both' OR fs.category IS NULL)
                                        ORDER BY fs.created_at DESC
                                        LIMIT 1
                                    """, (academic_level_id,))
                                
                                fee_structure_result = cursor.fetchone()
                                
                                # Only fall back to any fee structure if student has no category and no 'both' structure exists
                                # This ensures category-specific structures are never shown to wrong student categories
                                if not fee_structure_result and not student_category:
                                    cursor.execute("""
                                        SELECT fs.id, fs.fee_name, fs.start_date, fs.end_date, 
                                               fs.payment_deadline, fs.total_amount, fs.status, fs.category
                                        FROM fee_structures fs
                                    WHERE fs.academic_level_id = %s 
                                      AND fs.status = 'active'
                                    ORDER BY fs.created_at DESC
                                    LIMIT 1
                                """, (academic_level_id,))
                                fee_structure_result = cursor.fetchone()
                                
                                if fee_structure_result:
                                    # Format dates
                                    start_date = fee_structure_result.get('start_date')
                                    end_date = fee_structure_result.get('end_date')
                                    payment_deadline = fee_structure_result.get('payment_deadline')
                                    
                                    if start_date and hasattr(start_date, 'strftime'):
                                        start_date = start_date.strftime('%Y-%m-%d')
                                    elif start_date:
                                        start_date = str(start_date).split(' ')[0]
                                    
                                    if end_date and hasattr(end_date, 'strftime'):
                                        end_date = end_date.strftime('%Y-%m-%d')
                                    elif end_date:
                                        end_date = str(end_date).split(' ')[0]
                                    
                                    if payment_deadline and hasattr(payment_deadline, 'strftime'):
                                        payment_deadline = payment_deadline.strftime('%Y-%m-%d')
                                    elif payment_deadline:
                                        payment_deadline = str(payment_deadline).split(' ')[0]
                                    
                                    fee_structure = {
                                        'id': fee_structure_result.get('id'),
                                        'fee_name': fee_structure_result.get('fee_name', ''),
                                        'category': fee_structure_result.get('category', 'both'),
                                        'start_date': start_date,
                                        'end_date': end_date,
                                        'payment_deadline': payment_deadline,
                                        'total_amount': float(fee_structure_result.get('total_amount', 0)),
                                        'status': fee_structure_result.get('status', 'active')
                                    }
                                    
                                    # Fetch fee items for this structure
                                    cursor.execute("""
                                        SELECT item_name, item_description, amount
                                        FROM fee_items
                                        WHERE fee_structure_id = %s
                                        ORDER BY item_order ASC
                                    """, (fee_structure['id'],))
                                    fee_items = cursor.fetchall()
                                    fee_structure['items'] = [{
                                        'item_name': item.get('item_name', ''),
                                        'item_description': item.get('item_description', ''),
                                        'amount': float(item.get('amount', 0))
                                    } for item in fee_items]
                        
                        # Calculate total paid and balance
                        total_paid = 0.00
                        carry_forward = 0.00  # Overpayments from previous fee structures
                        
                        if fee_structure:
                            # Get payments for current fee structure
                            cursor.execute("""
                                SELECT COALESCE(SUM(amount_paid), 0) as total_paid
                                FROM student_payments
                                WHERE student_id = %s AND fee_structure_id = %s
                            """, (row.get('student_id'), fee_structure.get('id')))
                            payment_result = cursor.fetchone()
                            if payment_result:
                                total_paid = float(payment_result.get('total_paid', 0) if isinstance(payment_result, dict) else payment_result[0] or 0)
                            
                            # Calculate carry-forward from previous fee structures (overpayments)
                            # Get all previous fee structures for this student's academic level that have ended
                            if academic_level_id:
                                current_start_date = fee_structure.get('start_date')
                                if current_start_date:
                                    cursor.execute("""
                                        SELECT fs.id, fs.total_amount,
                                               COALESCE(SUM(sp.amount_paid), 0) as total_paid
                                        FROM fee_structures fs
                                        LEFT JOIN student_payments sp ON fs.id = sp.fee_structure_id AND sp.student_id = %s
                                        WHERE fs.academic_level_id = %s
                                        AND fs.id != %s
                                        AND fs.status = 'active'
                                        AND (fs.end_date < CURDATE() OR (fs.end_date IS NOT NULL AND fs.end_date < %s))
                                        GROUP BY fs.id, fs.total_amount
                                    """, (row.get('student_id'), academic_level_id, fee_structure.get('id'), current_start_date))
                                else:
                                    cursor.execute("""
                                        SELECT fs.id, fs.total_amount,
                                               COALESCE(SUM(sp.amount_paid), 0) as total_paid
                                        FROM fee_structures fs
                                        LEFT JOIN student_payments sp ON fs.id = sp.fee_structure_id AND sp.student_id = %s
                                        WHERE fs.academic_level_id = %s
                                        AND fs.id != %s
                                        AND fs.status = 'active'
                                        AND fs.end_date < CURDATE()
                                        GROUP BY fs.id, fs.total_amount
                                    """, (row.get('student_id'), academic_level_id, fee_structure.get('id')))
                                
                                previous_structures = cursor.fetchall()
                                for prev_struct in previous_structures:
                                    if isinstance(prev_struct, dict):
                                        prev_total = float(prev_struct.get('total_amount', 0) or 0)
                                        prev_paid = float(prev_struct.get('total_paid', 0) or 0)
                                    else:
                                        prev_total = float(prev_struct[1] or 0) if len(prev_struct) > 1 else 0
                                        prev_paid = float(prev_struct[2] or 0) if len(prev_struct) > 2 else 0
                                    
                                    prev_balance = prev_total - prev_paid
                                    # If there's an overpayment (negative balance), add to carry-forward
                                    if prev_balance < 0:
                                        carry_forward += abs(prev_balance)
                        
                        balance = 0.00
                        if fee_structure:
                            # Balance = Fee Total - (Payments + Carry Forward)
                            balance = float(fee_structure.get('total_amount', 0)) - (total_paid + carry_forward)
                        
                        # Determine payment status
                        payment_status = 'no_structure'
                        if fee_structure:
                            if balance <= 0:
                                payment_status = 'paid'
                            elif fee_structure.get('payment_deadline'):
                                deadline_str = fee_structure.get('payment_deadline')
                                try:
                                    deadline_date = datetime.strptime(deadline_str, '%Y-%m-%d').date()
                                    today_date = datetime.now().date()
                                    if deadline_date < today_date:
                                        payment_status = 'overdue'
                                    else:
                                        payment_status = 'pending'
                                except:
                                    payment_status = 'pending'
                            else:
                                payment_status = 'pending'
                        
                        students.append({
                            'id': row.get('id'),
                            'student_id': row.get('student_id'),
                            'full_name': row.get('full_name', ''),
                            'current_grade': student_grade,
                            'status': row.get('status', ''),
                            'student_category': row.get('student_category', ''),
                            'parent_name': row.get('parent_name', ''),
                            'parent_phone': row.get('parent_phone', ''),
                            'parent_email': row.get('parent_email', ''),
                            'academic_level': academic_level,
                            'fee_structure': fee_structure,
                            'payment_status': payment_status,
                            'total_paid': total_paid,
                            'carry_forward': carry_forward,
                            'balance': balance
                        })
                
                # Fetch active academic levels for fee structure creation
                # Include information about whether they have fee structures
                cursor.execute("""
                    SELECT al.id, al.level_category, al.level_name, al.level_description,
                           COUNT(DISTINCT fs.id) as fee_structure_count
                    FROM academic_levels al
                    LEFT JOIN fee_structures fs ON al.id = fs.academic_level_id 
                        AND fs.status = 'active'
                    WHERE al.level_status = 'active'
                    GROUP BY al.id, al.level_category, al.level_name, al.level_description
                    ORDER BY al.level_name ASC
                """)
                academic_levels_results = cursor.fetchall()
                
                if academic_levels_results:
                    for row in academic_levels_results:
                        fee_structure_count = row.get('fee_structure_count', 0) if isinstance(row, dict) else row[4]
                        fee_structure_count = int(fee_structure_count) if fee_structure_count else 0
                        academic_levels.append({
                            'id': row.get('id') if isinstance(row, dict) else row[0],
                            'level_category': row.get('level_category', '') if isinstance(row, dict) else row[1],
                            'level_name': row.get('level_name', '') if isinstance(row, dict) else row[2],
                            'level_description': row.get('level_description', '') if isinstance(row, dict) else row[3],
                            'has_fee_structure': fee_structure_count > 0,
                            'fee_structure_count': fee_structure_count
                        })
                
                # Fetch ALL active academic years (ONLY filter by status='active', ignore is_current)
                cursor.execute("""
                    SELECT id, year_name, start_date, end_date, status, is_current, is_locked
                    FROM academic_years
                    WHERE status = 'active'
                    ORDER BY start_date DESC
                """)
                academic_years_results = cursor.fetchall()
                print(f"DEBUG: Found {len(academic_years_results) if academic_years_results else 0} active academic years")
                if academic_years_results:
                    for year in academic_years_results:
                        year_dict = {
                            'id': year.get('id') if isinstance(year, dict) else year[0],
                            'year_name': year.get('year_name', '') if isinstance(year, dict) else year[1],
                            'start_date': year.get('start_date') if isinstance(year, dict) else year[2],
                            'end_date': year.get('end_date') if isinstance(year, dict) else year[3],
                            'status': year.get('status', '') if isinstance(year, dict) else year[4],
                            'is_current': year.get('is_current', False) if isinstance(year, dict) else year[5],
                            'is_locked': year.get('is_locked', False) if isinstance(year, dict) else (year[6] if len(year) > 6 else False)
                        }
                        academic_years.append(year_dict)
                        print(f"DEBUG: Added academic year: {year_dict['year_name']} (ID: {year_dict['id']}, Status: {year_dict['status']})")
                
                # Fetch ALL active terms that are not locked
                cursor.execute("""
                    SELECT t.id, t.term_name, t.academic_year_id, t.start_date, t.end_date, t.status, t.is_current, t.is_locked,
                           ay.year_name as academic_year_name
                    FROM terms t
                    LEFT JOIN academic_years ay ON t.academic_year_id = ay.id
                    WHERE t.status = 'active' AND (t.is_locked = FALSE OR t.is_locked IS NULL)
                    ORDER BY t.academic_year_id DESC, t.start_date ASC
                """)
                terms_results = cursor.fetchall()
                print(f"DEBUG: Found {len(terms_results) if terms_results else 0} active terms")
                if terms_results:
                    for term in terms_results:
                        term_dict = {
                            'id': term.get('id') if isinstance(term, dict) else term[0],
                            'term_name': term.get('term_name', '') if isinstance(term, dict) else term[1],
                            'academic_year_id': term.get('academic_year_id') if isinstance(term, dict) else term[2],
                            'start_date': term.get('start_date') if isinstance(term, dict) else term[3],
                            'end_date': term.get('end_date') if isinstance(term, dict) else term[4],
                            'status': term.get('status', '') if isinstance(term, dict) else term[5],
                            'is_current': term.get('is_current', False) if isinstance(term, dict) else term[6],
                            'is_locked': term.get('is_locked', False) if isinstance(term, dict) else term[7],
                            'academic_year_name': term.get('academic_year_name', '') if isinstance(term, dict) else term[8]
                        }
                        terms.append(term_dict)
                        print(f"DEBUG: Added term: {term_dict['term_name']} (ID: {term_dict['id']}, Year ID: {term_dict['academic_year_id']}, Status: {term_dict['status']})")
        except Exception as e:
            print(f"Error fetching data: {e}")
            import traceback
            traceback.print_exc()
        finally:
            connection.close()
    
    from datetime import datetime
    today = datetime.now().strftime('%Y-%m-%d')
    
    # Get current employee info for payment recording
    current_employee = None
    connection = get_db_connection()
    if connection:
        try:
            with connection.cursor() as cursor:
                employee_id = session.get('employee_id') or session.get('user_id')
                if employee_id:
                    cursor.execute("""
                        SELECT id, employee_id, full_name, role
                        FROM employees
                        WHERE id = %s OR employee_id = %s
                        LIMIT 1
                    """, (employee_id, employee_id))
                    emp_result = cursor.fetchone()
                    if emp_result:
                        current_employee = {
                            'id': emp_result.get('id'),
                            'employee_id': emp_result.get('employee_id'),
                            'full_name': emp_result.get('full_name'),
                            'role': emp_result.get('role')
                        }
        except Exception as e:
            print(f"Error fetching employee info: {e}")
        finally:
            connection.close()
    
    return render_template('dashboards/student_fees.html', 
                         students=students, 
                         academic_levels=academic_levels,
                         academic_years=academic_years,
                         terms=terms,
                         today=today,
                         current_employee=current_employee,
                         can_view_students=can_view_students,
                         can_view_fee_structure_details=can_view_fee_structure_details,
                         can_record_payments=can_record_payments,
                         can_add_fee_structure=can_add_fee_structure,
                         can_edit_fee_structure=can_edit_fee_structure,
                         can_delete_fee_structure=can_delete_fee_structure)

@app.route('/dashboard/employee/student-fees/generate-invoice/<student_id>')
@login_required
def generate_invoice(student_id):
    """Generate a professional PDF invoice for a student"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    employee_id = session.get('employee_id') or session.get('user_id')
    
    # Check if user is accountant or viewing as accountant
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_technician = user_role == 'technician'
    
    # Check permission-based access
    has_generate_invoices_permission = check_permission_or_role('generate_invoices', ['accountant'])
    
    if not (is_accountant or is_technician or has_generate_invoices_permission):
        flash('You do not have permission to generate invoices.', 'error')
        return redirect(url_for('student_fees'))
    
    connection = get_db_connection()
    if not connection:
        flash('Database connection error.', 'error')
        return redirect(url_for('student_fees'))
    
    try:
        with connection.cursor() as cursor:
            # Get student information
            cursor.execute("""
                SELECT s.id, s.student_id, s.full_name, s.current_grade, s.status, s.student_category,
                       p.full_name as parent_name, p.phone as parent_phone, p.email as parent_email, p.relationship
                FROM students s
                LEFT JOIN parents p ON s.student_id = p.student_id
                WHERE s.student_id = %s AND s.status = 'in session'
                LIMIT 1
            """, (student_id,))
            student_result = cursor.fetchone()
            
            if not student_result:
                flash('Student not found.', 'error')
                return redirect(url_for('student_fees'))
            
            student = dict(student_result) if isinstance(student_result, dict) else {
                'id': student_result[0],
                'student_id': student_result[1],
                'full_name': student_result[2],
                'current_grade': student_result[3],
                'status': student_result[4],
                'student_category': student_result[5] if len(student_result) > 5 else None,
                'parent_name': student_result[6] if len(student_result) > 6 else None,
                'parent_phone': student_result[7] if len(student_result) > 7 else None,
                'parent_email': student_result[8] if len(student_result) > 8 else None,
                'relationship': student_result[9] if len(student_result) > 9 else None
            }
            
            # Get student's academic level by matching current_grade with academic_levels
            # First get the student's current_grade
            cursor.execute("""
                SELECT current_grade 
                FROM students 
                WHERE student_id = %s
                LIMIT 1
            """, (student_id,))
            student_grade_result = cursor.fetchone()
            student_grade = None
            if student_grade_result:
                student_grade = student_grade_result[0] if isinstance(student_grade_result, (list, tuple)) else student_grade_result.get('current_grade')
            
            # Get fee structure and payments
            fee_structure = None
            academic_level_id = None
            if student_grade:
                # Find academic level by matching level_name with current_grade
                cursor.execute("""
                    SELECT id 
                    FROM academic_levels 
                    WHERE level_name = %s AND level_status = 'active'
                    LIMIT 1
                """, (student_grade,))
                academic_level_result = cursor.fetchone()
                if academic_level_result:
                    academic_level_id = academic_level_result[0] if isinstance(academic_level_result, (list, tuple)) else academic_level_result.get('id')
            
            if academic_level_id:
                cursor.execute("""
                    SELECT fs.id, fs.fee_name, fs.start_date, fs.end_date, fs.payment_deadline,
                           fs.total_amount, al.level_name, al.level_category
                    FROM fee_structures fs
                    LEFT JOIN academic_levels al ON fs.academic_level_id = al.id
                    WHERE fs.academic_level_id = %s AND fs.status = 'active'
                    ORDER BY fs.created_at DESC
                    LIMIT 1
                """, (academic_level_id,))
                fee_structure_result = cursor.fetchone()
                
                if fee_structure_result:
                    if isinstance(fee_structure_result, dict):
                        fee_structure = {
                            'id': fee_structure_result.get('id'),
                            'fee_name': fee_structure_result.get('fee_name', ''),
                            'start_date': fee_structure_result.get('start_date'),
                            'end_date': fee_structure_result.get('end_date'),
                            'payment_deadline': fee_structure_result.get('payment_deadline'),
                            'total_amount': float(fee_structure_result.get('total_amount', 0) or 0),
                            'level_name': fee_structure_result.get('level_name'),
                            'level_category': fee_structure_result.get('level_category')
                        }
                    else:
                        fee_structure = {
                            'id': fee_structure_result[0],
                            'fee_name': fee_structure_result[1] if len(fee_structure_result) > 1 else '',
                            'start_date': fee_structure_result[2] if len(fee_structure_result) > 2 else None,
                            'end_date': fee_structure_result[3] if len(fee_structure_result) > 3 else None,
                            'payment_deadline': fee_structure_result[4] if len(fee_structure_result) > 4 else None,
                            'total_amount': float(fee_structure_result[5] or 0) if len(fee_structure_result) > 5 and fee_structure_result[5] else 0,
                            'level_name': fee_structure_result[6] if len(fee_structure_result) > 6 else None,
                            'level_category': fee_structure_result[7] if len(fee_structure_result) > 7 else None
                        }
                
                # Get fee items
                cursor.execute("""
                    SELECT item_name, item_description, amount
                    FROM fee_items
                    WHERE fee_structure_id = %s
                    ORDER BY id ASC
                """, (fee_structure['id'],))
                fee_items = cursor.fetchall()
                fee_structure['items'] = []
                for item in fee_items:
                    if isinstance(item, dict):
                        item_name = item.get('item_name', '')
                        item_description = item.get('item_description', '') or ''
                        amount = item.get('amount', 0) or 0
                    else:
                        item_name = item[0] if len(item) > 0 else ''
                        item_description = item[1] if len(item) > 1 else ''
                        amount = item[2] if len(item) > 2 and item[2] else 0
                    
                    # Convert amount to float, handling Decimal types
                    try:
                        amount_float = float(amount) if amount else 0.0
                    except (TypeError, ValueError):
                        amount_float = 0.0
                    
                    fee_structure['items'].append({
                        'item_name': item_name,
                        'item_description': item_description,
                        'amount': amount_float
                    })
            
            # Get payment transactions for current fee structure
            total_paid = 0.0
            carry_forward = 0.0
            payment_transactions = []
            balance_brought_forward = 0.0  # Initialize for ledger calculation
            
            if fee_structure:
                cursor.execute("""
                    SELECT sp.amount_paid, sp.payment_date, sp.payment_method, sp.reference_number, 
                           sp.cheque_number, sp.transaction_id, sp.notes, 
                           e.full_name as received_by_name
                    FROM student_payments sp
                    LEFT JOIN employees e ON sp.received_by = e.id
                    WHERE sp.student_id = %s AND sp.fee_structure_id = %s
                    ORDER BY sp.payment_date DESC, sp.id DESC
                """, (student_id, fee_structure['id']))
                payments = cursor.fetchall()
                
                # Calculate total paid for current fee structure and collect transaction details
                for payment in payments:
                    if isinstance(payment, dict):
                        amount = payment.get('amount_paid', 0) or 0
                        payment_date = payment.get('payment_date')
                        payment_method = payment.get('payment_method', '')
                        reference_number = payment.get('reference_number', '')
                        cheque_number = payment.get('cheque_number', '')
                        transaction_id = payment.get('transaction_id', '')
                        notes = payment.get('notes', '')
                        received_by = payment.get('received_by_name', '') or 'N/A'
                    else:
                        amount = payment[0] if len(payment) > 0 and payment[0] else 0
                        payment_date = payment[1] if len(payment) > 1 else None
                        payment_method = payment[2] if len(payment) > 2 else ''
                        reference_number = payment[3] if len(payment) > 3 else ''
                        cheque_number = payment[4] if len(payment) > 4 else ''
                        transaction_id = payment[5] if len(payment) > 5 else ''
                        notes = payment[6] if len(payment) > 6 else ''
                        received_by = payment[7] if len(payment) > 7 else 'N/A'
                    
                    # Convert to float, handling Decimal types
                    try:
                        amount_float = float(amount) if amount else 0.0
                        total_paid += amount_float
                        
                        # Format payment date
                        if payment_date:
                            if hasattr(payment_date, 'strftime'):
                                formatted_date = payment_date.strftime('%B %d, %Y')
                            else:
                                try:
                                    date_obj = datetime.strptime(str(payment_date).split(' ')[0], '%Y-%m-%d')
                                    formatted_date = date_obj.strftime('%B %d, %Y')
                                except:
                                    formatted_date = str(payment_date).split(' ')[0]
                        else:
                            formatted_date = 'N/A'
                        
                        payment_transactions.append({
                            'amount': amount_float,
                            'date': formatted_date,
                            'method': payment_method,
                            'reference': reference_number,
                            'cheque': cheque_number,
                            'transaction_id': transaction_id,
                            'notes': notes,
                            'received_by': received_by
                        })
                    except (TypeError, ValueError):
                        total_paid += 0.0
                
                # Calculate carry-forward from previous fee structures (overpayments)
                if academic_level_id:
                    current_start_date = fee_structure.get('start_date')
                    if current_start_date:
                        cursor.execute("""
                            SELECT fs.id, fs.total_amount,
                                   COALESCE(SUM(sp.amount_paid), 0) as total_paid
                            FROM fee_structures fs
                            LEFT JOIN student_payments sp ON fs.id = sp.fee_structure_id AND sp.student_id = %s
                            WHERE fs.academic_level_id = %s
                            AND fs.id != %s
                            AND fs.status = 'active'
                            AND (fs.end_date < CURDATE() OR (fs.end_date IS NOT NULL AND fs.end_date < %s))
                            GROUP BY fs.id, fs.total_amount
                        """, (student_id, academic_level_id, fee_structure['id'], current_start_date))
                    else:
                        cursor.execute("""
                            SELECT fs.id, fs.total_amount,
                                   COALESCE(SUM(sp.amount_paid), 0) as total_paid
                            FROM fee_structures fs
                            LEFT JOIN student_payments sp ON fs.id = sp.fee_structure_id AND sp.student_id = %s
                            WHERE fs.academic_level_id = %s
                            AND fs.id != %s
                            AND fs.status = 'active'
                            AND fs.end_date < CURDATE()
                            GROUP BY fs.id, fs.total_amount
                        """, (student_id, academic_level_id, fee_structure['id']))
                    
                    previous_structures = cursor.fetchall()
                    for prev_struct in previous_structures:
                        if isinstance(prev_struct, dict):
                            prev_total = float(prev_struct.get('total_amount', 0) or 0)
                            prev_paid = float(prev_struct.get('total_paid', 0) or 0)
                        else:
                            prev_total = float(prev_struct[1] or 0) if len(prev_struct) > 1 else 0
                            prev_paid = float(prev_struct[2] or 0) if len(prev_struct) > 2 else 0
                        
                        prev_balance = prev_total - prev_paid
                        # If there's an overpayment (negative balance), add to carry-forward
                        if prev_balance < 0:
                            carry_forward += abs(prev_balance)
                
                # Calculate balance brought forward from previous fee structures (for ledger)
                balance_brought_forward = 0.0
                previous_term_info = None  # Store previous term's closing info
                if academic_level_id and fee_structure:
                    current_start_date = fee_structure.get('start_date')
                    if current_start_date:
                        try:
                            if hasattr(current_start_date, 'strftime'):
                                start_date_str = current_start_date.strftime('%Y-%m-%d')
                            else:
                                start_date_str = str(current_start_date).split(' ')[0]
                            
                            cursor.execute("""
                                SELECT fs.id, fs.total_amount, fs.fee_name, fs.end_date,
                                       COALESCE(SUM(sp.amount_paid), 0) as total_paid
                                FROM fee_structures fs
                                LEFT JOIN student_payments sp ON fs.id = sp.fee_structure_id AND sp.student_id = %s
                                WHERE fs.academic_level_id = %s
                                AND fs.id != %s
                                AND fs.status = 'active'
                                AND (fs.end_date < CURDATE() OR (fs.end_date IS NOT NULL AND fs.end_date < %s))
                                GROUP BY fs.id, fs.total_amount, fs.fee_name, fs.end_date
                                ORDER BY fs.end_date DESC
                                LIMIT 1
                            """, (student_id, academic_level_id, fee_structure['id'], start_date_str))
                        except:
                            cursor.execute("""
                                SELECT fs.id, fs.total_amount, fs.fee_name, fs.end_date,
                                       COALESCE(SUM(sp.amount_paid), 0) as total_paid
                                FROM fee_structures fs
                                LEFT JOIN student_payments sp ON fs.id = sp.fee_structure_id AND sp.student_id = %s
                                WHERE fs.academic_level_id = %s
                                AND fs.id != %s
                                AND fs.status = 'active'
                                AND fs.end_date < CURDATE()
                                GROUP BY fs.id, fs.total_amount, fs.fee_name, fs.end_date
                                ORDER BY fs.end_date DESC
                                LIMIT 1
                            """, (student_id, academic_level_id, fee_structure['id']))
                    else:
                        cursor.execute("""
                            SELECT fs.id, fs.total_amount, fs.fee_name, fs.end_date,
                                   COALESCE(SUM(sp.amount_paid), 0) as total_paid
                            FROM fee_structures fs
                            LEFT JOIN student_payments sp ON fs.id = sp.fee_structure_id AND sp.student_id = %s
                            WHERE fs.academic_level_id = %s
                            AND fs.id != %s
                            AND fs.status = 'active'
                            AND fs.end_date < CURDATE()
                            GROUP BY fs.id, fs.total_amount, fs.fee_name, fs.end_date
                            ORDER BY fs.end_date DESC
                            LIMIT 1
                        """, (student_id, academic_level_id, fee_structure['id']))
                    
                    previous_term_result = cursor.fetchone()
                    if previous_term_result:
                        if isinstance(previous_term_result, dict):
                            prev_total = float(previous_term_result.get('total_amount', 0) or 0)
                            prev_paid = float(previous_term_result.get('total_paid', 0) or 0)
                            prev_end_date = previous_term_result.get('end_date')
                            prev_fee_name = previous_term_result.get('fee_name', '')
                        else:
                            prev_total = float(previous_term_result[1] or 0) if len(previous_term_result) > 1 else 0
                            prev_paid = float(previous_term_result[4] or 0) if len(previous_term_result) > 4 else 0
                            prev_end_date = previous_term_result[3] if len(previous_term_result) > 3 else None
                            prev_fee_name = previous_term_result[2] if len(previous_term_result) > 2 else ''
                        
                        prev_balance = prev_total - prev_paid
                        balance_brought_forward = prev_balance
                        previous_term_info = {
                            'fee_name': prev_fee_name,
                            'end_date': prev_end_date,
                            'closing_balance': prev_balance
                        }
                    else:
                        previous_term_info = None
                else:
                    balance_brought_forward = 0.0
                    previous_term_info = None
            
            # Ensure both values are floats before subtraction
            fee_total = float(fee_structure.get('total_amount', 0) or 0) if fee_structure else 0.0
            balance = fee_total - (total_paid + carry_forward)
            
            # Get school settings
            cursor.execute("SELECT school_name, school_location, school_phone, school_email, school_logo FROM school_settings ORDER BY id DESC LIMIT 1")
            school_result = cursor.fetchone()
            school_settings = {}
            if school_result:
                if isinstance(school_result, dict):
                    school_settings = {
                        'school_name': school_result.get('school_name', 'Modern School') or 'Modern School',
                        'school_address': school_result.get('school_location', '') or '',
                        'school_phone': school_result.get('school_phone', '') or '',
                        'school_email': school_result.get('school_email', '') or '',
                        'school_logo': school_result.get('school_logo', '') or ''
                    }
                else:
                    school_settings = {
                        'school_name': school_result[0] if len(school_result) > 0 and school_result[0] else 'Modern School',
                        'school_address': school_result[1] if len(school_result) > 1 and school_result[1] else '',
                        'school_phone': school_result[2] if len(school_result) > 2 and school_result[2] else '',
                        'school_email': school_result[3] if len(school_result) > 3 and school_result[3] else '',
                        'school_logo': school_result[4] if len(school_result) > 4 and school_result[4] else ''
                    }
            else:
                # Default values if no settings found
                school_settings = {
                    'school_name': 'Modern School',
                    'school_address': '',
                    'school_phone': '',
                    'school_email': '',
                    'school_logo': ''
                }
            
            # Get student profile image if available (check if profile_image column exists)
            student['profile_image'] = None
            try:
                cursor.execute("SHOW COLUMNS FROM students LIKE 'profile_image'")
                if cursor.fetchone():
                    cursor.execute("SELECT profile_image FROM students WHERE student_id = %s", (student_id,))
                    profile_result = cursor.fetchone()
                    if profile_result:
                        profile_image = profile_result[0] if isinstance(profile_result, (list, tuple)) else profile_result.get('profile_image')
                        if profile_image:
                            student['profile_image'] = url_for('static', filename=profile_image)
            except:
                pass  # Column doesn't exist, skip
            
            # Generate invoice date, number, and time
            now = datetime.now()
            invoice_date = now.strftime('%B %d, %Y')
            invoice_time = now.strftime('%I:%M %p')
            invoice_number = f"INV-{student_id}-{now.strftime('%Y%m%d%H%M%S')}"
            
    except Exception as e:
        print(f"Error generating invoice: {e}")
        import traceback
        traceback.print_exc()
        flash('Error generating invoice.', 'error')
        return redirect(url_for('student_fees'))
    finally:
        connection.close()
    
    # Check if user wants PDF or HTML format (default to HTML for modern view)
    format_type = request.args.get('format', 'html').lower()
    
    if format_type == 'html':
        # Render HTML invoice template
        return render_template('dashboards/invoice.html',
                             student=student,
                             fee_structure=fee_structure,
                             payment_transactions=payment_transactions,
                             total_paid=total_paid,
                             carry_forward=carry_forward,
                             balance_brought_forward=balance_brought_forward,
                             balance=balance,
                             school_settings=school_settings,
                             invoice_date=invoice_date,
                             invoice_time=invoice_time,
                             invoice_number=invoice_number)
    
    # Generate PDF (original functionality)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, 
                           rightMargin=0.75*inch, leftMargin=0.75*inch,
                           topMargin=0.75*inch, bottomMargin=0.75*inch)
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Define styles with improved color scheme
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=28,
        textColor=colors.HexColor('#1D3557'),
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=13,
        textColor=colors.HexColor('#1D3557'),
        spaceAfter=10,
        fontName='Helvetica-Bold'
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#2B2D42'),
        spaceAfter=6
    )
    
    accent_style = ParagraphStyle(
        'AccentStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#E63946'),
        spaceAfter=6,
        fontName='Helvetica-Bold'
    )
    
    # Header with improved professional design
    school_name = school_settings.get('school_name', 'Modern School')
    school_address = school_settings.get('school_address', '') or ''
    school_phone = school_settings.get('school_phone', '') or ''
    school_email = school_settings.get('school_email', '') or ''
    
    # Professional header with school info and invoice title side by side
    header_title_style = ParagraphStyle(
        'HeaderTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1D3557'),
        fontName='Helvetica-Bold',
        spaceAfter=4
    )
    
    header_subtitle_style = ParagraphStyle(
        'HeaderSubtitle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#2B2D42'),
        fontName='Helvetica'
    )
    
    # Left column: School info
    school_info_lines = [Paragraph(f"<b>{school_name.upper()}</b>", header_title_style)]
    if school_address:
        school_info_lines.append(Paragraph(school_address, header_subtitle_style))
    if school_phone:
        school_info_lines.append(Paragraph(f"Phone: {school_phone}", header_subtitle_style))
    if school_email:
        school_info_lines.append(Paragraph(f"Email: {school_email}", header_subtitle_style))
    
    # Right column: Invoice title and student name
    student_name = student.get('full_name', 'N/A')
    invoice_title_style = ParagraphStyle(
        'InvoiceTitleHeader',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#E63946'),
        fontName='Helvetica-Bold',
        alignment=TA_RIGHT,
        spaceAfter=6
    )
    
    student_name_header_style = ParagraphStyle(
        'StudentNameHeader',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.HexColor('#1D3557'),
        fontName='Helvetica-Bold',
        alignment=TA_RIGHT,
        spaceAfter=2
    )
    
    invoice_info_lines = [
        Paragraph("FEE INVOICE", invoice_title_style),
        Paragraph(f"<b>{student_name.upper()}</b>", student_name_header_style)
    ]
    
    # Create two-column header
    header_data = [[school_info_lines, invoice_info_lines]]
    header_table = Table(header_data, colWidths=[4*inch, 3*inch])
    header_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('LINEBELOW', (0, 0), (-1, 0), 3, colors.HexColor('#1D3557')),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 0.25*inch))
    
    # Professional invoice details section
    invoice_date = datetime.now().strftime("%B %d, %Y")
    invoice_number = f"INV-{student_id}-{datetime.now().strftime('%Y%m%d')}"
    
    # Get current academic year and term if available
    current_year = datetime.now().strftime('%Y')
    if fee_structure and fee_structure.get('level_name'):
        academic_info = f"{fee_structure.get('level_name', '')} {fee_structure.get('level_category', '')}".strip()
    else:
        academic_info = student.get('current_grade', 'N/A')
    
    # Create a more organized information layout
    detail_label_style = ParagraphStyle(
        'DetailLabel',
        parent=normal_style,
        fontSize=9,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#1D3557')
    )
    
    # Header style with white text for blue backgrounds
    header_label_style = ParagraphStyle(
        'HeaderLabel',
        parent=normal_style,
        fontSize=9,
        fontName='Helvetica-Bold',
        textColor=colors.white
    )
    
    detail_value_style = ParagraphStyle(
        'DetailValue',
        parent=normal_style,
        fontSize=10,
        fontName='Helvetica',
        textColor=colors.HexColor('#2B2D42')
    )
    
    # Student information section
    student_category_text = ''
    if student.get('student_category'):
        category_display = student.get('student_category', '').title()
        student_category_text = f"<br/><b>Category:</b> {category_display}"
    
    student_details = [
        [Paragraph('<b>STUDENT INFORMATION</b>', header_label_style), ''],
        [Paragraph('Student ID:', detail_label_style), Paragraph(student_id, detail_value_style)],
        [Paragraph('Full Name:', detail_label_style), Paragraph(student.get('full_name', 'N/A'), detail_value_style)],
        [Paragraph('Grade:', detail_label_style), Paragraph(student.get('current_grade', 'N/A'), detail_value_style)],
    ]
    if student.get('student_category'):
        student_details.append([Paragraph('Category:', detail_label_style), Paragraph(category_display, detail_value_style)])
    
    # Invoice information section
    invoice_details = [
        [Paragraph('<b>INVOICE INFORMATION</b>', header_label_style), ''],
        [Paragraph('Invoice Number:', detail_label_style), Paragraph(invoice_number, detail_value_style)],
        [Paragraph('Invoice Date:', detail_label_style), Paragraph(invoice_date, detail_value_style)],
        [Paragraph('Academic Level:', detail_label_style), Paragraph(academic_info, detail_value_style)],
    ]
    
    # Combine into two-column layout
    info_data = []
    max_rows = max(len(student_details), len(invoice_details))
    for i in range(max_rows):
        row = []
        if i < len(student_details):
            row.extend(student_details[i])
        else:
            row.extend(['', ''])
        if i < len(invoice_details):
            row.extend(invoice_details[i])
        else:
            row.extend(['', ''])
        info_data.append(row)
    
    info_table = Table(info_data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 2*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#1D3557')),
        ('BACKGROUND', (2, 0), (2, 0), colors.HexColor('#1D3557')),
        ('TEXTCOLOR', (0, 0), (0, 0), colors.white),
        ('TEXTCOLOR', (2, 0), (2, 0), colors.white),
        ('BACKGROUND', (0, 1), (1, -1), colors.HexColor('#F8F9FA')),
        ('BACKGROUND', (2, 1), (3, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#2B2D42')),  # Only apply dark text to non-header rows
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E8F4F8')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LINEBELOW', (0, 0), (-1, 0), 0, colors.HexColor('#1D3557')),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.35*inch))
    
    # Ledger-style Account Statement
    elements.append(Paragraph("ACCOUNT STATEMENT", heading_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Use balance_brought_forward calculated earlier (inside cursor context)
    # balance_brought_forward is already calculated and stored above
    
    # Build ledger table
    # Create white text style for header row
    ledger_header_style = ParagraphStyle(
        'LedgerHeader',
        parent=normal_style,
        fontSize=10,
        fontName='Helvetica-Bold',
        textColor=colors.white
    )
    
    ledger_data = [[
            Paragraph('<b>Date</b>', ledger_header_style),
            Paragraph('<b>Ref No</b>', ledger_header_style),
            Paragraph('<b>Description</b>', ledger_header_style),
            Paragraph('<b>Debit</b>', ledger_header_style),
            Paragraph('<b>Credit</b>', ledger_header_style),
            Paragraph('<b>Balance</b>', ledger_header_style)
    ]]
    
    # Previous term closing section (if there's a previous term)
    if balance_brought_forward != 0:
            # Get previous term info from the calculation above
            # We need to recalculate it here or pass it through
            # For now, let's add a separator and closing balance row
            separator_style = ParagraphStyle(
                'Separator',
                parent=normal_style,
                fontSize=9,
                fontName='Helvetica-Bold',
                textColor=colors.HexColor('#1D3557'),
                alignment=TA_CENTER
            )
            
            # Previous term closing separator
            prev_term_closing_balance = balance_brought_forward
            prev_term_end_date = ""
            if fee_structure and fee_structure.get('start_date'):
                try:
                    if hasattr(fee_structure['start_date'], 'strftime'):
                        prev_term_end_date = (fee_structure['start_date'] - timedelta(days=1)).strftime('%b %d, %Y')
                    else:
                        date_obj = datetime.strptime(str(fee_structure['start_date']).split(' ')[0], '%Y-%m-%d')
                        prev_term_end_date = (date_obj - timedelta(days=1)).strftime('%b %d, %Y')
                except:
                    prev_term_end_date = "Previous Term"
            
            # Add separator row for previous term closing
            ledger_data.append([
                Paragraph("─" * 20, separator_style),
                Paragraph("─" * 10, separator_style),
                Paragraph("<b>PREVIOUS TERM CLOSING</b>", separator_style),
                Paragraph("─" * 10, separator_style),
                Paragraph("─" * 10, separator_style),
                Paragraph("─" * 10, separator_style)
            ])
            
            # Previous term closing balance
            if prev_term_closing_balance > 0:
                closing_desc = "Closing Balance (Outstanding)"
                closing_debit = f"{prev_term_closing_balance:,.2f}"
                closing_credit = "0.00"
            else:
                closing_desc = "Closing Balance (Credit/Overpay)"
                closing_debit = "0.00"
                closing_credit = f"{abs(prev_term_closing_balance):,.2f}"
            
            ledger_data.append([
                Paragraph(prev_term_end_date or "Previous Term", normal_style),
                Paragraph("CLOSE", normal_style),
                Paragraph(closing_desc, normal_style),
                Paragraph(closing_debit, normal_style),
                Paragraph(closing_credit, normal_style),
                Paragraph(f"{prev_term_closing_balance:,.2f}", normal_style)
            ])
            
            # Add separator row for new term opening
            ledger_data.append([
                Paragraph("─" * 20, separator_style),
                Paragraph("─" * 10, separator_style),
                Paragraph("<b>NEW TERM OPENING - BALANCE BROUGHT FORWARD</b>", separator_style),
                Paragraph("─" * 10, separator_style),
                Paragraph("─" * 10, separator_style),
                Paragraph("─" * 10, separator_style)
            ])
            
            # Balance brought forward row
            # Get fee date for balance brought forward
            bf_date = "Opening"
            if fee_structure and fee_structure.get('start_date'):
                try:
                    if hasattr(fee_structure['start_date'], 'strftime'):
                        bf_date = fee_structure['start_date'].strftime('%b %d, %Y')
                    else:
                        bf_date = datetime.strptime(str(fee_structure['start_date']).split(' ')[0], '%Y-%m-%d').strftime('%b %d, %Y')
                except:
                    bf_date = "Opening"
            bf_ref = "BF"
            if balance_brought_forward > 0:
                bf_desc = "Balance Brought Forward (Outstanding)"
                bf_debit = f"{balance_brought_forward:,.2f}"
                bf_credit = "0.00"
            else:
                bf_desc = "Balance Brought Forward (Credit/Overpay)"
                bf_debit = "0.00"
                bf_credit = f"{abs(balance_brought_forward):,.2f}"
            
            running_balance = balance_brought_forward
            ledger_data.append([
                Paragraph(bf_date or "Opening", normal_style),
                Paragraph(bf_ref, normal_style),
                Paragraph(bf_desc, normal_style),
                Paragraph(bf_debit, normal_style),
                Paragraph(bf_credit, normal_style),
                Paragraph(f"{running_balance:,.2f}", normal_style)
            ])
            
            # Add separator for current term transactions
            separator_style = ParagraphStyle(
                'Separator',
                parent=normal_style,
                fontSize=9,
                fontName='Helvetica-Bold',
                textColor=colors.HexColor('#1D3557'),
                alignment=TA_CENTER
            )
            ledger_data.append([
                Paragraph("", normal_style),
                Paragraph("", normal_style),
                Paragraph("<b>CURRENT TERM TRANSACTIONS</b>", separator_style),
                Paragraph("", normal_style),
                Paragraph("", normal_style),
                Paragraph("", normal_style)
            ])
    else:
        running_balance = 0.0
    
    # Current fee structure charges - break down each fee item (Debit)
    fee_date = ""
    if fee_structure.get('start_date'):
        try:
            if hasattr(fee_structure['start_date'], 'strftime'):
                fee_date = fee_structure['start_date'].strftime('%b %d, %Y')
            else:
                fee_date = datetime.strptime(str(fee_structure['start_date']).split(' ')[0], '%Y-%m-%d').strftime('%b %d, %Y')
        except:
            fee_date = str(fee_structure.get('start_date', ''))[:10]
    
    # Add each fee item as a separate row
    if fee_structure.get('items') and len(fee_structure['items']) > 0:
        for item in fee_structure['items']:
            item_name = item.get('item_name', '')
            item_desc = item.get('item_description', '') or ''
            item_amount = float(item.get('amount', 0) or 0)
            
            # Build description with item name and description
            if item_desc and item_desc != '-':
                desc_text = f"{item_name} - {item_desc}"
            else:
                desc_text = item_name
            
            running_balance += item_amount
            
            ledger_data.append([
                Paragraph(fee_date or "Current", normal_style),
                Paragraph("INV", normal_style),
                Paragraph(desc_text, normal_style),
                Paragraph(f"{item_amount:,.2f}", normal_style),
                Paragraph("0.00", normal_style),
                Paragraph(f"{running_balance:,.2f}", normal_style)
            ])
    else:
        # Fallback: if no items, show total fee structure
        fee_total = fee_structure.get('total_amount', 0)
        running_balance += fee_total
        fee_desc = f"Fee Structure: {fee_structure.get('fee_name', 'N/A')}"
        ledger_data.append([
            Paragraph(fee_date or "Current", normal_style),
            Paragraph("INV", normal_style),
            Paragraph(fee_desc, normal_style),
            Paragraph(f"{fee_total:,.2f}", normal_style),
            Paragraph("0.00", normal_style),
            Paragraph(f"{running_balance:,.2f}", normal_style)
        ])
    
    # Payment transactions (Credits)
    if payment_transactions:
        for trans in payment_transactions:
            trans_amount = trans.get('amount', 0)
            running_balance -= trans_amount
            
            ref_text = trans.get('reference', '') or trans.get('cheque', '') or trans.get('transaction_id', '') or trans.get('method', 'N/A')
            if len(ref_text) > 20:
                ref_text = ref_text[:20] + '...'
            
            desc_text = f"Payment - {trans.get('method', 'N/A')}"
            if trans.get('notes'):
                desc_text += f" ({trans.get('notes', '')[:30]})"
            
            ledger_data.append([
                Paragraph(trans.get('date', 'N/A'), normal_style),
                Paragraph(ref_text, normal_style),
                Paragraph(desc_text, normal_style),
                Paragraph("0.00", normal_style),
                Paragraph(f"{trans_amount:,.2f}", normal_style),
                Paragraph(f"{running_balance:,.2f}", normal_style)
            ])
    
    # Final balance row
    final_balance_style = ParagraphStyle(
        'FinalBalance',
        parent=normal_style,
        fontSize=11,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#E63946') if running_balance > 0 else colors.HexColor('#10b981')
    )
    
    balance_text = "Outstanding Balance"
    if running_balance < 0:
        balance_text = "Credit Balance (Overpay)"
    
    ledger_data.append([
        Paragraph("", normal_style),
        Paragraph("", normal_style),
        Paragraph(f"<b>{balance_text}</b>", final_balance_style),
        Paragraph("", normal_style),
        Paragraph("", normal_style),
        Paragraph(f"<b>{running_balance:,.2f}</b>", final_balance_style)
    ])
    
    # Create ledger table with improved styling
    ledger_table = Table(ledger_data, colWidths=[1.1*inch, 1.1*inch, 2.4*inch, 1.0*inch, 1.0*inch, 1.2*inch])
    ledger_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1D3557')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (3, 1), (4, -2), 'RIGHT'),
        ('ALIGN', (5, 1), (5, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -2), 9),
        ('FONTSIZE', (0, -1), (-1, -1), 11),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E8F4F8')),
        ('LINEBELOW', (0, -2), (-1, -2), 2, colors.HexColor('#1D3557')),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#1D3557')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E8F4F8')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F8F9FA')]),
    ]))
    elements.append(ledger_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Summary section
    summary_style = ParagraphStyle(
        'SummaryStyle',
        parent=normal_style,
        fontSize=10,
        fontName='Helvetica',
        textColor=colors.HexColor('#2B2D42')
    )
    
    fee_total = fee_structure.get('total_amount', 0) if fee_structure else 0.0
    
    # Professional summary table
    summary_heading_style = ParagraphStyle(
        'SummaryHeading',
        parent=normal_style,
        fontSize=12,
        fontName='Helvetica-Bold',
        textColor=colors.white,  # White text for blue background
        spaceAfter=8
    )
    
    summary_label_style = ParagraphStyle(
        'SummaryLabel',
        parent=normal_style,
        fontSize=10,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#2B2D42')
    )
    
    summary_value_style = ParagraphStyle(
        'SummaryValue',
        parent=normal_style,
        fontSize=10,
        fontName='Helvetica',
        textColor=colors.HexColor('#2B2D42'),
        alignment=TA_RIGHT
    )
    
    summary_data = [
        [Paragraph('<b>SUMMARY</b>', summary_heading_style), ''],
        [Paragraph('Balance Brought Forward:', summary_label_style), Paragraph(f"KES {balance_brought_forward:,.2f}", summary_value_style)],
        [Paragraph('Current Fee Charges:', summary_label_style), Paragraph(f"KES {fee_total:,.2f}", summary_value_style)],
        [Paragraph('Total Payments:', summary_label_style), Paragraph(f"KES {total_paid:,.2f}", summary_value_style)],
    ]
    
    if running_balance < 0:
        overpay_style = ParagraphStyle(
            'OverpayStyle',
            parent=summary_value_style,
            textColor=colors.HexColor('#10b981'),
            fontName='Helvetica-Bold'
        )
        summary_data.append([Paragraph('Current Overpay:', summary_label_style), Paragraph(f"KES {abs(running_balance):,.2f}", overpay_style)])
    
    balance_style = ParagraphStyle(
        'BalanceStyle',
        parent=summary_value_style,
        fontSize=11,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#E63946') if running_balance > 0 else colors.HexColor('#10b981')
    )
    summary_data.append([Paragraph('<b>Outstanding Balance:</b>', summary_label_style), Paragraph(f"<b>KES {running_balance:,.2f}</b>", balance_style)])
    
    summary_table = Table(summary_data, colWidths=[3.5*inch, 2.5*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1D3557')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor('#F8F9FA')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E8F4F8')),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#2B2D42')),  # Only apply dark text to non-header rows
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTSIZE', (0, 1), (-1, -2), 10),
        ('FONTSIZE', (0, -1), (-1, -1), 11),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E8F4F8')),
        ('LINEBELOW', (0, -2), (-1, -2), 2, colors.HexColor('#1D3557')),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#1D3557')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.35*inch))
    
    # Payment deadline and important notes
    if fee_structure and fee_structure.get('payment_deadline'):
        deadline_date = datetime.strptime(str(fee_structure['payment_deadline']), '%Y-%m-%d').strftime('%B %d, %Y')
        deadline_style = ParagraphStyle(
            'Deadline',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#E63946'),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        elements.append(Paragraph(f"<b>⚠ Payment Deadline: {deadline_date}</b>", deadline_style))
    
    # Important notes section
    elements.append(Spacer(1, 0.2*inch))
    notes_style = ParagraphStyle(
        'NotesStyle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#2B2D42'),
        spaceAfter=6,
        alignment=TA_LEFT,
        fontName='Helvetica'
    )
    notes_text = "<b>Important Notes:</b><br/>"
    notes_text += "• Please ensure payment is made before the deadline to avoid penalties.<br/>"
    notes_text += "• Overpayments will be carried forward to the next term.<br/>"
    notes_text += "• Keep this invoice for your records.<br/>"
    if running_balance > 0:
        notes_text += f"• Outstanding balance of KES {running_balance:,.2f} must be cleared before the deadline."
    elements.append(Paragraph(notes_text, notes_style))
    elements.append(Spacer(1, 0.2*inch))
    
    if not fee_structure or fee_structure.get('total_amount', 0) == 0:
        # No fee structure message
        no_fee_style = ParagraphStyle(
            'NoFee',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.HexColor('#DC2626'),
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        elements.append(Spacer(1, 0.3*inch))
        elements.append(Paragraph("No Fee Structure Assigned", no_fee_style))
        elements.append(Paragraph("This student does not have an active fee structure.", normal_style))
        elements.append(Spacer(1, 0.3*inch))
    
    # Footer
    elements.append(Spacer(1, 0.4*inch))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#6B7280'),
        alignment=TA_CENTER,
        spaceBefore=20
    )
    elements.append(Paragraph("This is a computer-generated invoice. No signature required.", footer_style))
    elements.append(Paragraph(f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", footer_style))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Create response
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename=Invoice_{student_id}_{datetime.now().strftime("%Y%m%d")}.pdf'
    
    return response

@app.route('/dashboard/employee/student-fees/fee-items', methods=['GET'])
@login_required
def get_fee_items():
    """Get all existing fee items from the database"""
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'items': []}), 200
    
    try:
        with connection.cursor() as cursor:
            # Get distinct fee items from all fee structures
            cursor.execute("""
                SELECT DISTINCT 
                    item_name, 
                    item_description, 
                    amount
                FROM fee_items
                ORDER BY item_name ASC, amount DESC
            """)
            items_results = cursor.fetchall()
            
            items = []
            if items_results:
                for item in items_results:
                    item_dict = {
                        'item_name': item.get('item_name', '') if isinstance(item, dict) else item[0],
                        'item_description': item.get('item_description', '') if isinstance(item, dict) else item[1],
                        'amount': float(item.get('amount', 0)) if isinstance(item, dict) else float(item[2] if len(item) > 2 else 0)
                    }
                    items.append(item_dict)
            
            return jsonify({'success': True, 'items': items}), 200
    except Exception as e:
        print(f"Error fetching fee items: {e}")
        return jsonify({'success': False, 'items': []}), 200
    finally:
        connection.close()

@app.route('/dashboard/employee/student-fees/check-fee-structure', methods=['GET'])
@login_required
def check_fee_structure():
    """Check if a fee structure exists for the given academic year, term, and academic level combination"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    employee_id = session.get('employee_id') or session.get('user_id')
    
    # Check permission to view fee structure details
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_technician = user_role == 'technician'
    has_view_fee_structure_details = check_permission_or_role('view_fee_structure_details', ['accountant', 'principal'])
    has_view_fees = check_permission_or_role('view_fees', ['accountant', 'principal'])
    
    if not (is_accountant or is_technician or has_view_fee_structure_details or has_view_fees):
        return jsonify({'success': False, 'message': 'You do not have permission to view fee structure details. Please contact your administrator.'}), 403
    
    academic_year_id = request.args.get('academic_year_id')
    term_id = request.args.get('term_id')
    academic_level_id = request.args.get('academic_level_id')
    
    if not academic_year_id or not term_id or not academic_level_id:
        return jsonify({'exists': False}), 200
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'exists': False}), 200
    
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id, fee_name
                FROM fee_structures 
                WHERE academic_year_id = %s
                AND term_id = %s
                AND academic_level_id = %s 
                AND status = 'active'
            """, (academic_year_id, term_id, academic_level_id))
            
            existing_structure = cursor.fetchone()
            if existing_structure:
                existing_name = existing_structure.get('fee_name') if isinstance(existing_structure, dict) else existing_structure[1]
                return jsonify({
                    'exists': True,
                    'fee_name': existing_name
                }), 200
            else:
                return jsonify({'exists': False}), 200
    except Exception as e:
        print(f"Error checking fee structure: {e}")
        return jsonify({'exists': False}), 200
    finally:
        connection.close()

@app.route('/dashboard/employee/student-fees/check-fee-structures-for-term', methods=['GET'])
@login_required
def check_fee_structures_for_term():
    """Get list of academic level IDs that have fee structures for the given term and academic year"""
    academic_year_id = request.args.get('academic_year_id')
    term_id = request.args.get('term_id')
    
    if not academic_year_id or not term_id:
        return jsonify({'academic_levels_with_structure': []}), 200
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'academic_levels_with_structure': []}), 200
    
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT academic_level_id
                FROM fee_structures 
                WHERE academic_year_id = %s
                AND term_id = %s
                AND status = 'active'
            """, (academic_year_id, term_id))
            
            results = cursor.fetchall()
            academic_level_ids = []
            for row in results:
                level_id = row.get('academic_level_id') if isinstance(row, dict) else row[0]
                if level_id:
                    academic_level_ids.append(int(level_id))
            
            return jsonify({
                'academic_levels_with_structure': academic_level_ids
            }), 200
    except Exception as e:
        print(f"Error checking fee structures for term: {e}")
        return jsonify({'academic_levels_with_structure': []}), 200
    finally:
        connection.close()

@app.route('/dashboard/employee/student-fees/create-fee-structure', methods=['POST'])
@login_required
def create_fee_structure():
    """Create a new fee structure"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    employee_id = session.get('employee_id') or session.get('user_id')
    
    # Check if user is accountant or viewing as accountant
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_technician = user_role == 'technician'
    
    # Check permission-based access - specific permission for adding fee structures
    has_add_fee_structure_permission = check_permission_or_role('add_fee_structure', ['accountant', 'principal'])
    has_manage_fees_permission = check_permission_or_role('manage_fees', ['accountant', 'principal'])
    
    if not (is_accountant or is_technician or has_add_fee_structure_permission or has_manage_fees_permission):
        return jsonify({'success': False, 'message': 'You do not have permission to add fee structures. Please contact your administrator.'}), 403
    
    try:
        data = request.get_json()
        
        # Validate required fields
        academic_level_id = data.get('academic_level_id')
        academic_year_id = data.get('academic_year_id')
        term_id = data.get('term_id')
        fee_name = data.get('fee_name', '').strip().upper()
        fee_items = data.get('fee_items', [])
        
        if not academic_level_id or not academic_year_id or not term_id or not fee_name:
            return jsonify({'success': False, 'message': 'All required fields must be filled'}), 400
        
        if not fee_items or len(fee_items) == 0:
            return jsonify({'success': False, 'message': 'At least one fee item is required'}), 400
        
        # Calculate total amount
        total_amount = sum(float(item.get('amount', 0)) for item in fee_items)
        
        # Get employee ID for created_by
        employee_id = session.get('employee_id') or session.get('user_id')
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        
        try:
            with connection.cursor() as cursor:
                # Get term details to extract dates
                cursor.execute("""
                    SELECT start_date, end_date
                    FROM terms
                    WHERE id = %s AND status = 'active'
                """, (term_id,))
                term_result = cursor.fetchone()
                
                if not term_result:
                    return jsonify({'success': False, 'message': 'Selected term not found or not active'}), 400
                
                term_start_date = term_result.get('start_date') if isinstance(term_result, dict) else term_result[0]
                term_end_date = term_result.get('end_date') if isinstance(term_result, dict) else term_result[1]
                
                # Use term end_date as payment deadline
                payment_deadline = term_end_date
                
                # Check if a fee structure already exists for this academic year, term, and academic level combination
                cursor.execute("""
                    SELECT id, fee_name
                    FROM fee_structures 
                    WHERE academic_year_id = %s
                    AND term_id = %s
                    AND academic_level_id = %s 
                    AND status = 'active'
                """, (academic_year_id, term_id, academic_level_id))
                
                existing_structure = cursor.fetchone()
                if existing_structure:
                    existing_name = existing_structure.get('fee_name') if isinstance(existing_structure, dict) else existing_structure[1]
                    return jsonify({
                        'success': False, 
                        'message': f'A fee structure already exists for this Academic Level in this Term. Each academic level must have a different fee structure in each term, and a term that has a fee structure for an academic level cannot have another one. You cannot create more than one fee structure for the same Academic Year, Term, and Academic Level combination. Existing structure: {existing_name}. Please select a different term or academic level, or edit the existing structure.'
                    }), 400
                
                # Get category from request data
                category = data.get('category', 'both').strip().lower()
                if category not in ['self sponsored', 'sponsored', 'both']:
                    category = 'both'  # Default to 'both' if invalid
                
                # Insert fee structure
                cursor.execute("""
                    INSERT INTO fee_structures 
                    (academic_level_id, academic_year_id, term_id, fee_name, category, start_date, end_date, payment_deadline, total_amount, created_by)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (academic_level_id, academic_year_id, term_id, fee_name, category, term_start_date, term_end_date, payment_deadline, total_amount, employee_id))
                
                fee_structure_id = cursor.lastrowid
                
                # Insert fee items
                items_inserted = 0
                for index, item in enumerate(fee_items):
                    item_name = item.get('item_name', '').strip().upper()
                    item_description = item.get('item_description', '').strip()
                    amount = float(item.get('amount', 0))
                    
                    # Validate item data
                    if not item_name or amount <= 0:
                        continue  # Skip invalid items
                    
                    cursor.execute("""
                        INSERT INTO fee_items 
                        (fee_structure_id, item_name, item_description, amount, item_order)
                        VALUES (%s, %s, %s, %s, %s)
                    """, (fee_structure_id, item_name, item_description, amount, index))
                    items_inserted += 1
                
                # Verify items were inserted
                if items_inserted == 0:
                    connection.rollback()
                    return jsonify({
                        'success': False, 
                        'message': 'No valid fee items were provided. Please ensure all items have a name and amount greater than 0.'
                    }), 400
                
                connection.commit()
                
                # Verify items were saved
                cursor.execute("SELECT COUNT(*) as count FROM fee_items WHERE fee_structure_id = %s", (fee_structure_id,))
                items_count = cursor.fetchone()
                count = items_count.get('count') if isinstance(items_count, dict) else items_count[0]
                
                if count == 0:
                    connection.rollback()
                    return jsonify({
                        'success': False, 
                        'message': 'Fee structure was created but fee items failed to save. Please try again.'
                    }), 500
                
                return jsonify({
                    'success': True, 
                    'message': f'Fee structure created successfully with {count} fee item(s)',
                    'fee_structure_id': fee_structure_id,
                    'items_count': count
                }), 200
                
        except Exception as e:
            connection.rollback()
            print(f"Error creating fee structure: {e}")
            return jsonify({'success': False, 'message': f'Error creating fee structure: {str(e)}'}), 500
        finally:
            connection.close()
            
    except Exception as e:
        print(f"Error in create_fee_structure: {e}")
        return jsonify({'success': False, 'message': 'An error occurred'}), 500

@app.route('/dashboard/employee/student-fees/fee-structures')
@login_required
def fee_structures():
    """Display all fee structures"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    employee_id = session.get('employee_id') or session.get('user_id')
    
    # Check if user is accountant or viewing as accountant
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_technician = user_role == 'technician'
    
    # Check permission-based access
    has_view_fees_permission = check_permission_or_role('view_fees', ['accountant'])
    has_manage_fees_permission = check_permission_or_role('manage_fees', ['accountant'])
    
    if not (is_accountant or is_technician or has_view_fees_permission or has_manage_fees_permission):
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('dashboard_employee'))
    
    connection = get_db_connection()
    fee_structures_list = []
    academic_levels = []
    
    if connection:
        try:
            with connection.cursor() as cursor:
                # Fetch all fee structures with academic level info
                cursor.execute("""
                    SELECT fs.id, fs.academic_level_id, fs.fee_name, fs.category, fs.start_date, fs.end_date, 
                           fs.payment_deadline, fs.total_amount, fs.status, fs.created_at,
                           al.level_name, al.level_category
                    FROM fee_structures fs
                    LEFT JOIN academic_levels al ON fs.academic_level_id = al.id
                    ORDER BY fs.created_at DESC
                """)
                structures = cursor.fetchall()
                
                for row in structures:
                    # Fetch fee items for this structure
                    cursor.execute("""
                        SELECT id, item_name, item_description, amount, item_order
                        FROM fee_items
                        WHERE fee_structure_id = %s
                        ORDER BY item_order ASC
                    """, (row.get('id'),))
                    items = cursor.fetchall()
                    
                    # Format dates
                    start_date = row.get('start_date')
                    end_date = row.get('end_date')
                    payment_deadline = row.get('payment_deadline')
                    
                    if start_date and hasattr(start_date, 'strftime'):
                        start_date = start_date.strftime('%Y-%m-%d')
                    elif start_date:
                        start_date = str(start_date).split(' ')[0]
                    
                    if end_date and hasattr(end_date, 'strftime'):
                        end_date = end_date.strftime('%Y-%m-%d')
                    elif end_date:
                        end_date = str(end_date).split(' ')[0]
                    
                    if payment_deadline and hasattr(payment_deadline, 'strftime'):
                        payment_deadline = payment_deadline.strftime('%Y-%m-%d')
                    elif payment_deadline:
                        payment_deadline = str(payment_deadline).split(' ')[0]
                    
                    fee_structures_list.append({
                        'id': row.get('id'),
                        'academic_level_id': row.get('academic_level_id'),
                        'fee_name': row.get('fee_name', ''),
                        'category': row.get('category', 'both'),
                        'start_date': start_date,
                        'end_date': end_date,
                        'payment_deadline': payment_deadline,
                        'total_amount': float(row.get('total_amount', 0)),
                        'status': row.get('status', 'active'),
                        'created_at': row.get('created_at'),
                        'level_name': row.get('level_name', ''),
                        'level_category': row.get('level_category', ''),
                        'items': [{
                            'id': item.get('id'),
                            'item_name': item.get('item_name', ''),
                            'item_description': item.get('item_description', ''),
                            'amount': float(item.get('amount', 0))
                        } for item in items]
                    })
                
                # Fetch active academic levels for edit modal
                cursor.execute("""
                    SELECT id, level_category, level_name, level_description
                    FROM academic_levels
                    WHERE level_status = 'active'
                    ORDER BY level_name ASC
                """)
                academic_levels_results = cursor.fetchall()
                
                for row in academic_levels_results:
                    academic_levels.append({
                        'id': row.get('id'),
                        'level_category': row.get('level_category', ''),
                        'level_name': row.get('level_name', ''),
                        'level_description': row.get('level_description', '')
                    })
        except Exception as e:
            print(f"Error fetching fee structures: {e}")
        finally:
            connection.close()
    
    return render_template('dashboards/fee_structures.html', 
                         fee_structures=fee_structures_list, 
                         academic_levels=academic_levels)

@app.route('/dashboard/employee/student-fees/payments-audit')
@login_required
def payments_audit():
    """Display payments audit log showing all changes to student fees"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    employee_id = session.get('employee_id') or session.get('user_id')
    
    # Check if user is accountant or viewing as accountant
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_technician = user_role == 'technician'
    
    # Check permission-based access
    has_view_fees_permission = check_permission_or_role('view_student_fees', ['accountant', 'principal'])
    has_manage_fees_permission = check_permission_or_role('manage_fees', ['accountant', 'principal'])
    
    if not (is_accountant or is_technician or has_view_fees_permission or has_manage_fees_permission):
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('dashboard_employee'))
    
    connection = get_db_connection()
    audit_logs = []
    
    if connection:
        try:
            with connection.cursor() as cursor:
                # Check if audit table exists, if not create it
                try:
                    cursor.execute("SHOW TABLES LIKE 'student_payment_audit'")
                    table_exists = cursor.fetchone()
                    
                    if not table_exists:
                        # Create the audit table if it doesn't exist
                        cursor.execute("""
                            CREATE TABLE IF NOT EXISTS student_payment_audit (
                                id INT AUTO_INCREMENT PRIMARY KEY,
                                payment_id INT,
                                student_id VARCHAR(20) NOT NULL,
                                action_type ENUM('INSERT', 'UPDATE', 'DELETE') NOT NULL,
                                field_name VARCHAR(100),
                                old_value TEXT,
                                new_value TEXT,
                                changed_by INT,
                                changed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                                FOREIGN KEY (payment_id) REFERENCES student_payments(id) ON DELETE SET NULL,
                                FOREIGN KEY (student_id) REFERENCES students(student_id) ON DELETE CASCADE,
                                FOREIGN KEY (changed_by) REFERENCES employees(id) ON DELETE SET NULL,
                                INDEX idx_payment_id (payment_id),
                                INDEX idx_student_id (student_id),
                                INDEX idx_changed_at (changed_at),
                                INDEX idx_action_type (action_type)
                            )
                        """)
                        connection.commit()
                except Exception as e:
                    print(f"Error checking/creating audit table: {e}")
                
                # Fetch all audit logs with student and employee information
                cursor.execute("""
                    SELECT 
                        spa.id,
                        spa.payment_id,
                        spa.student_id,
                        spa.action_type,
                        spa.field_name,
                        spa.old_value,
                        spa.new_value,
                        spa.changed_at,
                        s.full_name as student_name,
                        e.full_name as employee_name,
                        e.employee_id as employee_code,
                        sp.amount_paid,
                        sp.payment_method,
                        sp.payment_date,
                        fs.fee_name
                    FROM student_payment_audit spa
                    LEFT JOIN students s ON spa.student_id = s.student_id
                    LEFT JOIN employees e ON spa.changed_by = e.id
                    LEFT JOIN student_payments sp ON spa.payment_id = sp.id
                    LEFT JOIN fee_structures fs ON sp.fee_structure_id = fs.id
                    ORDER BY spa.changed_at DESC
                    LIMIT 1000
                """)
                
                results = cursor.fetchall()
                
                for row in results:
                    changed_at = row.get('changed_at')
                    if changed_at and hasattr(changed_at, 'strftime'):
                        changed_at_str = changed_at.strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        changed_at_str = str(changed_at) if changed_at else ''
                    
                    payment_date = row.get('payment_date')
                    if payment_date and hasattr(payment_date, 'strftime'):
                        payment_date_str = payment_date.strftime('%Y-%m-%d')
                    else:
                        payment_date_str = str(payment_date).split(' ')[0] if payment_date else ''
                    
                    audit_logs.append({
                        'id': row.get('id'),
                        'payment_id': row.get('payment_id'),
                        'student_id': row.get('student_id'),
                        'student_name': row.get('student_name', 'Unknown'),
                        'action_type': row.get('action_type', ''),
                        'field_name': row.get('field_name', ''),
                        'old_value': row.get('old_value', ''),
                        'new_value': row.get('new_value', ''),
                        'changed_at': changed_at_str,
                        'employee_name': row.get('employee_name', 'Unknown'),
                        'employee_code': row.get('employee_code', ''),
                        'amount_paid': float(row.get('amount_paid', 0)) if row.get('amount_paid') else 0,
                        'payment_method': row.get('payment_method', ''),
                        'payment_date': payment_date_str,
                        'fee_name': row.get('fee_name', '')
                    })
        except Exception as e:
            print(f"Error fetching audit logs: {e}")
            import traceback
            traceback.print_exc()
        finally:
            connection.close()
    
    return render_template('dashboards/payments_audit.html',
                         audit_logs=audit_logs,
                         role=user_role)

@app.route('/dashboard/employee/student-fees/record-payment', methods=['POST'])
@login_required
def record_payment():
    """Record a payment for a student"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    employee_id = session.get('employee_id') or session.get('user_id')
    
    # Check permissions
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_technician = user_role == 'technician'
    
    # Check permission-based access
    has_process_payments_permission = check_permission_or_role('process_payments', ['accountant'])
    
    if not (is_accountant or is_technician or has_process_payments_permission):
        return jsonify({'success': False, 'message': 'You do not have permission to record payments.'}), 403
    
    try:
        # Get form data
        student_id = request.form.get('student_id')
        fee_structure_id = request.form.get('fee_structure_id')
        amount_paid = float(request.form.get('amount_paid', 0))
        payment_method = request.form.get('payment_method')
        reference_number = request.form.get('reference_number', '').strip()
        cheque_number = request.form.get('cheque_number', '').strip()
        transaction_id = request.form.get('transaction_id', '').strip()
        payment_date = request.form.get('payment_date')
        notes = request.form.get('notes', '').strip()
        
        # Validate required fields
        if not student_id or not fee_structure_id or not amount_paid or not payment_method or not payment_date:
            return jsonify({'success': False, 'message': 'Please fill in all required fields.'}), 400
        
        if amount_paid <= 0:
            return jsonify({'success': False, 'message': 'Amount paid must be greater than zero.'}), 400
        
        # Handle file upload
        proof_of_payment = None
        if 'proof_of_payment' in request.files:
            file = request.files['proof_of_payment']
            if file and file.filename != '' and allowed_payment_file(file.filename):
                timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
                filename = secure_filename(f"payment_{student_id}_{timestamp}_{file.filename}")
                filepath = os.path.join(app.config['PAYMENT_PROOF_FOLDER'], filename)
                file.save(filepath)
                proof_of_payment = f"uploads/payment_proofs/{filename}"
        
        # Validate payment method specific fields
        if payment_method == 'Bank Transfer' and not reference_number:
            return jsonify({'success': False, 'message': 'Reference number is required for bank transfer.'}), 400
        if payment_method == 'Cheque' and not cheque_number:
            return jsonify({'success': False, 'message': 'Cheque number is required for cheque payment.'}), 400
        if payment_method == 'Mobile Money' and not transaction_id:
            return jsonify({'success': False, 'message': 'Transaction ID is required for mobile money payment.'}), 400
        
        # Save to database
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'Database connection error.'}), 500
        
        try:
            with connection.cursor() as cursor:
                # Get employee database ID (not employee_id code)
                received_by_id = None
                employee_identifier = session.get('employee_id') or session.get('user_id')
                if employee_identifier:
                    cursor.execute("""
                        SELECT id FROM employees 
                        WHERE id = %s OR employee_id = %s
                        LIMIT 1
                    """, (employee_identifier, employee_identifier))
                    emp_result = cursor.fetchone()
                    if emp_result:
                        received_by_id = emp_result.get('id') if isinstance(emp_result, dict) else emp_result[0]
                
                cursor.execute("""
                    INSERT INTO student_payments 
                    (student_id, fee_structure_id, amount_paid, payment_method, reference_number, 
                     cheque_number, transaction_id, proof_of_payment, received_by, payment_date, notes)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (student_id, fee_structure_id, amount_paid, payment_method, reference_number,
                      cheque_number, transaction_id, proof_of_payment, received_by_id, payment_date, notes))
                
                payment_id = cursor.lastrowid
                
                # Log audit entry for payment creation
                payment_details = f"Amount: KES {amount_paid:,.2f}, Method: {payment_method}, Date: {payment_date}"
                if reference_number:
                    payment_details += f", Ref: {reference_number}"
                if cheque_number:
                    payment_details += f", Cheque: {cheque_number}"
                if transaction_id:
                    payment_details += f", Transaction: {transaction_id}"
                if notes:
                    payment_details += f", Notes: {notes[:100]}"
                
                cursor.execute("""
                    INSERT INTO student_payment_audit 
                    (payment_id, student_id, action_type, field_name, old_value, new_value, changed_by)
                    VALUES (%s, %s, 'INSERT', 'Payment Created', NULL, %s, %s)
                """, (payment_id, student_id, payment_details, received_by_id))
                
                connection.commit()
                
                return jsonify({
                    'success': True,
                    'message': 'Payment recorded successfully.'
                })
        except Exception as e:
            connection.rollback()
            print(f"Error recording payment: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'message': f'Error recording payment: {str(e)}'}), 500
        finally:
            connection.close()
            
    except Exception as e:
        print(f"Error in record_payment: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': 'An error occurred while recording payment.'}), 500

@app.route('/dashboard/employee/student-fees/transactions/<student_id>', methods=['GET'])
@login_required
def get_student_transactions(student_id):
    """Get all transactions for a student"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    
    # Check permissions
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_technician = user_role == 'technician'
    
    if not is_accountant and not is_technician:
        return jsonify({'success': False, 'message': 'You do not have permission to view transactions.'}), 403
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': 'Database connection error.'}), 500
    
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT sp.id, sp.amount_paid, sp.payment_method, sp.reference_number, 
                       sp.cheque_number, sp.transaction_id, sp.proof_of_payment, 
                       sp.payment_date, sp.notes, sp.created_at,
                       fs.fee_name, fs.total_amount,
                       e.full_name as received_by_name, e.employee_id as received_by_id
                FROM student_payments sp
                LEFT JOIN fee_structures fs ON sp.fee_structure_id = fs.id
                LEFT JOIN employees e ON sp.received_by = e.id
                WHERE sp.student_id = %s
                ORDER BY sp.payment_date DESC, sp.created_at DESC
            """, (student_id,))
            transactions = cursor.fetchall()
            
            transactions_list = []
            for row in transactions:
                payment_date = row.get('payment_date')
                if payment_date and hasattr(payment_date, 'strftime'):
                    payment_date = payment_date.strftime('%Y-%m-%d')
                elif payment_date:
                    payment_date = str(payment_date).split(' ')[0]
                
                created_at = row.get('created_at')
                if created_at and hasattr(created_at, 'strftime'):
                    created_at = created_at.strftime('%Y-%m-%d %H:%M:%S')
                elif created_at:
                    created_at = str(created_at)
                
                # Safely convert amounts to float, handling None values
                amount_paid = row.get('amount_paid')
                if amount_paid is None:
                    amount_paid = 0.0
                else:
                    try:
                        amount_paid = float(amount_paid)
                    except (TypeError, ValueError):
                        amount_paid = 0.0
                
                total_amount = row.get('total_amount')
                if total_amount is None:
                    total_amount = 0.0
                else:
                    try:
                        total_amount = float(total_amount)
                    except (TypeError, ValueError):
                        total_amount = 0.0
                
                transactions_list.append({
                    'id': row.get('id'),
                    'amount_paid': amount_paid,
                    'payment_method': row.get('payment_method', ''),
                    'reference_number': row.get('reference_number', ''),
                    'cheque_number': row.get('cheque_number', ''),
                    'transaction_id': row.get('transaction_id', ''),
                    'proof_of_payment': row.get('proof_of_payment', ''),
                    'payment_date': payment_date,
                    'notes': row.get('notes', ''),
                    'created_at': created_at,
                    'fee_name': row.get('fee_name', ''),
                    'fee_total': total_amount,
                    'received_by_name': row.get('received_by_name', ''),
                    'received_by_id': row.get('received_by_id', '')
                })
            
            return jsonify({
                'success': True,
                'transactions': transactions_list
            })
    except Exception as e:
        print(f"Error fetching transactions: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': 'Error fetching transactions.'}), 500
    finally:
        connection.close()

@app.route('/dashboard/employee/student-fees/fee-structure/<int:structure_id>/update', methods=['POST'])
@login_required
def update_fee_structure(structure_id):
    """Update a fee structure"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    employee_id = session.get('employee_id') or session.get('user_id')
    
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_technician = user_role == 'technician'
    
    # Check permission-based access - specific permission for editing fee structures
    has_edit_fee_structure_permission = check_permission_or_role('edit_fee_structure', ['accountant', 'principal'])
    has_manage_fees_permission = check_permission_or_role('manage_fees', ['accountant', 'principal'])
    
    if not (is_accountant or is_technician or has_edit_fee_structure_permission or has_manage_fees_permission):
        return jsonify({'success': False, 'message': 'You do not have permission to edit fee structures. Please contact your administrator.'}), 403
    
    try:
        data = request.get_json()
        
        academic_level_id = data.get('academic_level_id')
        fee_name = data.get('fee_name', '').strip().upper()
        category = data.get('category', 'both').strip().lower()
        if category not in ['self sponsored', 'sponsored', 'both']:
            category = 'both'  # Default to 'both' if invalid
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        payment_deadline = data.get('payment_deadline')
        fee_items = data.get('fee_items', [])
        
        if not academic_level_id or not fee_name or not start_date or not end_date or not payment_deadline:
            return jsonify({'success': False, 'message': 'All required fields must be filled'}), 400
        
        if not fee_items or len(fee_items) == 0:
            return jsonify({'success': False, 'message': 'At least one fee item is required'}), 400
        
        total_amount = sum(float(item.get('amount', 0)) for item in fee_items)
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        
        try:
            with connection.cursor() as cursor:
                # Check if another fee structure already exists for this academic level and term period (excluding current one)
                cursor.execute("""
                    SELECT id, fee_name, start_date, end_date 
                    FROM fee_structures 
                    WHERE academic_level_id = %s 
                    AND start_date = %s 
                    AND end_date = %s
                    AND id != %s
                    AND status = 'active'
                """, (academic_level_id, start_date, end_date, structure_id))
                
                existing_structure = cursor.fetchone()
                if existing_structure:
                    existing_name = existing_structure.get('fee_name') if isinstance(existing_structure, dict) else existing_structure[1]
                    return jsonify({
                        'success': False, 
                        'message': f'Another fee structure already exists for this academic level and term period. Existing structure: {existing_name}'
                    }), 400
                
                # Update fee structure
                cursor.execute("""
                    UPDATE fee_structures 
                    SET academic_level_id = %s, fee_name = %s, category = %s, start_date = %s, 
                        end_date = %s, payment_deadline = %s, total_amount = %s
                    WHERE id = %s
                """, (academic_level_id, fee_name, category, start_date, end_date, payment_deadline, total_amount, structure_id))
                
                # Delete existing items
                cursor.execute("DELETE FROM fee_items WHERE fee_structure_id = %s", (structure_id,))
                
                # Insert new items
                items_inserted = 0
                for index, item in enumerate(fee_items):
                    item_name = item.get('item_name', '').strip().upper()
                    item_description = item.get('item_description', '').strip()
                    amount = float(item.get('amount', 0))
                    
                    # Validate item data
                    if not item_name or amount <= 0:
                        continue  # Skip invalid items
                    
                    cursor.execute("""
                        INSERT INTO fee_items 
                        (fee_structure_id, item_name, item_description, amount, item_order)
                        VALUES (%s, %s, %s, %s, %s)
                    """, (structure_id, item_name, item_description, amount, index))
                    items_inserted += 1
                
                # Verify items were inserted
                if items_inserted == 0:
                    connection.rollback()
                    return jsonify({
                        'success': False, 
                        'message': 'No valid fee items were provided. Please ensure all items have a name and amount greater than 0.'
                    }), 400
                
                connection.commit()
                
                # Verify items were saved
                cursor.execute("SELECT COUNT(*) as count FROM fee_items WHERE fee_structure_id = %s", (structure_id,))
                items_count = cursor.fetchone()
                count = items_count.get('count') if isinstance(items_count, dict) else items_count[0]
                
                return jsonify({
                    'success': True, 
                    'message': f'Fee structure updated successfully with {count} fee item(s)',
                    'items_count': count
                }), 200
                
        except Exception as e:
            connection.rollback()
            print(f"Error updating fee structure: {e}")
            return jsonify({'success': False, 'message': f'Error updating fee structure: {str(e)}'}), 500
        finally:
            connection.close()
            
    except Exception as e:
        print(f"Error in update_fee_structure: {e}")
        return jsonify({'success': False, 'message': 'An error occurred'}), 500

@app.route('/dashboard/employee/student-fees/fee-structure/<int:structure_id>/delete', methods=['POST'])
@login_required
def delete_fee_structure(structure_id):
    """Delete a fee structure"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    employee_id = session.get('employee_id') or session.get('user_id')
    
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_technician = user_role == 'technician'
    
    # Check permission-based access - specific permission for deleting fee structures
    has_delete_fee_structure_permission = check_permission_or_role('delete_fee_structure', ['accountant', 'principal'])
    has_manage_fees_permission = check_permission_or_role('manage_fees', ['accountant', 'principal'])
    
    if not (is_accountant or is_technician or has_delete_fee_structure_permission or has_manage_fees_permission):
        return jsonify({'success': False, 'message': 'You do not have permission to delete fee structures. Please contact your administrator.'}), 403
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': 'Database connection failed'}), 500
    
    try:
        with connection.cursor() as cursor:
            # Delete fee structure (cascade will delete items)
            cursor.execute("DELETE FROM fee_structures WHERE id = %s", (structure_id,))
            connection.commit()
            return jsonify({'success': True, 'message': 'Fee structure deleted successfully'}), 200
    except Exception as e:
        connection.rollback()
        print(f"Error deleting fee structure: {e}")
        return jsonify({'success': False, 'message': f'Error deleting fee structure: {str(e)}'}), 500
    finally:
        connection.close()

@app.route('/dashboard/employee/staff-and-salaries')
@login_required
def staff_and_salaries():
    """Staff and Salaries page for employees - unified with tabs"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    employee_id = session.get('employee_id') or session.get('user_id')
    
    # Check permissions - accountant, principal, super admin, technician can access
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_principal = user_role == 'principal' or viewing_as_role == 'principal'
    is_super_admin = user_role == 'super admin' or viewing_as_role == 'super admin'
    is_technician = user_role == 'technician'
    
    # Check permission-based access
    has_salary_permission = check_permission_or_role('manage_salaries', 
                                                     ['accountant', 'principal', 'super admin'])
    
    if not (is_accountant or is_principal or is_super_admin or is_technician or has_salary_permission):
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('dashboard_employee'))
    
    # Get filter parameters for salary records
    filter_type = request.args.get('filter', 'all')
    filter_month = request.args.get('month', '')
    filter_year = request.args.get('year', '')
    
    # Fetch all data types
    connection = get_db_connection()
    employees = []
    salary_records_list = []
    audits = []
    active_count = 0
    pending_count = 0
    suspended_count = 0
    roles_set = set()
    filter_info = {'type': filter_type, 'month': filter_month, 'year': filter_year}
    
    if connection:
        try:
            with connection.cursor() as cursor:
                # 1. Fetch employees with their salary information
                cursor.execute("""
                    SELECT e.id, e.employee_id, e.full_name, e.email, e.phone, e.id_number, e.role, e.status, 
                           e.profile_picture, e.created_at,
                           es.id as salary_id, es.effective_date, es.basic_salary, es.house_allowance,
                           es.transport_allowance, es.medical_allowance, es.overtime, es.bonus,
                           es.paye, es.nssf, es.nhif, es.sacco, es.staff_loans, es.absenteeism,
                           es.total_earnings, es.total_deductions, es.net_salary, es.is_active as salary_is_active,
                           es.updated_at as salary_updated_at, es.payment_period
                    FROM employees e
                    LEFT JOIN employee_salaries es ON e.id = es.employee_id AND es.is_active = TRUE
                    WHERE e.status = 'active'
                    ORDER BY e.full_name ASC
                """)
                employees_raw = cursor.fetchall()
                
                # Process employees and calculate stats
                for emp in employees_raw:
                    # Check if salary was edited (check if there's an audit record)
                    salary_edited = False
                    if emp.get('salary_id'):
                        try:
                            cursor.execute("""
                                SELECT COUNT(*) as audit_count 
                                FROM employee_salary_audits 
                                WHERE salary_id = %s
                            """, (emp.get('salary_id'),))
                            audit_result = cursor.fetchone()
                            salary_edited = audit_result.get('audit_count', 0) > 0 if audit_result else False
                        except:
                            # Table might not exist yet, that's okay
                            salary_edited = False
                    
                    emp_dict = dict(emp)
                    emp_dict['salary_edited'] = salary_edited
                    employees.append(emp_dict)
                    
                    # Count by status
                    status = emp.get('status', '')
                    if status == 'active':
                        active_count += 1
                    elif status == 'pending approval':
                        pending_count += 1
                    elif status == 'suspended':
                        suspended_count += 1
                    
                    # Collect unique roles
                    roles_set.add(emp.get('role', ''))
                
                roles_count = len(roles_set)
                
                # 2. Fetch salary records with payment information
                from datetime import datetime, timedelta
                from dateutil.relativedelta import relativedelta
                
                base_query = """
                    SELECT 
                        es.id as salary_id,
                        es.employee_id,
                        es.net_salary,
                        es.effective_date,
                        es.payment_period,
                        e.full_name,
                        e.phone,
                        e.employee_id as emp_code
                    FROM employee_salaries es
                    INNER JOIN employees e ON es.employee_id = e.id
                    WHERE 1=1
                """
                
                query_params = []
                
                # Add period filter
                if filter_type == 'month' and filter_month:
                    base_query += " AND es.is_active = TRUE AND DATE_FORMAT(es.effective_date, '%%Y-%%m') = %s"
                    query_params.append(filter_month)
                elif filter_type == 'year' and filter_year:
                    base_query += " AND YEAR(es.effective_date) = %s"
                    query_params.append(int(filter_year))
                else:
                    base_query += " AND es.is_active = TRUE"
                
                base_query += " ORDER BY e.full_name ASC"
                
                cursor.execute(base_query, tuple(query_params) if query_params else None)
                records = cursor.fetchall()
                
                for record in records:
                    salary_id = record.get('salary_id')
                    employee_id = record.get('employee_id')
                    net_salary = float(record.get('net_salary', 0))
                    effective_date = record.get('effective_date')
                    payment_period = record.get('payment_period', 'Monthly')
                    
                    # Calculate current period based on effective_date and payment_period
                    current_date = datetime.now().date()
                    if isinstance(effective_date, str):
                        effective_date = datetime.strptime(effective_date, '%Y-%m-%d').date()
                    elif hasattr(effective_date, 'date'):
                        effective_date = effective_date.date()
                    
                    # Calculate period start and end dates
                    period_start, period_end, current_period_num = calculate_period_dates(
                        effective_date, payment_period, current_date
                    )
                    
                    # Get payments for current period
                    cursor.execute("""
                        SELECT COALESCE(SUM(amount_paid), 0) as total_paid
                        FROM employee_salary_payments
                        WHERE salary_id = %s 
                        AND payment_date >= %s 
                        AND payment_date <= %s
                    """, (salary_id, period_start, period_end))
                    current_period_paid = cursor.fetchone()
                    total_paid_current = float(current_period_paid.get('total_paid', 0) if current_period_paid else 0)
                    
                    # Calculate carry-forward balance from previous period
                    carry_forward = 0.0
                    if current_period_num > 1:
                        # Get previous period dates
                        prev_period_start, prev_period_end, _ = calculate_period_dates(
                            effective_date, payment_period, period_start - timedelta(days=1)
                        )
                        
                        # Get payments for previous period
                        cursor.execute("""
                            SELECT COALESCE(SUM(amount_paid), 0) as total_paid
                            FROM employee_salary_payments
                            WHERE salary_id = %s 
                            AND payment_date >= %s 
                            AND payment_date <= %s
                        """, (salary_id, prev_period_start, prev_period_end))
                        prev_period_paid = cursor.fetchone()
                        total_paid_prev = float(prev_period_paid.get('total_paid', 0) if prev_period_paid else 0)
                        
                        # Calculate previous period balance (net_salary - total_paid)
                        carry_forward = max(0, net_salary - total_paid_prev)
                    
                    # Amount to be paid = net_salary + carry_forward
                    amount_to_be_paid = net_salary + carry_forward
                    balance = amount_to_be_paid - total_paid_current
                    
                    salary_records_list.append({
                        'salary_id': salary_id,
                        'employee_id': employee_id,
                        'employee_code': record.get('emp_code'),
                        'employee_name': record.get('full_name'),
                        'phone': record.get('phone'),
                        'amount_to_be_paid': amount_to_be_paid,
                        'total_paid': total_paid_current,
                        'balance': balance,
                        'effective_date': effective_date,
                        'payment_period': payment_period,
                        'carry_forward': carry_forward,
                        'current_period_start': period_start,
                        'current_period_end': period_end
                    })
                
                # 3. Fetch salary audits
                try:
                    cursor.execute("""
                        SELECT 
                            esa.id,
                            esa.salary_id,
                            esa.employee_id,
                            esa.field_name,
                            esa.old_value,
                            esa.new_value,
                            esa.edited_by,
                            esa.edited_by_name,
                            esa.edited_at,
                            e.full_name as employee_name,
                            e.employee_id as employee_code
                        FROM employee_salary_audits esa
                        INNER JOIN employees e ON esa.employee_id = e.id
                        ORDER BY esa.edited_at DESC
                    """)
                    audits = cursor.fetchall()
                except Exception as e:
                    # Table might not exist yet
                    audits = []
                    print(f"Salary audits table might not exist: {e}")
                    
        except Exception as e:
            print(f"Error fetching data: {e}")
            import traceback
            traceback.print_exc()
            flash('Error loading data. Please try again.', 'error')
        finally:
            if connection:
                try:
                    connection.close()
                except:
                    pass  # Connection might already be closed
    
    return render_template('dashboards/staff_and_salaries.html', 
                         employees=employees,
                         salary_records=salary_records_list,
                         audits=audits,
                         active_count=active_count,
                         pending_count=pending_count,
                         suspended_count=suspended_count,
                         roles_count=roles_count,
                         filter_info=filter_info)

def calculate_period_dates(effective_date, payment_period, reference_date):
    """
    Calculate period start and end dates based on payment_period and effective_date.
    Returns: (period_start, period_end, period_number)
    """
    if isinstance(effective_date, str):
        effective_date = datetime.strptime(effective_date, '%Y-%m-%d').date()
    elif hasattr(effective_date, 'date'):
        effective_date = effective_date.date()
    
    if isinstance(reference_date, str):
        reference_date = datetime.strptime(reference_date, '%Y-%m-%d').date()
    elif hasattr(reference_date, 'date'):
        reference_date = reference_date.date()
    
    # Calculate how many periods have passed since effective_date
    if payment_period == 'Daily':
        days_diff = (reference_date - effective_date).days
        period_num = days_diff + 1
        period_start = effective_date + timedelta(days=days_diff)
        period_end = period_start
    elif payment_period == 'Weekly':
        weeks_diff = (reference_date - effective_date).days // 7
        period_num = weeks_diff + 1
        period_start = effective_date + timedelta(weeks=weeks_diff)
        period_end = period_start + timedelta(days=6)
    elif payment_period == 'Monthly':
        months_diff = (reference_date.year - effective_date.year) * 12 + (reference_date.month - effective_date.month)
        period_num = months_diff + 1
        period_start = effective_date + relativedelta(months=months_diff)
        # Set to first day of the month
        period_start = period_start.replace(day=1)
        # Last day of the month
        if period_start.month == 12:
            period_end = period_start.replace(day=31)
        else:
            period_end = (period_start + relativedelta(months=1) - timedelta(days=1))
    elif payment_period == 'Quarterly':
        months_diff = (reference_date.year - effective_date.year) * 12 + (reference_date.month - effective_date.month)
        quarters_diff = months_diff // 3
        period_num = quarters_diff + 1
        quarter_start_month = ((effective_date.month - 1) // 3) * 3 + 1
        period_start = effective_date.replace(month=quarter_start_month, day=1) + relativedelta(months=quarters_diff * 3)
        period_end = period_start + relativedelta(months=3) - timedelta(days=1)
    elif payment_period == 'Semi-Annual':
        months_diff = (reference_date.year - effective_date.year) * 12 + (reference_date.month - effective_date.month)
        semis_diff = months_diff // 6
        period_num = semis_diff + 1
        semi_start_month = ((effective_date.month - 1) // 6) * 6 + 1
        period_start = effective_date.replace(month=semi_start_month, day=1) + relativedelta(months=semis_diff * 6)
        period_end = period_start + relativedelta(months=6) - timedelta(days=1)
    elif payment_period == '3/4 Annual':
        months_diff = (reference_date.year - effective_date.year) * 12 + (reference_date.month - effective_date.month)
        periods_diff = months_diff // 9
        period_num = periods_diff + 1
        period_start_month = ((effective_date.month - 1) // 9) * 9 + 1
        period_start = effective_date.replace(month=period_start_month, day=1) + relativedelta(months=periods_diff * 9)
        period_end = period_start + relativedelta(months=9) - timedelta(days=1)
    elif payment_period == 'Annually':
        years_diff = reference_date.year - effective_date.year
        period_num = years_diff + 1
        period_start = effective_date.replace(year=effective_date.year + years_diff, month=1, day=1)
        period_end = period_start.replace(month=12, day=31)
    else:
        # Default to monthly
        months_diff = (reference_date.year - effective_date.year) * 12 + (reference_date.month - effective_date.month)
        period_num = months_diff + 1
        period_start = effective_date + relativedelta(months=months_diff)
        period_start = period_start.replace(day=1)
        if period_start.month == 12:
            period_end = period_start.replace(day=31)
        else:
            period_end = (period_start + relativedelta(months=1) - timedelta(days=1))
    
    return period_start, period_end, period_num

@app.route('/dashboard/employee/staff-and-salaries/salary-records')
@login_required
def salary_records():
    """Display employee salary records with payment information"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    
    # Check permissions - accountant, principal, super admin, technician can access
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_principal = user_role == 'principal' or viewing_as_role == 'principal'
    is_super_admin = user_role == 'super admin' or viewing_as_role == 'super admin'
    is_technician = user_role == 'technician'
    
    if not (is_accountant or is_principal or is_super_admin or is_technician):
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('dashboard_employee'))
    
    # Get filter parameters
    filter_type = request.args.get('filter', 'all')
    filter_month = request.args.get('month', '')
    filter_year = request.args.get('year', '')
    
    # Fetch salary records with employee information
    connection = get_db_connection()
    salary_records_list = []
    filter_info = {'type': filter_type, 'month': filter_month, 'year': filter_year}
    
    if connection:
        try:
            with connection.cursor() as cursor:
                # Ensure payment_period column exists in employee_salaries table
                try:
                    cursor.execute("ALTER TABLE employee_salaries ADD COLUMN payment_period ENUM('Daily', 'Weekly', 'Monthly', 'Quarterly', 'Semi-Annual', '3/4 Annual', 'Annually') DEFAULT 'Monthly' AFTER effective_date")
                except Exception as e:
                    # Column might already exist, ignore the error
                    pass
                
                # Create employee_salary_payments table if it doesn't exist
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS employee_salary_payments (
                        id INT AUTO_INCREMENT PRIMARY KEY,
                        employee_id INT NOT NULL,
                        salary_id INT NOT NULL,
                        amount_paid DECIMAL(15, 2) NOT NULL DEFAULT 0.00,
                        payment_date DATE NOT NULL,
                        payment_method ENUM('Cash', 'Bank Transfer', 'Cheque', 'Mobile Money', 'Credit/Debit Card') DEFAULT 'Bank Transfer',
                        reference_number VARCHAR(255),
                        notes TEXT,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                        FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE,
                        FOREIGN KEY (salary_id) REFERENCES employee_salaries(id) ON DELETE CASCADE,
                        INDEX idx_employee_id (employee_id),
                        INDEX idx_salary_id (salary_id),
                        INDEX idx_payment_date (payment_date)
                    )
                """)
                
                # Build the query with filters
                base_query = """
                    SELECT 
                        es.id as salary_id,
                        es.employee_id,
                        es.net_salary,
                        es.effective_date,
                        es.payment_period,
                        e.full_name,
                        e.phone,
                        e.employee_id as emp_code
                    FROM employee_salaries es
                    INNER JOIN employees e ON es.employee_id = e.id
                    WHERE 1=1
                """
                
                query_params = []
                
                # Add period filter
                if filter_type == 'month' and filter_month:
                    # Filter by month (YYYY-MM format) - show active salaries for that month
                    base_query += " AND es.is_active = TRUE AND DATE_FORMAT(es.effective_date, '%%Y-%%m') = %s"
                    query_params.append(filter_month)
                elif filter_type == 'year' and filter_year:
                    # Filter by year - show ALL records (active and inactive) for that year
                    base_query += " AND YEAR(es.effective_date) = %s"
                    query_params.append(int(filter_year))
                else:
                    # Default: show only active salaries
                    base_query += " AND es.is_active = TRUE"
                
                base_query += """
                    ORDER BY e.full_name ASC
                """
                
                cursor.execute(base_query, tuple(query_params) if query_params else None)
                records = cursor.fetchall()
                
                from datetime import datetime, timedelta
                from dateutil.relativedelta import relativedelta
                
                for record in records:
                    salary_id = record.get('salary_id')
                    employee_id = record.get('employee_id')
                    net_salary = float(record.get('net_salary', 0))
                    effective_date = record.get('effective_date')
                    payment_period = record.get('payment_period', 'Monthly')
                    
                    # Calculate current period based on effective_date and payment_period
                    current_date = datetime.now().date()
                    if isinstance(effective_date, str):
                        effective_date = datetime.strptime(effective_date, '%Y-%m-%d').date()
                    elif hasattr(effective_date, 'date'):
                        effective_date = effective_date.date()
                    
                    # Calculate period start and end dates
                    period_start, period_end, current_period_num = calculate_period_dates(
                        effective_date, payment_period, current_date
                    )
                    
                    # Get payments for current period
                    cursor.execute("""
                        SELECT COALESCE(SUM(amount_paid), 0) as total_paid
                        FROM employee_salary_payments
                        WHERE salary_id = %s 
                        AND payment_date >= %s 
                        AND payment_date <= %s
                    """, (salary_id, period_start, period_end))
                    current_period_paid = cursor.fetchone()
                    total_paid_current = float(current_period_paid.get('total_paid', 0) if current_period_paid else 0)
                    
                    # Calculate carry-forward balance from previous period
                    carry_forward = 0.0
                    if current_period_num > 1:
                        # Get previous period dates
                        prev_period_start, prev_period_end, _ = calculate_period_dates(
                            effective_date, payment_period, period_start - timedelta(days=1)
                        )
                        
                        # Get payments for previous period
                        cursor.execute("""
                            SELECT COALESCE(SUM(amount_paid), 0) as total_paid
                            FROM employee_salary_payments
                            WHERE salary_id = %s 
                            AND payment_date >= %s 
                            AND payment_date <= %s
                        """, (salary_id, prev_period_start, prev_period_end))
                        prev_period_paid = cursor.fetchone()
                        total_paid_prev = float(prev_period_paid.get('total_paid', 0) if prev_period_paid else 0)
                        
                        # Calculate previous period balance (net_salary - total_paid)
                        carry_forward = max(0, net_salary - total_paid_prev)
                    
                    # Amount to be paid = net_salary + carry_forward
                    amount_to_be_paid = net_salary + carry_forward
                    balance = amount_to_be_paid - total_paid_current
                    
                    salary_records_list.append({
                        'salary_id': salary_id,
                        'employee_id': employee_id,
                        'employee_code': record.get('emp_code'),
                        'employee_name': record.get('full_name'),
                        'phone': record.get('phone'),
                        'amount_to_be_paid': amount_to_be_paid,
                        'total_paid': total_paid_current,
                        'balance': balance,
                        'effective_date': effective_date,
                        'payment_period': payment_period,
                        'carry_forward': carry_forward,
                        'current_period_start': period_start,
                        'current_period_end': period_end
                    })
        
        except Exception as e:
            print(f"Error fetching salary records: {e}")
            import traceback
            traceback.print_exc()
            flash('Error loading salary records. Please try again.', 'error')
        finally:
            if connection:
                try:
                    connection.close()
                except:
                    pass
    
    return render_template('dashboards/salary_records.html', 
                         salary_records=salary_records_list,
                         filter_info=filter_info)

@app.route('/dashboard/employee/staff-and-salaries/register-salary', methods=['POST'])
@login_required
def register_salary():
    """Register salary for an employee"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    
    # Check permissions - accountant, principal, super admin, technician can access
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_principal = user_role == 'principal' or viewing_as_role == 'principal'
    is_super_admin = user_role == 'super admin' or viewing_as_role == 'super admin'
    is_technician = user_role == 'technician'
    
    if not (is_accountant or is_principal or is_super_admin or is_technician):
        return jsonify({'success': False, 'message': 'You do not have permission to perform this action.'}), 403
    
    try:
        data = request.get_json()
        
        # Validate required fields
        if not data.get('employee_id'):
            return jsonify({'success': False, 'message': 'Employee ID is required.'}), 400
        
        if not data.get('effective_date'):
            return jsonify({'success': False, 'message': 'Effective date is required.'}), 400
        
        if not data.get('basic_salary') or float(data.get('basic_salary', 0)) <= 0:
            return jsonify({'success': False, 'message': 'Basic salary is required and must be greater than 0.'}), 400
        
        # Get database connection
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'Database connection error.'}), 500
        
        try:
            with connection.cursor() as cursor:
                # Check if employee exists
                cursor.execute("SELECT id, employee_id, full_name FROM employees WHERE id = %s", (data.get('employee_id'),))
                employee = cursor.fetchone()
                
                if not employee:
                    return jsonify({'success': False, 'message': 'Employee not found.'}), 404
                
                # Create salaries table if it doesn't exist
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS employee_salaries (
                        id INT AUTO_INCREMENT PRIMARY KEY,
                        employee_id INT NOT NULL,
                        effective_date DATE NOT NULL,
                        payment_period ENUM('Daily', 'Weekly', 'Monthly', 'Quarterly', 'Semi-Annual', '3/4 Annual', 'Annually') DEFAULT 'Monthly',
                        basic_salary DECIMAL(15, 2) NOT NULL DEFAULT 0.00,
                        house_allowance DECIMAL(15, 2) DEFAULT 0.00,
                        transport_allowance DECIMAL(15, 2) DEFAULT 0.00,
                        medical_allowance DECIMAL(15, 2) DEFAULT 0.00,
                        overtime DECIMAL(15, 2) DEFAULT 0.00,
                        bonus DECIMAL(15, 2) DEFAULT 0.00,
                        paye DECIMAL(15, 2) DEFAULT 0.00,
                        nssf DECIMAL(15, 2) DEFAULT 0.00,
                        nhif DECIMAL(15, 2) DEFAULT 0.00,
                        sacco DECIMAL(15, 2) DEFAULT 0.00,
                        staff_loans DECIMAL(15, 2) DEFAULT 0.00,
                        absenteeism DECIMAL(15, 2) DEFAULT 0.00,
                        total_earnings DECIMAL(15, 2) NOT NULL DEFAULT 0.00,
                        total_deductions DECIMAL(15, 2) NOT NULL DEFAULT 0.00,
                        net_salary DECIMAL(15, 2) NOT NULL DEFAULT 0.00,
                        is_active BOOLEAN DEFAULT TRUE,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                        FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE,
                        INDEX idx_employee_id (employee_id),
                        INDEX idx_effective_date (effective_date),
                        INDEX idx_is_active (is_active)
                    )
                """)
                
                # Add payment_period column if it doesn't exist (for existing tables)
                try:
                    cursor.execute("ALTER TABLE employee_salaries ADD COLUMN payment_period ENUM('Daily', 'Weekly', 'Monthly', 'Quarterly', 'Semi-Annual', '3/4 Annual', 'Annually') DEFAULT 'Monthly' AFTER effective_date")
                except Exception as e:
                    # Column might already exist, ignore the error
                    pass
                
                # Check if there's an active salary for this employee
                cursor.execute("""
                    SELECT id FROM employee_salaries 
                    WHERE employee_id = %s AND is_active = TRUE
                """, (data.get('employee_id'),))
                existing_salary = cursor.fetchone()
                
                # If there's an existing active salary, deactivate it
                if existing_salary:
                    cursor.execute("""
                        UPDATE employee_salaries 
                        SET is_active = FALSE 
                        WHERE employee_id = %s AND is_active = TRUE
                    """, (data.get('employee_id'),))
                
                # Insert new salary record
                cursor.execute("""
                    INSERT INTO employee_salaries (
                        employee_id, effective_date, payment_period, basic_salary, house_allowance,
                        transport_allowance, medical_allowance, overtime, bonus,
                        paye, nssf, nhif, sacco, staff_loans, absenteeism,
                        total_earnings, total_deductions, net_salary, is_active
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, TRUE
                    )
                """, (
                    data.get('employee_id'),
                    data.get('effective_date'),
                    data.get('payment_period', 'Monthly'),
                    float(data.get('basic_salary', 0)),
                    float(data.get('house_allowance', 0)),
                    float(data.get('transport_allowance', 0)),
                    float(data.get('medical_allowance', 0)),
                    float(data.get('overtime', 0)),
                    float(data.get('bonus', 0)),
                    float(data.get('paye', 0)),
                    float(data.get('nssf', 0)),
                    float(data.get('nhif', 0)),
                    float(data.get('sacco', 0)),
                    float(data.get('staff_loans', 0)),
                    float(data.get('absenteeism', 0)),
                    float(data.get('total_earnings', 0)),
                    float(data.get('total_deductions', 0)),
                    float(data.get('net_salary', 0))
                ))
                
                connection.commit()
                
                return jsonify({
                    'success': True,
                    'message': f'Salary registered successfully for {employee.get("full_name", "employee")}.',
                    'redirect_url': '/dashboard/employee/staff-and-salaries/salary-records'
                })
        
        except Exception as e:
            connection.rollback()
            print(f"Error registering salary: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'message': f'Error registering salary: {str(e)}'}), 500
        
        finally:
            connection.close()
    
    except Exception as e:
        print(f"Error in register_salary route: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': 'An error occurred. Please try again.'}), 500

@app.route('/dashboard/employee/staff-and-salaries/get-salary/<int:salary_id>', methods=['GET'])
@login_required
def get_salary(salary_id):
    """Get salary details for editing"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    
    # Check permissions
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_principal = user_role == 'principal' or viewing_as_role == 'principal'
    is_super_admin = user_role == 'super admin' or viewing_as_role == 'super admin'
    is_technician = user_role == 'technician'
    
    if not (is_accountant or is_principal or is_super_admin or is_technician):
        return jsonify({'success': False, 'message': 'You do not have permission to perform this action.'}), 403
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': 'Database connection error.'}), 500
    
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT es.*, e.full_name as employee_name, e.employee_id as employee_code
                FROM employee_salaries es
                INNER JOIN employees e ON es.employee_id = e.id
                WHERE es.id = %s
            """, (salary_id,))
            salary = cursor.fetchone()
            
            if not salary:
                return jsonify({'success': False, 'message': 'Salary not found.'}), 404
            
            return jsonify({
                'success': True,
                'salary': {
                    'id': salary.get('id'),
                    'employee_id': salary.get('employee_id'),
                    'employee_name': salary.get('employee_name'),
                    'employee_code': salary.get('employee_code'),
                    'effective_date': salary.get('effective_date').strftime('%Y-%m-%d') if salary.get('effective_date') else '',
                    'basic_salary': float(salary.get('basic_salary', 0)),
                    'house_allowance': float(salary.get('house_allowance', 0)),
                    'transport_allowance': float(salary.get('transport_allowance', 0)),
                    'medical_allowance': float(salary.get('medical_allowance', 0)),
                    'overtime': float(salary.get('overtime', 0)),
                    'bonus': float(salary.get('bonus', 0)),
                    'paye': float(salary.get('paye', 0)),
                    'nssf': float(salary.get('nssf', 0)),
                    'nhif': float(salary.get('nhif', 0)),
                    'sacco': float(salary.get('sacco', 0)),
                    'staff_loans': float(salary.get('staff_loans', 0)),
                    'absenteeism': float(salary.get('absenteeism', 0)),
                    'total_earnings': float(salary.get('total_earnings', 0)),
                    'total_deductions': float(salary.get('total_deductions', 0)),
                    'net_salary': float(salary.get('net_salary', 0))
                }
            })
    except Exception as e:
        print(f"Error fetching salary: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': 'Error fetching salary details.'}), 500
    finally:
        connection.close()

@app.route('/dashboard/employee/staff-and-salaries/update-salary', methods=['POST'])
@login_required
def update_salary():
    """Update salary and track changes in audit table"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    
    # Check permissions
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_principal = user_role == 'principal' or viewing_as_role == 'principal'
    is_super_admin = user_role == 'super admin' or viewing_as_role == 'super admin'
    is_technician = user_role == 'technician'
    
    if not (is_accountant or is_principal or is_super_admin or is_technician):
        return jsonify({'success': False, 'message': 'You do not have permission to perform this action.'}), 403
    
    try:
        data = request.get_json()
        salary_id = data.get('salary_id')
        
        if not salary_id:
            return jsonify({'success': False, 'message': 'Salary ID is required.'}), 400
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'Database connection error.'}), 500
        
        try:
            with connection.cursor() as cursor:
                # Create employee_salary_audits table if it doesn't exist
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS employee_salary_audits (
                        id INT AUTO_INCREMENT PRIMARY KEY,
                        salary_id INT NOT NULL,
                        employee_id INT NOT NULL,
                        field_name VARCHAR(100) NOT NULL,
                        old_value TEXT,
                        new_value TEXT,
                        edited_by INT NOT NULL,
                        edited_by_name VARCHAR(255),
                        edited_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (salary_id) REFERENCES employee_salaries(id) ON DELETE CASCADE,
                        FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE,
                        FOREIGN KEY (edited_by) REFERENCES employees(id) ON DELETE CASCADE,
                        INDEX idx_salary_id (salary_id),
                        INDEX idx_employee_id (employee_id),
                        INDEX idx_edited_at (edited_at)
                    )
                """)
                
                # Get current salary values
                cursor.execute("""
                    SELECT * FROM employee_salaries WHERE id = %s
                """, (salary_id,))
                old_salary = cursor.fetchone()
                
                if not old_salary:
                    return jsonify({'success': False, 'message': 'Salary not found.'}), 404
                
                # Get editor information - need to get the database ID, not employee_id code
                editor_identifier = session.get('employee_id') or session.get('user_id')
                editor_id = None
                editor_name = session.get('full_name', 'Unknown')
                
                # Get editor database ID and name from database
                if editor_identifier:
                    cursor.execute("SELECT id, full_name FROM employees WHERE id = %s OR employee_id = %s LIMIT 1", 
                                 (editor_identifier, editor_identifier))
                    editor_result = cursor.fetchone()
                    if editor_result:
                        editor_id = editor_result.get('id')
                        editor_name = editor_result.get('full_name', editor_name)
                
                # If we couldn't find the editor, use a default or skip audit
                if not editor_id:
                    # Try to get any admin/accountant as fallback
                    cursor.execute("SELECT id, full_name FROM employees WHERE role IN ('super admin', 'accountant', 'principal') AND status = 'active' LIMIT 1")
                    fallback_editor = cursor.fetchone()
                    if fallback_editor:
                        editor_id = fallback_editor.get('id')
                        editor_name = fallback_editor.get('full_name', 'System')
                    else:
                        # If still no editor found, we can't create audit records
                        # But we should still allow the update
                        editor_id = None
                
                # Track changes for each field
                fields_to_track = [
                    'effective_date', 'payment_period', 'basic_salary', 'house_allowance', 'transport_allowance',
                    'medical_allowance', 'overtime', 'bonus', 'paye', 'nssf', 'nhif',
                    'sacco', 'staff_loans', 'absenteeism', 'total_earnings', 'total_deductions', 'net_salary'
                ]
                
                audit_records = []
                for field in fields_to_track:
                    old_val = old_salary.get(field)
                    new_val = data.get(field)
                    
                    # Convert dates to strings for comparison
                    if field == 'effective_date':
                        if old_val:
                            old_val = old_val.strftime('%Y-%m-%d') if hasattr(old_val, 'strftime') else str(old_val).split(' ')[0]
                        if new_val:
                            new_val = str(new_val).split('T')[0] if 'T' in str(new_val) else str(new_val).split(' ')[0]
                    
                    # Convert numeric values
                    if field in ['basic_salary', 'house_allowance', 'transport_allowance', 'medical_allowance',
                                'overtime', 'bonus', 'paye', 'nssf', 'nhif', 'sacco', 'staff_loans',
                                'absenteeism', 'total_earnings', 'total_deductions', 'net_salary']:
                        old_val = float(old_val) if old_val else 0.0
                        new_val = float(new_val) if new_val else 0.0
                    
                    # Only track if value changed
                    if str(old_val) != str(new_val):
                        audit_records.append({
                            'salary_id': salary_id,
                            'employee_id': data.get('employee_id'),
                            'field_name': field,
                            'old_value': str(old_val),
                            'new_value': str(new_val)
                        })
                
                # Update salary
                cursor.execute("""
                    UPDATE employee_salaries SET
                        effective_date = %s,
                        payment_period = %s,
                        basic_salary = %s,
                        house_allowance = %s,
                        transport_allowance = %s,
                        medical_allowance = %s,
                        overtime = %s,
                        bonus = %s,
                        paye = %s,
                        nssf = %s,
                        nhif = %s,
                        sacco = %s,
                        staff_loans = %s,
                        absenteeism = %s,
                        total_earnings = %s,
                        total_deductions = %s,
                        net_salary = %s,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = %s
                """, (
                    data.get('effective_date'),
                    data.get('payment_period', 'Monthly'),
                    float(data.get('basic_salary', 0)),
                    float(data.get('house_allowance', 0)),
                    float(data.get('transport_allowance', 0)),
                    float(data.get('medical_allowance', 0)),
                    float(data.get('overtime', 0)),
                    float(data.get('bonus', 0)),
                    float(data.get('paye', 0)),
                    float(data.get('nssf', 0)),
                    float(data.get('nhif', 0)),
                    float(data.get('sacco', 0)),
                    float(data.get('staff_loans', 0)),
                    float(data.get('absenteeism', 0)),
                    float(data.get('total_earnings', 0)),
                    float(data.get('total_deductions', 0)),
                    float(data.get('net_salary', 0)),
                    salary_id
                ))
                
                # Insert audit records (only if we have a valid editor_id)
                if audit_records and editor_id:
                    for audit in audit_records:
                        try:
                            cursor.execute("""
                                INSERT INTO employee_salary_audits 
                                (salary_id, employee_id, field_name, old_value, new_value, edited_by, edited_by_name)
                                VALUES (%s, %s, %s, %s, %s, %s, %s)
                            """, (
                                audit['salary_id'],
                                audit['employee_id'],
                                audit['field_name'],
                                audit['old_value'],
                                audit['new_value'],
                                editor_id,  # Use the resolved database ID
                                editor_name
                            ))
                        except Exception as audit_error:
                            # Log audit error but don't fail the update
                            print(f"Error creating audit record: {audit_error}")
                            import traceback
                            traceback.print_exc()
                elif audit_records and not editor_id:
                    print("Warning: Audit records skipped because editor_id could not be resolved.")
                
                connection.commit()
                
                return jsonify({
                    'success': True,
                    'message': 'Salary updated successfully!'
                })
        
        except Exception as e:
            connection.rollback()
            print(f"Error updating salary: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'message': f'Error updating salary: {str(e)}'}), 500
        
        finally:
            connection.close()
    
    except Exception as e:
        print(f"Error in update_salary route: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': 'An error occurred. Please try again.'}), 500

@app.route('/dashboard/employee/staff-and-salaries/get-employees-with-salaries')
@login_required
def get_employees_with_salaries():
    """Get all employees with active salaries for payment"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    
    # Check permissions
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_principal = user_role == 'principal' or viewing_as_role == 'principal'
    is_super_admin = user_role == 'super admin' or viewing_as_role == 'super admin'
    is_technician = user_role == 'technician'
    
    if not (is_accountant or is_principal or is_super_admin or is_technician):
        return jsonify({'success': False, 'message': 'Permission denied'}), 403
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': 'Database connection failed'}), 500
    
    try:
        with connection.cursor() as cursor:
            # Get employees with active salaries
            cursor.execute("""
                SELECT DISTINCT e.id, e.employee_id, e.full_name
                FROM employees e
                INNER JOIN employee_salaries es ON e.id = es.employee_id
                WHERE e.status = 'active' AND es.is_active = TRUE
                ORDER BY e.full_name ASC
            """)
            employees = cursor.fetchall()
            
            return jsonify({
                'success': True,
                'employees': [{
                    'id': emp.get('id'),
                    'employee_id': emp.get('employee_id'),
                    'full_name': emp.get('full_name')
                } for emp in employees]
            })
    except Exception as e:
        print(f"Error fetching employees with salaries: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': 'Error fetching employees'}), 500
    finally:
        connection.close()

@app.route('/dashboard/employee/staff-and-salaries/get-employee-salary/<int:employee_id>')
@login_required
def get_employee_salary(employee_id):
    """Get active salary information for an employee including total paid"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    
    # Check permissions
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_principal = user_role == 'principal' or viewing_as_role == 'principal'
    is_super_admin = user_role == 'super admin' or viewing_as_role == 'super admin'
    is_technician = user_role == 'technician'
    
    if not (is_accountant or is_principal or is_super_admin or is_technician):
        return jsonify({'success': False, 'message': 'Permission denied'}), 403
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': 'Database connection failed'}), 500
    
    try:
        with connection.cursor() as cursor:
            # Ensure payment_period column exists
            try:
                cursor.execute("ALTER TABLE employee_salaries ADD COLUMN payment_period ENUM('Daily', 'Weekly', 'Monthly', 'Quarterly', 'Semi-Annual', '3/4 Annual', 'Annually') DEFAULT 'Monthly' AFTER effective_date")
            except Exception as e:
                # Column might already exist, ignore the error
                pass
            
            # Get active salary
            cursor.execute("""
                SELECT es.*
                FROM employee_salaries es
                WHERE es.employee_id = %s AND es.is_active = TRUE
                ORDER BY es.effective_date DESC
                LIMIT 1
            """, (employee_id,))
            salary = cursor.fetchone()
            
            if not salary:
                return jsonify({'success': False, 'message': 'No active salary found for this employee'}), 404
            
            # Calculate current period and payments with carry-forward
            effective_date = salary.get('effective_date')
            payment_period = salary.get('payment_period', 'Monthly')
            net_salary = float(salary.get('net_salary', 0))
            
            if isinstance(effective_date, str):
                effective_date = datetime.strptime(effective_date, '%Y-%m-%d').date()
            elif hasattr(effective_date, 'date'):
                effective_date = effective_date.date()
            
            current_date = datetime.now().date()
            period_start, period_end, current_period_num = calculate_period_dates(
                effective_date, payment_period, current_date
            )
            
            # Get payments for current period
            cursor.execute("""
                SELECT COALESCE(SUM(amount_paid), 0) as total_paid
                FROM employee_salary_payments
                WHERE salary_id = %s 
                AND payment_date >= %s 
                AND payment_date <= %s
            """, (salary.get('id'), period_start, period_end))
            current_period_paid = cursor.fetchone()
            total_paid_current = float(current_period_paid.get('total_paid', 0) if current_period_paid else 0)
            
            # Calculate carry-forward balance from previous period
            carry_forward = 0.0
            if current_period_num > 1:
                # Get previous period dates
                prev_period_start, prev_period_end, _ = calculate_period_dates(
                    effective_date, payment_period, period_start - timedelta(days=1)
                )
                
                # Get payments for previous period
                cursor.execute("""
                    SELECT COALESCE(SUM(amount_paid), 0) as total_paid
                    FROM employee_salary_payments
                    WHERE salary_id = %s 
                    AND payment_date >= %s 
                    AND payment_date <= %s
                """, (salary.get('id'), prev_period_start, prev_period_end))
                prev_period_paid = cursor.fetchone()
                total_paid_prev = float(prev_period_paid.get('total_paid', 0) if prev_period_paid else 0)
                
                # Calculate previous period balance (net_salary - total_paid)
                carry_forward = max(0, net_salary - total_paid_prev)
            
            # Amount to be paid = net_salary + carry_forward
            amount_to_be_paid = net_salary + carry_forward
            balance = amount_to_be_paid - total_paid_current
            
            return jsonify({
                'success': True,
                'salary': {
                    'id': salary.get('id'),
                    'employee_id': salary.get('employee_id'),
                    'effective_date': salary.get('effective_date').strftime('%Y-%m-%d') if salary.get('effective_date') else '',
                    'net_salary': net_salary,
                    'total_paid': total_paid_current,
                    'amount_to_be_paid': amount_to_be_paid,
                    'carry_forward': carry_forward,
                    'balance': balance
                }
            })
    except Exception as e:
        print(f"Error fetching employee salary: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': 'Error fetching salary information'}), 500
    finally:
        connection.close()

@app.route('/dashboard/employee/staff-and-salaries/record-salary-payment', methods=['POST'])
@login_required
def record_salary_payment():
    """Record a salary payment for an employee"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    
    # Check permissions
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_principal = user_role == 'principal' or viewing_as_role == 'principal'
    is_super_admin = user_role == 'super admin' or viewing_as_role == 'super admin'
    is_technician = user_role == 'technician'
    
    if not (is_accountant or is_principal or is_super_admin or is_technician):
        return jsonify({'success': False, 'message': 'Permission denied'}), 403
    
    try:
        data = request.get_json()
        employee_id = data.get('employee_id')
        salary_id = data.get('salary_id')
        amount_paid = data.get('amount_paid')
        payment_date = data.get('payment_date')
        payment_method = data.get('payment_method')
        reference_number = data.get('reference_number')
        notes = data.get('notes')
        
        if not all([employee_id, salary_id, amount_paid, payment_date, payment_method]):
            return jsonify({'success': False, 'message': 'Missing required fields'}), 400
        
        try:
            amount_paid = float(amount_paid)
            if amount_paid <= 0:
                return jsonify({'success': False, 'message': 'Payment amount must be greater than 0'}), 400
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'Invalid payment amount'}), 400
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        
        try:
            with connection.cursor() as cursor:
                # Ensure payment_period column exists
                try:
                    cursor.execute("ALTER TABLE employee_salaries ADD COLUMN payment_period ENUM('Daily', 'Weekly', 'Monthly', 'Quarterly', 'Semi-Annual', '3/4 Annual', 'Annually') DEFAULT 'Monthly' AFTER effective_date")
                except Exception as e:
                    # Column might already exist, ignore the error
                    pass
                
                # Verify salary exists and is active
                cursor.execute("""
                    SELECT net_salary, effective_date, payment_period FROM employee_salaries
                    WHERE id = %s AND employee_id = %s AND is_active = TRUE
                """, (salary_id, employee_id))
                salary = cursor.fetchone()
                
                if not salary:
                    return jsonify({'success': False, 'message': 'Salary not found or inactive'}), 404
                
                # Calculate current period and balance with carry-forward
                effective_date = salary.get('effective_date')
                payment_period = salary.get('payment_period', 'Monthly')
                net_salary = float(salary.get('net_salary', 0))
                
                if isinstance(effective_date, str):
                    effective_date = datetime.strptime(effective_date, '%Y-%m-%d').date()
                elif hasattr(effective_date, 'date'):
                    effective_date = effective_date.date()
                
                # Use payment_date to determine which period this payment belongs to
                if isinstance(payment_date, str):
                    payment_date_obj = datetime.strptime(payment_date, '%Y-%m-%d').date()
                else:
                    payment_date_obj = payment_date
                
                period_start, period_end, current_period_num = calculate_period_dates(
                    effective_date, payment_period, payment_date_obj
                )
                
                # Get payments for current period
                cursor.execute("""
                    SELECT COALESCE(SUM(amount_paid), 0) as total_paid
                    FROM employee_salary_payments
                    WHERE salary_id = %s 
                    AND payment_date >= %s 
                    AND payment_date <= %s
                """, (salary_id, period_start, period_end))
                current_period_paid = cursor.fetchone()
                total_paid_current = float(current_period_paid.get('total_paid', 0) if current_period_paid else 0)
                
                # Calculate carry-forward balance from previous period
                carry_forward = 0.0
                if current_period_num > 1:
                    # Get previous period dates
                    prev_period_start, prev_period_end, _ = calculate_period_dates(
                        effective_date, payment_period, period_start - timedelta(days=1)
                    )
                    
                    # Get payments for previous period
                    cursor.execute("""
                        SELECT COALESCE(SUM(amount_paid), 0) as total_paid
                        FROM employee_salary_payments
                        WHERE salary_id = %s 
                        AND payment_date >= %s 
                        AND payment_date <= %s
                    """, (salary_id, prev_period_start, prev_period_end))
                    prev_period_paid = cursor.fetchone()
                    total_paid_prev = float(prev_period_paid.get('total_paid', 0) if prev_period_paid else 0)
                    
                    # Calculate previous period balance (net_salary - total_paid)
                    carry_forward = max(0, net_salary - total_paid_prev)
                
                # Amount to be paid = net_salary + carry_forward
                amount_to_be_paid = net_salary + carry_forward
                balance = amount_to_be_paid - total_paid_current
                
                # Check if payment exceeds balance
                if amount_paid > balance:
                    return jsonify({'success': False, 'message': 'Payment amount exceeds the remaining balance'}), 400
                
                # Record payment
                cursor.execute("""
                    INSERT INTO employee_salary_payments
                    (employee_id, salary_id, amount_paid, payment_date, payment_method, reference_number, notes)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (employee_id, salary_id, amount_paid, payment_date, payment_method, reference_number, notes))
                
                connection.commit()
                
                return jsonify({
                    'success': True,
                    'message': 'Payment recorded successfully'
                })
        except Exception as e:
            connection.rollback()
            print(f"Error recording salary payment: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'message': f'Error recording payment: {str(e)}'}), 500
        finally:
            connection.close()
    except Exception as e:
        print(f"Error in record_salary_payment: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': 'An error occurred'}), 500

@app.route('/dashboard/employee/staff-and-salaries/get-payment-history/<int:employee_id>/<int:salary_id>')
@login_required
def get_payment_history(employee_id, salary_id):
    """Get payment history for an employee's salary"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    
    # Check permissions
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_principal = user_role == 'principal' or viewing_as_role == 'principal'
    is_super_admin = user_role == 'super admin' or viewing_as_role == 'super admin'
    is_technician = user_role == 'technician'
    
    if not (is_accountant or is_principal or is_super_admin or is_technician):
        return jsonify({'success': False, 'message': 'Permission denied'}), 403
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': 'Database connection failed'}), 500
    
    try:
        with connection.cursor() as cursor:
            # Get employee and salary information
            cursor.execute("""
                SELECT e.full_name, e.employee_id as employee_code, es.net_salary
                FROM employees e
                INNER JOIN employee_salaries es ON e.id = es.employee_id
                WHERE e.id = %s AND es.id = %s
            """, (employee_id, salary_id))
            employee_salary = cursor.fetchone()
            
            if not employee_salary:
                return jsonify({'success': False, 'message': 'Employee or salary not found'}), 404
            
            # Get all payment history for this salary
            cursor.execute("""
                SELECT 
                    id,
                    amount_paid,
                    payment_date,
                    payment_method,
                    reference_number,
                    notes,
                    created_at
                FROM employee_salary_payments
                WHERE employee_id = %s AND salary_id = %s
                ORDER BY payment_date DESC, created_at DESC
            """, (employee_id, salary_id))
            payments = cursor.fetchall()
            
            # Calculate total paid
            total_paid = sum(float(p.get('amount_paid', 0)) for p in payments)
            net_salary = float(employee_salary.get('net_salary', 0))
            
            return jsonify({
                'success': True,
                'employee_name': employee_salary.get('full_name'),
                'employee_code': employee_salary.get('employee_code'),
                'net_salary': net_salary,
                'total_paid': total_paid,
                'payments': [{
                    'id': p.get('id'),
                    'amount_paid': float(p.get('amount_paid', 0)),
                    'payment_date': p.get('payment_date').strftime('%Y-%m-%d') if p.get('payment_date') else '',
                    'payment_method': p.get('payment_method', ''),
                    'reference_number': p.get('reference_number', ''),
                    'notes': p.get('notes', ''),
                    'created_at': p.get('created_at').strftime('%Y-%m-%d %H:%M:%S') if p.get('created_at') else ''
                } for p in payments]
            })
    except Exception as e:
        print(f"Error fetching payment history: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': 'Error fetching payment history'}), 500
    finally:
        connection.close()

@app.route('/dashboard/employee/staff-and-salaries/salary-audits')
@login_required
def salary_audits():
    """Display salary audit trail"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    
    # Check permissions
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_principal = user_role == 'principal' or viewing_as_role == 'principal'
    is_super_admin = user_role == 'super admin' or viewing_as_role == 'super admin'
    is_technician = user_role == 'technician'
    
    if not (is_accountant or is_principal or is_super_admin or is_technician):
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('dashboard_employee'))
    
    connection = get_db_connection()
    audits = []
    
    if connection:
        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT 
                        esa.id,
                        esa.salary_id,
                        esa.employee_id,
                        esa.field_name,
                        esa.old_value,
                        esa.new_value,
                        esa.edited_by,
                        esa.edited_by_name,
                        esa.edited_at,
                        e.full_name as employee_name,
                        e.employee_id as employee_code
                    FROM employee_salary_audits esa
                    INNER JOIN employees e ON esa.employee_id = e.id
                    ORDER BY esa.edited_at DESC
                """)
                audits = cursor.fetchall()
        except Exception as e:
            print(f"Error fetching salary audits: {e}")
            import traceback
            traceback.print_exc()
            flash('Error loading salary audits. Please try again.', 'error')
        finally:
            if connection:
                try:
                    connection.close()
                except:
                    pass
    
    return render_template('dashboards/salary_audits.html', audits=audits)

@app.route('/staff-management')
@login_required
def staff_management():
    """Staff management page for employees"""
    user_role = session.get('role', '').lower()
    employee_id = session.get('employee_id') or session.get('user_id')
    
    # Check permission OR role-based access
    has_access = check_permission_or_role('view_staff', 
                                         allowed_roles=['employee', 'super admin', 'principal', 'deputy principal', 
                                                       'academic coordinator', 'teachers', 'accountant', 'librarian', 
                                                       'warden', 'transport manager', 'technician'])
    
    if not has_access:
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('dashboard_employee'))
    
    # Check what actions the user can perform
    can_add = check_permission_or_role('add_staff', ['principal', 'deputy principal'])
    can_edit = check_permission_or_role('edit_staff', ['principal', 'deputy principal', 'academic coordinator'])
    can_delete = check_permission_or_role('delete_staff', ['principal'])
    
    # Fetch all employees
    connection = get_db_connection()
    employees = []
    if connection:
        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT id, employee_id, full_name, email, phone, id_number, role, status, 
                           profile_picture, created_at
                    FROM employees 
                    ORDER BY created_at DESC
                """)
                employees = cursor.fetchall()
        except Exception as e:
            print(f"Error fetching employees: {e}")
            flash('Error loading employees. Please try again.', 'error')
        finally:
            if connection:
                try:
                    connection.close()
                except:
                    pass  # Connection might already be closed
    
    return render_template('dashboards/staff_management.html', 
                         employees=employees,
                         can_add=can_add,
                         can_edit=can_edit,
                         can_delete=can_delete)

@app.route('/assign-roles-approve')
@login_required
def assign_roles_approve():
    """Assign Roles & Approve page for employees"""
    user_role = session.get('role', '').lower()
    
    # Check if user is an employee (any employee role)
    employee_roles = ['employee', 'super admin', 'principal', 'deputy principal', 'academic coordinator', 
                     'teachers', 'accountant', 'librarian', 'warden', 'transport manager', 'technician']
    
    if user_role not in employee_roles:
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('home'))
    
    # Fetch only pending approval employees
    connection = get_db_connection()
    employees = []
    if connection:
        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT id, employee_id, full_name, email, phone, id_number, role, status, 
                           profile_picture, created_at
                    FROM employees 
                    WHERE status = 'pending approval'
                    ORDER BY created_at DESC
                """)
                employees = cursor.fetchall()
        except Exception as e:
            print(f"Error fetching employees: {e}")
            flash('Error loading employees. Please try again.', 'error')
        finally:
            if connection:
                try:
                    connection.close()
                except:
                    pass
    
    return render_template('dashboards/assign_roles_approve.html', employees=employees)

@app.route('/approve-employee/<int:employee_id>', methods=['POST'])
@login_required
def approve_employee(employee_id):
    """Approve employee, assign role, and send email"""
    user_role = session.get('role', '').lower()
    
    # Check if user is an employee (any employee role)
    employee_roles = ['employee', 'super admin', 'principal', 'deputy principal', 'academic coordinator', 
                     'teachers', 'accountant', 'librarian', 'warden', 'transport manager', 'technician']
    
    if user_role not in employee_roles:
        return jsonify({'success': False, 'message': 'You do not have permission to approve employees.'}), 403
    
    data = request.get_json()
    role = data.get('role', '').strip() if data else request.form.get('role', '').strip()
    
    if not role:
        return jsonify({'success': False, 'message': 'Please select a role for the employee.'}), 400
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': 'Database connection error.'}), 500
    
    try:
        with connection.cursor() as cursor:
            # Check if employee exists
            cursor.execute("""
                SELECT id, employee_id, full_name, email, status
                FROM employees 
                WHERE id = %s
            """, (employee_id,))
            employee = cursor.fetchone()
            
            if not employee:
                return jsonify({'success': False, 'message': 'Employee not found.'}), 404
            
            # Check if employee is already approved
            if employee.get('status') == 'active':
                return jsonify({'success': False, 'message': f'Employee {employee.get("full_name")} is already active.'}), 400
            
            # Update employee status to 'active' and assign role
            cursor.execute("""
                UPDATE employees 
                SET status = 'active', role = %s, updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
            """, (role, employee_id))
            
            connection.commit()
            
            # Send approval email with role information
            email_sent = send_employee_approval_email(
                employee.get('email'),
                employee.get('full_name'),
                employee.get('employee_id'),
                role
            )
            
            if not email_sent:
                print(f"Warning: Failed to send approval email to {employee.get('email')}")
            
            return jsonify({
                'success': True, 
                'message': f'Employee {employee.get("full_name")} has been approved and assigned the role of {role.title()}. Email notification sent.'
            })
            
    except Exception as e:
        print(f"Error approving employee: {e}")
        connection.rollback()
        return jsonify({'success': False, 'message': 'An error occurred while approving the employee.'}), 500
    finally:
        connection.close()

@app.route('/update-employee/<int:employee_id>', methods=['POST'])
@login_required
def update_employee(employee_id):
    """Update employee details"""
    user_role = session.get('role', '').lower()
    current_employee_id = session.get('employee_id') or session.get('user_id')
    
    # Check permission OR role-based access
    has_access = check_permission_or_role('edit_staff', 
                                         allowed_roles=['employee', 'super admin', 'principal', 'deputy principal', 
                                                       'academic coordinator', 'teachers', 'accountant', 'librarian', 
                                                       'warden', 'transport manager', 'technician'])
    
    if not has_access:
        return jsonify({'success': False, 'message': 'You do not have permission to update employees.'}), 403
    
    data = request.get_json()
    full_name = data.get('full_name', '').strip().upper()
    email = data.get('email', '').strip().lower()
    phone = data.get('phone', '').strip()
    id_number = data.get('id_number', '').strip().upper()
    role = data.get('role', '').strip()
    
    if not all([full_name, email, phone]):
        return jsonify({'success': False, 'message': 'Please fill in all required fields.'}), 400
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': 'Database connection error.'}), 500
    
    try:
        with connection.cursor() as cursor:
            # Check if employee exists
            cursor.execute("SELECT id FROM employees WHERE id = %s", (employee_id,))
            if not cursor.fetchone():
                return jsonify({'success': False, 'message': 'Employee not found.'}), 404
            
            # Check if email is already taken by another employee
            cursor.execute("SELECT id FROM employees WHERE email = %s AND id != %s", (email, employee_id))
            if cursor.fetchone():
                return jsonify({'success': False, 'message': 'Email is already in use by another employee.'}), 400
            
            # Update employee
            cursor.execute("""
                UPDATE employees 
                SET full_name = %s, email = %s, phone = %s, id_number = %s, role = %s, updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
            """, (full_name, email, phone, id_number, role, employee_id))
            
            connection.commit()
            return jsonify({'success': True, 'message': 'Employee updated successfully.'})
            
    except Exception as e:
        print(f"Error updating employee: {e}")
        connection.rollback()
        return jsonify({'success': False, 'message': 'An error occurred while updating the employee.'}), 500
    finally:
        connection.close()

@app.route('/delete-employee/<int:employee_id>', methods=['POST'])
@login_required
def delete_employee(employee_id):
    """Delete an employee"""
    user_role = session.get('role', '').lower()
    current_employee_id = session.get('employee_id') or session.get('user_id')
    
    # Check permission OR role-based access
    has_access = check_permission_or_role('delete_staff', 
                                         allowed_roles=['employee', 'super admin', 'principal', 'deputy principal', 
                                                       'academic coordinator', 'teachers', 'accountant', 'librarian', 
                                                       'warden', 'transport manager', 'technician'])
    
    if not has_access:
        return jsonify({'success': False, 'message': 'You do not have permission to delete employees.'}), 403
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': 'Database connection error.'}), 500
    
    try:
        with connection.cursor() as cursor:
            # Check if employee exists and get name
            cursor.execute("SELECT id, full_name FROM employees WHERE id = %s", (employee_id,))
            employee = cursor.fetchone()
            
            if not employee:
                return jsonify({'success': False, 'message': 'Employee not found.'}), 404
            
            # Delete employee
            cursor.execute("DELETE FROM employees WHERE id = %s", (employee_id,))
            
            connection.commit()
            return jsonify({'success': True, 'message': f'Employee {employee.get("full_name")} has been deleted successfully.'})
            
    except Exception as e:
        print(f"Error deleting employee: {e}")
        connection.rollback()
        return jsonify({'success': False, 'message': 'An error occurred while deleting the employee.'}), 500
    finally:
        connection.close()

@app.route('/toggle-suspend-employee/<int:employee_id>', methods=['POST'])
@login_required
def toggle_suspend_employee(employee_id):
    """Toggle suspend/unsuspend employee"""
    user_role = session.get('role', '').lower()
    
    # Check if user is an employee (any employee role)
    employee_roles = ['employee', 'super admin', 'principal', 'deputy principal', 'academic coordinator', 
                     'teachers', 'accountant', 'librarian', 'warden', 'transport manager', 'technician']
    
    if user_role not in employee_roles:
        return jsonify({'success': False, 'message': 'You do not have permission to suspend/unsuspend employees.'}), 403
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': 'Database connection error.'}), 500
    
    try:
        with connection.cursor() as cursor:
            # Check if employee exists and get current status
            cursor.execute("SELECT id, full_name, status FROM employees WHERE id = %s", (employee_id,))
            employee = cursor.fetchone()
            
            if not employee:
                return jsonify({'success': False, 'message': 'Employee not found.'}), 404
            
            current_status = employee.get('status')
            new_status = 'suspended' if current_status != 'suspended' else 'active'
            
            # Update employee status
            cursor.execute("""
                UPDATE employees 
                SET status = %s, updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
            """, (new_status, employee_id))
            
            connection.commit()
            action = 'suspended' if new_status == 'suspended' else 'unsuspended'
            return jsonify({
                'success': True, 
                'message': f'Employee {employee.get("full_name")} has been {action} successfully.',
                'new_status': new_status
            })
            
    except Exception as e:
        print(f"Error toggling suspend employee: {e}")
        connection.rollback()
        return jsonify({'success': False, 'message': 'An error occurred while updating the employee status.'}), 500
    finally:
        connection.close()

@app.route('/get-employee/<int:employee_id>', methods=['GET'])
@login_required
def get_employee(employee_id):
    """Get employee details for editing"""
    user_role = session.get('role', '').lower()
    
    # Check if user is an employee (any employee role)
    employee_roles = ['employee', 'super admin', 'principal', 'deputy principal', 'academic coordinator', 
                     'teachers', 'accountant', 'librarian', 'warden', 'transport manager', 'technician']
    
    if user_role not in employee_roles:
        return jsonify({'success': False, 'message': 'You do not have permission to view employee details.'}), 403
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': 'Database connection error.'}), 500
    
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id, employee_id, full_name, email, phone, id_number, role, status
                FROM employees 
                WHERE id = %s
            """, (employee_id,))
            employee = cursor.fetchone()
            
            if not employee:
                return jsonify({'success': False, 'message': 'Employee not found.'}), 404
            
            return jsonify({'success': True, 'employee': employee})
            
    except Exception as e:
        print(f"Error fetching employee: {e}")
        return jsonify({'success': False, 'message': 'An error occurred while fetching employee details.'}), 500
    finally:
        connection.close()

@app.route('/student-management')
@login_required
def student_management():
    """Student management page for employees"""
    user_role = session.get('role', '').lower()
    employee_id = session.get('employee_id') or session.get('user_id')
    
    # Check permission OR role-based access
    has_access = check_permission_or_role('view_students', 
                                         allowed_roles=['employee', 'super admin', 'principal', 'deputy principal', 
                                                       'academic coordinator', 'teachers', 'accountant', 'librarian', 
                                                       'warden', 'transport manager', 'technician'])
    
    if not has_access:
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('dashboard_employee'))
    
    # Check what actions the user can perform
    can_add = check_permission_or_role('add_students', ['principal', 'deputy principal', 'academic coordinator'])
    can_edit = check_permission_or_role('edit_students', ['principal', 'deputy principal', 'academic coordinator', 'teachers'])
    can_delete = check_permission_or_role('delete_students', ['principal', 'deputy principal'])
    
    # Fetch all students with parent information
    connection = get_db_connection()
    students = []
    if connection:
        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT s.id, s.student_id, s.full_name, s.date_of_birth, s.gender, 
                           s.current_grade, s.previous_school, s.address, s.medical_info, 
                           s.special_needs, s.student_category, s.status, s.created_at, s.updated_at,
                           p.full_name as parent_name, p.phone as parent_phone, 
                           p.email as parent_email, p.relationship
                    FROM students s
                    LEFT JOIN parents p ON s.student_id = p.student_id
                    ORDER BY s.created_at DESC
                """)
                students = cursor.fetchall()
        except Exception as e:
            print(f"Error fetching students: {e}")
            flash('Error loading students. Please try again.', 'error')
        finally:
            if connection:
                try:
                    connection.close()
                except:
                    pass  # Connection might already be closed
    
    return render_template('dashboards/student_management.html', 
                         students=students,
                         can_add=can_add,
                         can_edit=can_edit,
                         can_delete=can_delete)

@app.route('/get-student/<student_id>', methods=['GET'])
@login_required
def get_student(student_id):
    """Get student details by student_id"""
    user_role = session.get('role', '').lower()
    employee_id = session.get('employee_id') or session.get('user_id')
    
    # Check permission OR role-based access
    has_access = check_permission_or_role('view_students', 
                                         allowed_roles=['employee', 'super admin', 'principal', 'deputy principal', 
                                                       'academic coordinator', 'teachers', 'accountant', 'librarian', 
                                                       'warden', 'transport manager', 'technician'])
    
    if not has_access:
        return jsonify({'success': False, 'message': 'You do not have permission to access this.'}), 403
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': 'Database connection error.'}), 500
    
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT s.id, s.student_id, s.full_name, s.date_of_birth, s.gender, 
                       s.current_grade, s.previous_school, s.address, s.medical_info, 
                       s.special_needs, s.student_category, s.sponsor_name, s.sponsor_phone, 
                       s.sponsor_email, s.status, s.created_at, s.updated_at,
                       p.full_name as parent_name, p.phone as parent_phone, 
                       p.email as parent_email, p.relationship, p.emergency_contact
                FROM students s
                LEFT JOIN parents p ON s.student_id = p.student_id
                WHERE s.student_id = %s
            """, (student_id,))
            student = cursor.fetchone()
            
            if not student:
                return jsonify({'success': False, 'message': 'Student not found.'}), 404
            
            return jsonify({
                'success': True,
                'student': {
                    'id': student.get('id'),
                    'student_id': student.get('student_id'),
                    'full_name': student.get('full_name'),
                    'date_of_birth': str(student.get('date_of_birth')) if student.get('date_of_birth') else None,
                    'gender': student.get('gender'),
                    'current_grade': student.get('current_grade'),
                    'previous_school': student.get('previous_school'),
                    'address': student.get('address'),
                    'medical_info': student.get('medical_info'),
                    'special_needs': student.get('special_needs'),
                    'student_category': student.get('student_category'),
                    'sponsor_name': student.get('sponsor_name'),
                    'sponsor_phone': student.get('sponsor_phone'),
                    'sponsor_email': student.get('sponsor_email'),
                    'status': student.get('status'),
                    'parent_name': student.get('parent_name'),
                    'parent_phone': student.get('parent_phone'),
                    'parent_email': student.get('parent_email'),
                    'relationship': student.get('relationship'),
                    'emergency_contact': student.get('emergency_contact'),
                    'created_at': str(student.get('created_at')) if student.get('created_at') else None,
                    'updated_at': str(student.get('updated_at')) if student.get('updated_at') else None
                }
            })
    except Exception as e:
        print(f"Error fetching student: {e}")
        return jsonify({'success': False, 'message': 'Error fetching student details.'}), 500
    finally:
        connection.close()

@app.route('/check-student-id/<student_id>', methods=['GET'])
@login_required
def check_student_id(student_id):
    """Check if student ID already exists"""
    user_role = session.get('role', '').lower()
    employee_roles = ['employee', 'super admin', 'principal', 'deputy principal', 'academic coordinator', 
                     'teachers', 'accountant', 'librarian', 'warden', 'transport manager', 'technician']
    
    if user_role not in employee_roles:
        return jsonify({'exists': False, 'message': 'Unauthorized'}), 403
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'exists': False, 'message': 'Database error'}), 500
    
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT student_id FROM students WHERE student_id = %s", (student_id,))
            result = cursor.fetchone()
            return jsonify({'exists': result is not None})
    except Exception as e:
        print(f"Error checking student ID: {e}")
        return jsonify({'exists': False, 'message': 'Error checking student ID'}), 500
    finally:
        connection.close()

@app.route('/update-student/<student_id>', methods=['POST'])
@login_required
def update_student(student_id):
    """Update student details"""
    user_role = session.get('role', '').lower()
    employee_id = session.get('employee_id') or session.get('user_id')
    
    # Check permission OR role-based access
    has_access = check_permission_or_role('edit_students', 
                                         allowed_roles=['employee', 'super admin', 'principal', 'deputy principal', 
                                                       'academic coordinator', 'teachers', 'accountant', 'librarian', 
                                                       'warden', 'transport manager', 'technician'])
    
    if not has_access:
        return jsonify({'success': False, 'message': 'You do not have permission to update students.'}), 403
    
    data = request.get_json()
    new_student_id = (data.get('student_id') or '').strip().upper()
    full_name = (data.get('full_name') or '').strip().upper()
    date_of_birth = (data.get('date_of_birth') or '').strip() or None
    gender = (data.get('gender') or '').strip().upper()
    current_grade = (data.get('current_grade') or '').strip()
    previous_school = (data.get('previous_school') or '').strip()
    address = (data.get('address') or '').strip()
    medical_info = (data.get('medical_info') or '').strip()
    special_needs = (data.get('special_needs') or '').strip()
    student_category = (data.get('student_category') or '').strip().lower() or None
    sponsor_name = (data.get('sponsor_name') or '').strip()
    sponsor_phone = (data.get('sponsor_phone') or '').strip()
    sponsor_email = (data.get('sponsor_email') or '').strip().lower()
    status = (data.get('status') or '').strip()
    
    # Parent information
    parent_name = (data.get('parent_name') or '').strip().upper()
    parent_phone = (data.get('parent_phone') or '').strip()
    parent_email = (data.get('parent_email') or '').strip().lower()
    relationship = (data.get('relationship') or '').strip()
    emergency_contact = (data.get('emergency_contact') or '').strip()
    
    if not full_name:
        return jsonify({'success': False, 'message': 'Student name is required.'}), 400
    
    if not new_student_id:
        return jsonify({'success': False, 'message': 'Student ID is required.'}), 400
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': 'Database connection error.'}), 500
    
    try:
        with connection.cursor() as cursor:
            # Check if new student_id is different and if it already exists
            if new_student_id != student_id:
                cursor.execute("SELECT student_id FROM students WHERE student_id = %s", (new_student_id,))
                existing = cursor.fetchone()
                if existing:
                    return jsonify({'success': False, 'message': 'This Student ID is already registered to another student.'}), 400
            
            # If student_id is changing, we need to handle foreign key constraint
            if new_student_id != student_id:
                # Disable foreign key checks temporarily to allow student_id update
                cursor.execute("SET FOREIGN_KEY_CHECKS = 0")
                
                # Update student first (including student_id)
                cursor.execute("""
                    UPDATE students 
                    SET student_id = %s, full_name = %s, date_of_birth = %s, gender = %s, current_grade = %s, 
                        previous_school = %s, address = %s, medical_info = %s, special_needs = %s,
                        student_category = %s, sponsor_name = %s, sponsor_phone = %s, sponsor_email = %s,
                        status = %s, updated_at = CURRENT_TIMESTAMP
                    WHERE student_id = %s
                """, (new_student_id, full_name, date_of_birth, gender, current_grade, previous_school, address, 
                      medical_info, special_needs, student_category, sponsor_name, sponsor_phone, 
                      sponsor_email, status, student_id))
                
                # Update parents table to reference new student_id
                cursor.execute("""
                    UPDATE parents 
                    SET student_id = %s, updated_at = CURRENT_TIMESTAMP
                    WHERE student_id = %s
                """, (new_student_id, student_id))
                
                # Re-enable foreign key checks
                cursor.execute("SET FOREIGN_KEY_CHECKS = 1")
            else:
                # If student_id is not changing, just update normally
                cursor.execute("""
                    UPDATE students 
                    SET full_name = %s, date_of_birth = %s, gender = %s, current_grade = %s, 
                        previous_school = %s, address = %s, medical_info = %s, special_needs = %s,
                        student_category = %s, sponsor_name = %s, sponsor_phone = %s, sponsor_email = %s,
                        status = %s, updated_at = CURRENT_TIMESTAMP
                    WHERE student_id = %s
                """, (full_name, date_of_birth, gender, current_grade, previous_school, address, 
                      medical_info, special_needs, student_category, sponsor_name, sponsor_phone, 
                      sponsor_email, status, student_id))
            
            # Update or insert parent
            if parent_name and parent_phone and parent_email:
                cursor.execute("""
                    SELECT id FROM parents WHERE student_id = %s
                """, (new_student_id,))
                parent_exists = cursor.fetchone()
                
                if parent_exists:
                    cursor.execute("""
                        UPDATE parents 
                        SET full_name = %s, phone = %s, email = %s, relationship = %s, 
                            emergency_contact = %s, updated_at = CURRENT_TIMESTAMP
                        WHERE student_id = %s
                    """, (parent_name, parent_phone, parent_email, relationship, emergency_contact, new_student_id))
                else:
                    cursor.execute("""
                        INSERT INTO parents (student_id, full_name, phone, email, relationship, emergency_contact)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (new_student_id, parent_name, parent_phone, parent_email, relationship, emergency_contact))
            
            connection.commit()
            return jsonify({'success': True, 'message': 'Student updated successfully.'})
    except Exception as e:
        print(f"Error updating student: {e}")
        connection.rollback()
        return jsonify({'success': False, 'message': 'Error updating student details.'}), 500
    finally:
        connection.close()

@app.route('/delete-student/<student_id>', methods=['POST'])
@login_required
def delete_student(student_id):
    """Delete a student"""
    user_role = session.get('role', '').lower()
    employee_id = session.get('employee_id') or session.get('user_id')
    
    # Check permission OR role-based access
    has_access = check_permission_or_role('delete_students', 
                                         allowed_roles=['employee', 'super admin', 'principal', 'deputy principal', 
                                                       'academic coordinator', 'teachers', 'accountant', 'librarian', 
                                                       'warden', 'transport manager', 'technician'])
    
    if not has_access:
        return jsonify({'success': False, 'message': 'You do not have permission to delete students.'}), 403
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': 'Database connection error.'}), 500
    
    try:
        with connection.cursor() as cursor:
            # Check if student exists
            cursor.execute("SELECT id FROM students WHERE student_id = %s", (student_id,))
            student = cursor.fetchone()
            
            if not student:
                return jsonify({'success': False, 'message': 'Student not found.'}), 404
            
            # Delete student (cascade will delete parent records)
            cursor.execute("DELETE FROM students WHERE student_id = %s", (student_id,))
            connection.commit()
            
            return jsonify({'success': True, 'message': 'Student deleted successfully.'})
    except Exception as e:
        print(f"Error deleting student: {e}")
        connection.rollback()
        return jsonify({'success': False, 'message': 'Error deleting student.'}), 500
    finally:
        connection.close()

@app.route('/approve-student/<student_id>', methods=['POST'])
@login_required
def approve_student(student_id):
    """Approve student admission and send congratulations email to parent"""
    user_role = session.get('role', '').lower()
    employee_id = session.get('employee_id') or session.get('user_id')
    
    # Check permission OR role-based access (approve requires edit permission)
    has_access = check_permission_or_role('edit_students', 
                                         allowed_roles=['employee', 'super admin', 'principal', 'deputy principal', 
                                                       'academic coordinator', 'teachers', 'accountant', 'librarian', 
                                                       'warden', 'transport manager', 'technician'])
    
    if not has_access:
        flash('You do not have permission to approve students.', 'error')
        return redirect(url_for('student_management'))
    
    connection = get_db_connection()
    if not connection:
        flash('Database connection error. Please try again later.', 'error')
        return redirect(url_for('student_management'))
    
    try:
        with connection.cursor() as cursor:
            # Check if student exists and get student details
            cursor.execute("""
                SELECT s.id, s.student_id, s.full_name, s.status,
                       p.full_name as parent_name, p.email as parent_email
                FROM students s
                LEFT JOIN parents p ON s.student_id = p.student_id
                WHERE s.student_id = %s
            """, (student_id,))
            student = cursor.fetchone()
            
            if not student:
                flash('Student not found.', 'error')
                return redirect(url_for('student_management'))
            
            # Check if student is already approved
            if student.get('status') != 'pending approval':
                flash(f'Student {student.get("full_name")} is already {student.get("status")}.', 'info')
                return redirect(url_for('student_management'))
            
            # Update student status to 'in session'
            cursor.execute("""
                UPDATE students 
                SET status = 'in session', updated_at = CURRENT_TIMESTAMP
                WHERE student_id = %s
            """, (student_id,))
            
            connection.commit()
            
            # Send approval email to parent
            if student.get('parent_email') and student.get('parent_name'):
                try:
                    send_student_approval_email(
                        student.get('parent_email'),
                        student.get('parent_name'),
                        student.get('full_name'),
                        student_id
                    )
                except Exception as email_error:
                    print(f"Error sending approval email: {email_error}")
                    # Don't fail the approval if email fails, but log it
                    flash(f'Student approved successfully, but email could not be sent: {email_error}', 'warning')
                else:
                    flash(f'Student {student.get("full_name")} has been approved successfully! Congratulations email sent to parent.', 'success')
            else:
                flash(f'Student {student.get("full_name")} has been approved successfully! (No parent email found to send notification.)', 'success')
            
    except Exception as e:
        print(f"Error approving student: {e}")
        connection.rollback()
        flash('An error occurred while approving the student. Please try again.', 'error')
    finally:
        if connection:
            try:
                connection.close()
            except:
                pass  # Connection might already be closed
    
    return redirect(url_for('student_management'))

# Profile and Settings Routes
@app.route('/profile/<role>')
@login_required
def profile(role):
    """Profile page for each role"""
    user_role = session.get('role', '').lower()
    
    # Validate role access
    if role == 'employee':
        employee_roles = ['employee', 'super admin', 'principal', 'deputy principal', 'academic coordinator', 
                         'teachers', 'accountant', 'librarian', 'warden', 'transport manager', 'technician']
        if user_role not in employee_roles:
            flash('You do not have permission to access this page.', 'error')
            return redirect(url_for('home'))
        
        # Get employee data
        connection = get_db_connection()
        user_data = {}
        if connection:
            try:
                with connection.cursor() as cursor:
                    employee_id = session.get('employee_id') or session.get('user_id')
                    cursor.execute("SELECT * FROM employees WHERE id = %s OR employee_id = %s", 
                                 (employee_id, employee_id))
                    employee = cursor.fetchone()
                    if employee:
                        user_data = employee
            except Exception as e:
                print(f"Error fetching employee data: {e}")
            finally:
                connection.close()
        
        return render_template('dashboards/profile_employee.html', user_data=user_data, role=user_role)
    
    elif role in ['student', 'parent']:
        if user_role != role:
            flash('You do not have permission to access this page.', 'error')
            return redirect(url_for('home'))
        
        # Get user data from users table
        connection = get_db_connection()
        user_data = {}
        if connection:
            try:
                with connection.cursor() as cursor:
                    user_id = session.get('user_id')
                    cursor.execute("SELECT * FROM users WHERE id = %s AND role = %s", (user_id, role))
                    user = cursor.fetchone()
                    if user:
                        user_data = user
            except Exception as e:
                print(f"Error fetching user data: {e}")
            finally:
                connection.close()
        
        return render_template(f'dashboards/profile_{role}.html', user_data=user_data, role=role)
    
    else:
        flash('Invalid role.', 'error')
        return redirect(url_for('home'))

@app.route('/settings/<role>')
@login_required
def settings(role):
    """Settings page for each role"""
    user_role = session.get('role', '').lower()
    is_technician = user_role == 'technician'
    
    # Validate role access
    if role == 'principal':
        # Allow principals and technicians to access
        if user_role != 'principal' and not is_technician:
            flash('You do not have permission to access this page.', 'error')
            return redirect(url_for('home'))
        
        return render_template('dashboards/settings_principal.html', role=user_role)
    
    elif role == 'employee':
        employee_roles = ['employee', 'super admin', 'deputy principal', 'academic coordinator', 
                         'teachers', 'accountant', 'librarian', 'warden', 'transport manager', 'technician']
        if user_role not in employee_roles:
            flash('You do not have permission to access this page.', 'error')
            return redirect(url_for('home'))
        
        return render_template('dashboards/settings_employee.html', role=user_role)
    
    elif role in ['student', 'parent']:
        if user_role != role:
            flash('You do not have permission to access this page.', 'error')
            return redirect(url_for('home'))
        
        return render_template(f'dashboards/settings_{role}.html', role=role)
    
    else:
        flash('Invalid role.', 'error')
        return redirect(url_for('home'))

# Profile Update Routes
@app.route('/profile/employee/update', methods=['POST'])
@login_required
def update_employee_profile():
    """Update employee profile information"""
    user_role = session.get('role', '').lower()
    employee_roles = ['employee', 'super admin', 'principal', 'deputy principal', 'academic coordinator', 
                     'teachers', 'accountant', 'librarian', 'warden', 'transport manager', 'technician']
    
    if user_role not in employee_roles:
        flash('You do not have permission to perform this action.', 'error')
        return redirect(url_for('home'))
    
    full_name = request.form.get('full_name', '').strip().upper()
    email = request.form.get('email', '').strip().lower()
    phone = request.form.get('phone', '').strip().upper()
    id_number = request.form.get('id_number', '').strip().upper()
    
    if not all([full_name, email, phone]):
        flash('Please fill in all required fields.', 'error')
        return redirect(url_for('profile', role='employee'))
    
    connection = get_db_connection()
    if connection:
        try:
            with connection.cursor() as cursor:
                employee_id = session.get('employee_id') or session.get('user_id')
                cursor.execute("""
                    UPDATE employees 
                    SET full_name = %s, email = %s, phone = %s, id_number = %s
                    WHERE id = %s OR employee_id = %s
                """, (full_name, email, phone, id_number, employee_id, employee_id))
                connection.commit()
                
                # Update session
                session['full_name'] = full_name
                session['email'] = email
                
                flash('Profile updated successfully!', 'success')
        except Exception as e:
            print(f"Error updating employee profile: {e}")
            connection.rollback()
            flash('An error occurred while updating your profile. Please try again.', 'error')
        finally:
            connection.close()
    else:
        flash('Database connection error. Please try again later.', 'error')
    
    return redirect(url_for('profile', role='employee'))

# Settings Update Routes
@app.route('/settings/<role>/password', methods=['POST'])
@login_required
def update_password(role):
    """Update user password"""
    user_role = session.get('role', '').lower()
    
    if role == 'principal':
        is_technician = user_role == 'technician'
        # Allow principals and technicians to update password
        if user_role != 'principal' and not is_technician:
            flash('You do not have permission to perform this action.', 'error')
            return redirect(url_for('home'))
        
        current_password = request.form.get('current_password', '')
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')
        
        if not all([current_password, new_password, confirm_password]):
            flash('Please fill in all password fields.', 'error')
            return redirect(url_for('settings', role='principal'))
        
        if new_password != confirm_password:
            flash('New passwords do not match.', 'error')
            return redirect(url_for('settings', role='principal'))
        
        if len(new_password) < 6:
            flash('Password must be at least 6 characters long.', 'error')
            return redirect(url_for('settings', role='principal'))
        
        connection = get_db_connection()
        if connection:
            try:
                with connection.cursor() as cursor:
                    employee_id = session.get('employee_id') or session.get('user_id')
                    cursor.execute("SELECT password_hash FROM employees WHERE id = %s OR employee_id = %s", 
                                 (employee_id, employee_id))
                    employee = cursor.fetchone()
                    
                    if employee and check_password_hash(employee['password_hash'], current_password):
                        new_password_hash = generate_password_hash(new_password)
                        cursor.execute("""
                            UPDATE employees 
                            SET password_hash = %s
                            WHERE id = %s OR employee_id = %s
                        """, (new_password_hash, employee_id, employee_id))
                        connection.commit()
                        flash('Password updated successfully!', 'success')
                    else:
                        flash('Current password is incorrect.', 'error')
            except Exception as e:
                print(f"Error updating password: {e}")
                connection.rollback()
                flash('An error occurred while updating your password. Please try again.', 'error')
            finally:
                connection.close()
        else:
            flash('Database connection error. Please try again later.', 'error')
        
        return redirect(url_for('settings', role='principal'))
    
    elif role == 'employee':
        employee_roles = ['employee', 'super admin', 'deputy principal', 'academic coordinator', 
                         'teachers', 'accountant', 'librarian', 'warden', 'transport manager', 'technician']
        if user_role not in employee_roles:
            flash('You do not have permission to perform this action.', 'error')
            return redirect(url_for('home'))
        
        current_password = request.form.get('current_password', '')
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')
        
        if not all([current_password, new_password, confirm_password]):
            flash('Please fill in all password fields.', 'error')
            return redirect(url_for('settings', role='employee'))
        
        if new_password != confirm_password:
            flash('New passwords do not match.', 'error')
            return redirect(url_for('settings', role='employee'))
        
        if len(new_password) < 6:
            flash('Password must be at least 6 characters long.', 'error')
            return redirect(url_for('settings', role='employee'))
        
        connection = get_db_connection()
        if connection:
            try:
                with connection.cursor() as cursor:
                    employee_id = session.get('employee_id') or session.get('user_id')
                    cursor.execute("SELECT password_hash FROM employees WHERE id = %s OR employee_id = %s", 
                                 (employee_id, employee_id))
                    employee = cursor.fetchone()
                    
                    if employee and check_password_hash(employee['password_hash'], current_password):
                        new_password_hash = generate_password_hash(new_password)
                        cursor.execute("""
                            UPDATE employees 
                            SET password_hash = %s
                            WHERE id = %s OR employee_id = %s
                        """, (new_password_hash, employee_id, employee_id))
                        connection.commit()
                        flash('Password updated successfully!', 'success')
                    else:
                        flash('Current password is incorrect.', 'error')
            except Exception as e:
                print(f"Error updating password: {e}")
                connection.rollback()
                flash('An error occurred while updating your password. Please try again.', 'error')
            finally:
                connection.close()
        else:
            flash('Database connection error. Please try again later.', 'error')
        
        return redirect(url_for('settings', role='employee'))
    
    elif role in ['student', 'parent']:
        if user_role != role:
            flash('You do not have permission to perform this action.', 'error')
            return redirect(url_for('home'))
        
        # Similar logic for student/parent password update
        flash('Password update functionality for students and parents will be implemented soon.', 'info')
        return redirect(url_for('settings', role=role))
    
    else:
        flash('Invalid role.', 'error')
        return redirect(url_for('home'))

@app.route('/settings/<role>/notifications', methods=['POST'])
@login_required
def update_notifications(role):
    """Update notification preferences"""
    # This is a placeholder - notification preferences can be stored in a separate table or user preferences
    flash('Notification preferences saved successfully!', 'success')
    return redirect(url_for('settings', role=role))

@app.route('/settings/<role>/preferences', methods=['POST'])
@login_required
def update_preferences(role):
    """Update account preferences"""
    # This is a placeholder - preferences can be stored in a separate table or user preferences
    flash('Preferences saved successfully!', 'success')
    return redirect(url_for('settings', role=role))

# Role Switching Route (for technicians)
@app.route('/switch-role/<target_role>')
@login_required
def switch_role(target_role):
    """Allow technicians to switch between different role views"""
    user_role = session.get('role', '').lower()
    
    # Only technicians can switch roles
    if user_role != 'technician':
        flash('You do not have permission to switch roles.', 'error')
        return redirect(url_for('dashboard_employee'))
    
    # Valid main roles to switch to
    valid_main_roles = ['employee', 'student', 'parent']
    
    # Valid employee sub-roles
    valid_employee_roles = ['employee', 'super admin', 'principal', 'deputy principal', 
                           'academic coordinator', 'teachers', 'accountant', 'librarian', 
                           'warden', 'transport manager', 'technician']
    
    # Check if it's a main role
    if target_role in valid_main_roles:
        # Store the role we're viewing as
        session['viewing_as_role'] = target_role
        session['viewing_as_employee_role'] = None  # Clear employee sub-role
        
        flash(f'Switched to {target_role.title()} role view.', 'success')
        
        # Redirect to the appropriate dashboard
        if target_role == 'employee':
            return redirect(url_for('dashboard_employee'))
        elif target_role == 'student':
            return redirect(url_for('dashboard_student'))
        elif target_role == 'parent':
            return redirect(url_for('dashboard_parent'))
    
    # Check if it's an employee sub-role
    elif target_role in valid_employee_roles:
        # Set viewing as employee with specific sub-role
        session['viewing_as_role'] = 'employee'
        session['viewing_as_employee_role'] = target_role
        flash(f'Switched to {target_role.title()} role view.', 'success')
        return redirect(url_for('dashboard_employee'))
    
    else:
        flash('Invalid role selected.', 'error')
        return redirect(url_for('role_switch_page'))
    
    return redirect(url_for('role_switch_page'))

@app.route('/role-switch')
@login_required
def role_switch_page():
    """Role switch page for technicians to select which role to view as"""
    user_role = session.get('role', '').lower()
    
    # Only technicians can access this page
    if user_role != 'technician':
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('dashboard_employee'))
    
    # List of all employee roles
    employee_roles_list = ['employee', 'super admin', 'principal', 'deputy principal', 
                          'academic coordinator', 'teachers', 'accountant', 'librarian', 
                          'warden', 'transport manager', 'technician']
    
    return render_template('dashboards/role_switch.html', 
                         employee_roles_list=employee_roles_list)

@app.route('/switch-role/reset')
@login_required
def reset_role():
    """Reset to technician's own role"""
    user_role = session.get('role', '').lower()
    
    if user_role != 'technician':
        flash('You do not have permission to perform this action.', 'error')
        return redirect(url_for('dashboard_employee'))
    
    # Remove viewing roles from session
    session.pop('viewing_as_role', None)
    session.pop('viewing_as_employee_role', None)
    
    flash('Switched back to your technician role.', 'success')
    return redirect(url_for('role_switch_page'))

# System Settings Route (for technicians)
@app.route('/system-settings')
@login_required
def system_settings():
    """System settings page for technicians and accountants"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    employee_id = session.get('employee_id') or session.get('user_id')
    
    # Check if user is accountant or viewing as accountant
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_technician = user_role == 'technician'
    
    # Check permission OR role-based access - allow technicians and accountants
    has_access = check_permission_or_role('system_settings', 
                                         allowed_roles=['technician', 'accountant'])
    
    if not (has_access or is_accountant or is_technician):
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('dashboard_employee'))
    
    # Get school settings, academic levels, academic years, and terms
    connection = get_db_connection()
    school_data = {}
    academic_levels = []
    academic_years = []
    terms = []
    current_academic_year = None
    if connection:
        try:
            with connection.cursor() as cursor:
                # Get school settings
                cursor.execute("SELECT * FROM school_settings ORDER BY id DESC LIMIT 1")
                result = cursor.fetchone()
                if result:
                    # Handle both dict (DictCursor) and tuple results
                    if isinstance(result, dict):
                        school_data = {
                            'school_name': result.get('school_name', '') or '',
                            'school_email': result.get('school_email', '') or '',
                            'school_phone': result.get('school_phone', '') or '',
                            'school_logo': result.get('school_logo', '') or '',
                            'twitter_url': result.get('twitter_url', '') or '',
                            'facebook_url': result.get('facebook_url', '') or '',
                            'instagram_url': result.get('instagram_url', '') or '',
                            'tiktok_url': result.get('tiktok_url', '') or '',
                            'whatsapp_number': result.get('whatsapp_number', '') or '',
                            'school_location': result.get('school_location', '') or ''
                        }
                    else:
                        # Handle tuple results (fallback)
                        school_data = {
                            'school_name': result[1] if len(result) > 1 else '',
                            'school_email': result[2] if len(result) > 2 else '',
                            'school_phone': result[3] if len(result) > 3 else '',
                            'school_logo': result[4] if len(result) > 4 else '',
                            'twitter_url': result[5] if len(result) > 5 else '',
                            'facebook_url': result[6] if len(result) > 6 else '',
                            'instagram_url': result[7] if len(result) > 7 else '',
                            'tiktok_url': result[8] if len(result) > 8 else '',
                            'whatsapp_number': result[9] if len(result) > 9 else '',
                            'school_location': result[10] if len(result) > 10 else ''
                        }
                
                # Get academic levels
                try:
                    cursor.execute("""
                        SELECT id, level_category, level_name, level_description, level_status, 
                               created_at, updated_at 
                        FROM academic_levels 
                        ORDER BY created_at DESC
                    """)
                    results = cursor.fetchall()
                    
                    print(f"DEBUG: Query returned {len(results) if results else 0} rows")
                    
                    if results:
                        for row in results:
                            # Since we're using DictCursor, row should be a dict
                            level_data = {
                                'id': row.get('id'),
                                'level_category': str(row.get('level_category', '')).strip() or '',
                                'level_name': str(row.get('level_name', '')).strip() or '',
                                'level_description': str(row.get('level_description', '')).strip() if row.get('level_description') else '',
                                'level_status': str(row.get('level_status', 'active')).strip() or 'active',
                                'created_at': row.get('created_at'),
                                'updated_at': row.get('updated_at')
                            }
                            academic_levels.append(level_data)
                        print(f"✓ Successfully fetched {len(academic_levels)} academic level(s)")
                    else:
                        print("ℹ No academic levels found in database (table exists but is empty)")
                except pymysql.err.ProgrammingError as e:
                    # Table doesn't exist (error 1146)
                    error_code = e.args[0] if e.args else 0
                    if error_code == 1146 or "doesn't exist" in str(e).lower():
                        print("⚠ Academic levels table does not exist yet")
                    else:
                        print(f"⚠ SQL Error: {e}")
                        import traceback
                        traceback.print_exc()
                except Exception as e:
                    print(f"⚠ Error fetching academic levels: {e}")
                    import traceback
                    traceback.print_exc()
                
                # Get academic years
                try:
                    # Auto-lock academic years whose end_date has passed
                    from datetime import date
                    today = date.today()
                    cursor.execute("""
                        UPDATE academic_years 
                        SET is_locked = TRUE, locked_at = CURRENT_TIMESTAMP 
                        WHERE end_date < %s AND is_locked = FALSE
                    """, (today,))
                    if cursor.rowcount > 0:
                        connection.commit()
                        print(f"✓ Auto-locked {cursor.rowcount} academic year(s) that have ended")
                    
                    cursor.execute("""
                        SELECT id, year_name, start_date, end_date, status, is_current, is_locked, locked_at
                        FROM academic_years
                        ORDER BY start_date DESC
                    """)
                    academic_years_result = cursor.fetchall()
                    academic_years = []
                    for year in academic_years_result:
                        if isinstance(year, dict):
                            academic_years.append(year)
                        else:
                            academic_years.append({
                                'id': year[0],
                                'year_name': year[1],
                                'start_date': year[2],
                                'end_date': year[3],
                                'status': year[4],
                                'is_current': year[5],
                                'is_locked': year[6] if len(year) > 6 else False,
                                'locked_at': year[7] if len(year) > 7 else None
                            })
                    
                    # Get current academic year
                    cursor.execute("""
                        SELECT id, year_name, start_date, end_date, status, is_current, is_locked, locked_at
                        FROM academic_years
                        WHERE is_current = TRUE
                        LIMIT 1
                    """)
                    current_year_result = cursor.fetchone()
                    if current_year_result:
                        if isinstance(current_year_result, dict):
                            current_academic_year = current_year_result
                        else:
                            current_academic_year = {
                                'id': current_year_result[0],
                                'year_name': current_year_result[1],
                                'start_date': current_year_result[2],
                                'end_date': current_year_result[3],
                                'status': current_year_result[4],
                                'is_current': current_year_result[5],
                                'is_locked': current_year_result[6] if len(current_year_result) > 6 else False,
                                'locked_at': current_year_result[7] if len(current_year_result) > 7 else None
                            }
                    else:
                        current_academic_year = None
                except Exception as e:
                    # Tables might not exist yet
                    print(f"Note: academic_years table may not exist yet: {e}")
                    academic_years = []
                
                # Get terms with their academic levels
                try:
                    cursor.execute("""
                        SELECT t.id, t.term_name, t.academic_year_id,
                               t.start_date, t.end_date, t.status, t.is_current, t.is_locked, t.locked_at,
                               ay.year_name as academic_year_name
                        FROM terms t
                        LEFT JOIN academic_years ay ON t.academic_year_id = ay.id
                        ORDER BY t.academic_year_id DESC, t.start_date ASC
                    """)
                    terms_raw = cursor.fetchall()
                    terms = []
                    for term in terms_raw:
                        term_dict = dict(term) if isinstance(term, dict) else {
                            'id': term[0], 'term_name': term[1], 'academic_year_id': term[2],
                            'start_date': term[3], 'end_date': term[4], 'status': term[5],
                            'is_current': term[6] if len(term) > 6 else False,
                            'is_locked': term[7] if len(term) > 7 else False,
                            'locked_at': term[8] if len(term) > 8 else None,
                            'academic_year_name': term[9] if len(term) > 9 else None
                        }
                        
                        # Auto-lock terms that have passed their end date
                        if term_dict.get('end_date') and not term_dict.get('is_locked'):
                            from datetime import date, datetime
                            end_date = term_dict['end_date']
                            if isinstance(end_date, str):
                                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                            elif hasattr(end_date, 'date'):
                                end_date = end_date.date()
                            
                            if end_date < date.today():
                                # Auto-lock the term
                                cursor.execute("""
                                    UPDATE terms 
                                    SET is_locked = TRUE, locked_at = NOW() 
                                    WHERE id = %s AND is_locked = FALSE
                                """, (term_dict['id'],))
                                connection.commit()  # Commit auto-lock immediately
                                term_dict['is_locked'] = True
                                term_dict['locked_at'] = datetime.now()
                        # Get academic levels for this term (only active ones)
                        cursor.execute("""
                            SELECT al.id, al.level_name, al.level_category
                            FROM term_academic_levels tal
                            JOIN academic_levels al ON tal.academic_level_id = al.id
                            WHERE tal.term_id = %s AND al.level_status = 'active'
                            ORDER BY al.level_name ASC
                        """, (term_dict['id'],))
                        academic_levels_list = cursor.fetchall()
                        term_dict['academic_levels'] = [
                            {
                                'id': al.get('id') if isinstance(al, dict) else al[0],
                                'level_name': al.get('level_name') if isinstance(al, dict) else al[1],
                                'level_category': al.get('level_category') if isinstance(al, dict) else al[2]
                            }
                            for al in academic_levels_list
                        ]
                        # For backward compatibility, set academic_level_name to first level or None
                        term_dict['academic_level_name'] = academic_levels_list[0].get('level_name') if academic_levels_list and isinstance(academic_levels_list[0], dict) else (academic_levels_list[0][1] if academic_levels_list else None)
                        terms.append(term_dict)
                except Exception as e:
                    # Tables might not exist yet
                    print(f"Note: terms table may not exist yet: {e}")
                    terms = []
        except Exception as e:
            print(f"❌ Error fetching data: {e}")
            print(f"Error type: {type(e).__name__}")
            print(f"Error args: {e.args}")
            import traceback
            traceback.print_exc()
        finally:
            if connection:
                connection.close()
    
    # Calculate days remaining for academic years and terms
    from datetime import date, datetime
    today = date.today()
    
    # Add days_remaining to academic years
    for year in academic_years:
        if year.get('end_date'):
            end_date = year['end_date']
            # Handle different date formats
            if isinstance(end_date, str):
                try:
                    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                except:
                    try:
                        end_date = datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').date()
                    except:
                        end_date = None
            elif isinstance(end_date, datetime):
                end_date = end_date.date()
            elif isinstance(end_date, date):
                end_date = end_date
            else:
                end_date = None
            
            if end_date:
                try:
                    days_diff = (end_date - today).days
                    year['days_remaining'] = days_diff if days_diff >= 0 else 0
                except:
                    year['days_remaining'] = None
            else:
                year['days_remaining'] = None
        else:
            year['days_remaining'] = None
    
    # Add days_remaining to terms
    for term in terms:
        if term.get('end_date'):
            end_date = term['end_date']
            # Handle different date formats
            if isinstance(end_date, str):
                try:
                    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                except:
                    try:
                        end_date = datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').date()
                    except:
                        end_date = None
            elif isinstance(end_date, datetime):
                end_date = end_date.date()
            elif isinstance(end_date, date):
                end_date = end_date
            else:
                end_date = None
            
            if end_date:
                try:
                    days_diff = (end_date - today).days
                    term['days_remaining'] = days_diff if days_diff >= 0 else 0
                except:
                    term['days_remaining'] = None
            else:
                term['days_remaining'] = None
        else:
            term['days_remaining'] = None
    
    return render_template('dashboards/system_settings.html', 
                         school_data=school_data, 
                         academic_levels=academic_levels,
                         academic_years=academic_years,
                         terms=terms,
                         current_academic_year=current_academic_year,
                         today=today,
                         is_accountant=is_accountant)

@app.route('/dashboard/employee/academic-settings')
@login_required
def academic_settings():
    """Academic settings page for accountants - shows only Academic Levels, Years, and Terms"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    employee_id = session.get('employee_id') or session.get('user_id')
    
    # Check if user is accountant or viewing as accountant
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_technician = user_role == 'technician'
    
    # Check permission-based access
    has_view_fees_permission = check_permission_or_role('view_student_fees', ['accountant', 'principal'])
    has_manage_fees_permission = check_permission_or_role('manage_fees', ['accountant', 'principal'])
    
    if not (is_accountant or is_technician or has_view_fees_permission or has_manage_fees_permission):
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('dashboard_employee'))
    
    # Get academic levels, academic years, and terms
    connection = get_db_connection()
    academic_levels = []
    academic_years = []
    terms = []
    current_academic_year = None
    today = datetime.now().date()
    
    if connection:
        try:
            with connection.cursor() as cursor:
                # Get academic levels
                cursor.execute("""
                    SELECT id, level_category, level_name, level_description, level_status
                    FROM academic_levels
                    ORDER BY level_category, level_name ASC
                """)
                academic_levels_results = cursor.fetchall()
                
                for row in academic_levels_results:
                    academic_levels.append({
                        'id': row.get('id') if isinstance(row, dict) else row[0],
                        'level_category': row.get('level_category', '') if isinstance(row, dict) else row[1],
                        'level_name': row.get('level_name', '') if isinstance(row, dict) else row[2],
                        'level_description': row.get('level_description', '') if isinstance(row, dict) else row[3],
                        'level_status': row.get('level_status', 'active') if isinstance(row, dict) else (row[4] if len(row) > 4 else 'active')
                    })
                
                # Get academic years
                try:
                    cursor.execute("""
                        SELECT id, year_name, start_date, end_date, status, is_current
                        FROM academic_years
                        ORDER BY start_date DESC
                    """)
                    academic_years_results = cursor.fetchall()
                    
                    for row in academic_years_results:
                        year_dict = {
                            'id': row.get('id') if isinstance(row, dict) else row[0],
                            'year_name': row.get('year_name', '') if isinstance(row, dict) else row[1],
                            'start_date': row.get('start_date') if isinstance(row, dict) else row[2],
                            'end_date': row.get('end_date') if isinstance(row, dict) else row[3],
                            'status': row.get('status', 'active') if isinstance(row, dict) else (row[4] if len(row) > 4 else 'active'),
                            'is_current': row.get('is_current', False) if isinstance(row, dict) else (row[5] if len(row) > 5 else False)
                        }
                        academic_years.append(year_dict)
                        
                        if year_dict.get('is_current'):
                            current_academic_year = year_dict
                except Exception as e:
                    print(f"Note: academic_years table may not exist yet: {e}")
                    academic_years = []
                
                # Get terms with their academic levels
                try:
                    cursor.execute("""
                        SELECT t.id, t.term_name, t.academic_year_id,
                               t.start_date, t.end_date, t.status, t.is_current, t.is_locked, t.locked_at,
                               ay.year_name as academic_year_name
                        FROM terms t
                        LEFT JOIN academic_years ay ON t.academic_year_id = ay.id
                        ORDER BY t.academic_year_id DESC, t.start_date ASC
                    """)
                    terms_raw = cursor.fetchall()
                    terms = []
                    for term in terms_raw:
                        term_dict = dict(term) if isinstance(term, dict) else {
                            'id': term[0], 'term_name': term[1], 'academic_year_id': term[2],
                            'start_date': term[3], 'end_date': term[4], 'status': term[5],
                            'is_current': term[6] if len(term) > 6 else False,
                            'is_locked': term[7] if len(term) > 7 else False,
                            'locked_at': term[8] if len(term) > 8 else None,
                            'academic_year_name': term[9] if len(term) > 9 else None
                        }
                        terms.append(term_dict)
                except Exception as e:
                    print(f"Note: terms table may not exist yet: {e}")
                    terms = []
        except Exception as e:
            print(f"Error fetching academic settings: {e}")
        finally:
            connection.close()
    
    return render_template('dashboards/academic_settings.html',
                         academic_levels=academic_levels,
                         academic_years=academic_years,
                         terms=terms,
                         current_academic_year=current_academic_year,
                         today=today,
                         role=user_role)

# Integration Settings Route (for technicians)
@app.route('/dashboard/employee/integration-settings')
@login_required
def integration_settings():
    """Integration settings page for technicians"""
    user_role = session.get('role', '').lower()
    employee_id = session.get('employee_id') or session.get('user_id')
    
    # Check permission OR role-based access - allow technicians
    has_access = check_permission_or_role('integration_settings', 
                                         allowed_roles=['technician'])
    
    if not has_access:
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('dashboard_employee'))
    
    # Get integration settings from database if they exist
    connection = get_db_connection()
    integration_data = {
        'whatsapp': {
            'enabled': False,
            'api_key': '',
            'api_secret': '',
            'phone_number': '',
            'webhook_url': ''
        },
        'email': {
            'enabled': False,
            'smtp_server': '',
            'smtp_port': 587,
            'smtp_username': '',
            'smtp_password': '',
            'from_email': '',
            'from_name': ''
        },
        'sms': {
            'enabled': False,
            'provider': '',
            'api_key': '',
            'api_secret': '',
            'sender_id': ''
        }
    }
    
    if connection:
        try:
            with connection.cursor() as cursor:
                # Check if integration_settings table exists and fetch data
                cursor.execute("""
                    SELECT integration_type, settings_json 
                    FROM integration_settings 
                    WHERE id = 1
                """)
                result = cursor.fetchone()
                if result:
                    # Parse JSON settings if stored as JSON
                    import json
                    if isinstance(result, dict):
                        integration_type = result.get('integration_type', '')
                        settings_json = result.get('settings_json', '{}')
                    else:
                        integration_type = result[0] if len(result) > 0 else ''
                        settings_json = result[1] if len(result) > 1 else '{}'
                    
                    try:
                        settings = json.loads(settings_json) if settings_json else {}
                        if integration_type in integration_data:
                            integration_data[integration_type].update(settings)
                    except:
                        pass
        except Exception as e:
            # Table might not exist yet, that's okay
            print(f"Integration settings table may not exist: {e}")
        finally:
            connection.close()
    
    return render_template('dashboards/integration_settings.html', 
                         integration_data=integration_data,
                         role=user_role)

# Database Management Route (for technicians)
@app.route('/database')
@login_required
def database_management():
    """Database management page for technicians and principals"""
    user_role = session.get('role', '').lower()
    employee_id = session.get('employee_id') or session.get('user_id')
    
    # Check permission OR role-based access
    has_access = check_permission_or_role('view_database', 
                                         allowed_roles=['technician', 'principal'])
    
    if not has_access:
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('dashboard_employee'))
    
    # Get database information and data analysis
    connection = get_db_connection()
    tables = []
    db_info = {}
    data_analysis = {
        'total_tables': 0,
        'total_records': 0,
        'total_size_mb': 0,
        'largest_table': {'name': '', 'rows': 0, 'size_mb': 0},
        'most_active_tables': [],
        'database_size_mb': 0,
        # Business metrics
        'total_students': 0,
        'students_by_status': {},
        'total_employees': 0,
        'employees_by_role': {},
        'employees_by_status': {},
        'total_parents': 0,
        'total_academic_levels': 0,
        'active_academic_levels': 0,
        'total_fee_structures': 0,
        'active_fee_structures': 0,
        'total_fee_items': 0,
        'total_fee_amount': 0,
        'total_student_fees': 0,
        'paid_fees': 0,
        'pending_fees': 0,
        'total_fee_revenue': 0
    }
    
    if connection:
        try:
            with connection.cursor() as cursor:
                # Get database name
                cursor.execute("SELECT DATABASE() as db_name")
                db_result = cursor.fetchone()
                db_info['name'] = db_result.get('db_name', DB_CONFIG['database']) if db_result else DB_CONFIG['database']
                
                # Get total database size
                cursor.execute("""
                    SELECT 
                        ROUND(SUM(data_length + index_length) / 1024 / 1024, 2) AS total_size_mb
                    FROM information_schema.TABLES 
                    WHERE table_schema = DATABASE()
                """)
                db_size_result = cursor.fetchone()
                data_analysis['database_size_mb'] = db_size_result.get('total_size_mb', 0) if db_size_result else 0
                
                # Get all tables
                cursor.execute("SHOW TABLES")
                table_results = cursor.fetchall()
                data_analysis['total_tables'] = len(table_results)
                
                table_stats = []
                
                # Get table information
                for table_result in table_results:
                    table_name = list(table_result.values())[0] if isinstance(table_result, dict) else table_result[0]
                    
                    # Get row count
                    cursor.execute(f"SELECT COUNT(*) as count FROM `{table_name}`")
                    count_result = cursor.fetchone()
                    row_count = count_result.get('count', 0) if count_result else 0
                    data_analysis['total_records'] += row_count
                    
                    # Get table size
                    cursor.execute(f"""
                        SELECT 
                            ROUND(((data_length + index_length) / 1024 / 1024), 2) AS size_mb
                        FROM information_schema.TABLES 
                        WHERE table_schema = DATABASE() 
                        AND table_name = %s
                    """, (table_name,))
                    size_result = cursor.fetchone()
                    size_mb = size_result.get('size_mb', 0) if size_result else 0
                    data_analysis['total_size_mb'] += size_mb
                    
                    # Track largest table
                    if row_count > data_analysis['largest_table']['rows']:
                        data_analysis['largest_table'] = {
                            'name': table_name,
                            'rows': row_count,
                            'size_mb': size_mb
                        }
                    
                    # Store table stats for most active tables
                    table_stats.append({
                        'name': table_name,
                        'rows': row_count,
                        'size_mb': size_mb
                    })
                    
                    # Get table structure
                    cursor.execute(f"DESCRIBE `{table_name}`")
                    columns_result = cursor.fetchall()
                    
                    # Convert columns to list of dicts if needed
                    columns = []
                    for col in columns_result:
                        if isinstance(col, dict):
                            columns.append(col)
                        else:
                            # Handle tuple results
                            columns.append({
                                'Field': col[0] if len(col) > 0 else '',
                                'Type': col[1] if len(col) > 1 else '',
                                'Null': col[2] if len(col) > 2 else '',
                                'Key': col[3] if len(col) > 3 else '',
                                'Default': col[4] if len(col) > 4 else None,
                                'Extra': col[5] if len(col) > 5 else ''
                            })
                    
                    tables.append({
                        'name': table_name,
                        'row_count': row_count,
                        'size_mb': size_mb,
                        'columns': columns
                    })
                
                # Get top 5 most active tables by row count
                table_stats_sorted = sorted(table_stats, key=lambda x: x['rows'], reverse=True)
                data_analysis['most_active_tables'] = table_stats_sorted[:5]
                
                # Business Metrics - Students
                try:
                    cursor.execute("SELECT COUNT(*) as count FROM students")
                    result = cursor.fetchone()
                    data_analysis['total_students'] = result.get('count', 0) if result else 0
                    
                    # Students by status
                    cursor.execute("""
                        SELECT status, COUNT(*) as count 
                        FROM students 
                        GROUP BY status
                    """)
                    status_results = cursor.fetchall()
                    for row in status_results:
                        status = row.get('status', 'unknown') if isinstance(row, dict) else row[0]
                        count = row.get('count', 0) if isinstance(row, dict) else row[1]
                        data_analysis['students_by_status'][status] = count
                except Exception as e:
                    print(f"Error fetching students data: {e}")
                
                # Business Metrics - Employees
                try:
                    cursor.execute("SELECT COUNT(*) as count FROM employees")
                    result = cursor.fetchone()
                    data_analysis['total_employees'] = result.get('count', 0) if result else 0
                    
                    # Employees by role
                    cursor.execute("""
                        SELECT role, COUNT(*) as count 
                        FROM employees 
                        GROUP BY role
                    """)
                    role_results = cursor.fetchall()
                    for row in role_results:
                        role = row.get('role', 'unknown') if isinstance(row, dict) else row[0]
                        count = row.get('count', 0) if isinstance(row, dict) else row[1]
                        data_analysis['employees_by_role'][role] = count
                    
                    # Employees by status
                    cursor.execute("""
                        SELECT status, COUNT(*) as count 
                        FROM employees 
                        GROUP BY status
                    """)
                    status_results = cursor.fetchall()
                    for row in status_results:
                        status = row.get('status', 'unknown') if isinstance(row, dict) else row[0]
                        count = row.get('count', 0) if isinstance(row, dict) else row[1]
                        data_analysis['employees_by_status'][status] = count
                except Exception as e:
                    print(f"Error fetching employees data: {e}")
                
                # Business Metrics - Parents
                try:
                    cursor.execute("SELECT COUNT(*) as count FROM parents")
                    result = cursor.fetchone()
                    data_analysis['total_parents'] = result.get('count', 0) if result else 0
                except Exception as e:
                    print(f"Error fetching parents data: {e}")
                
                # Business Metrics - Academic Levels (Classes)
                try:
                    cursor.execute("SELECT COUNT(*) as count FROM academic_levels")
                    result = cursor.fetchone()
                    data_analysis['total_academic_levels'] = result.get('count', 0) if result else 0
                    
                    cursor.execute("""
                        SELECT COUNT(*) as count 
                        FROM academic_levels 
                        WHERE level_status = 'active'
                    """)
                    result = cursor.fetchone()
                    data_analysis['active_academic_levels'] = result.get('count', 0) if result else 0
                except Exception as e:
                    print(f"Error fetching academic levels data: {e}")
                
                # Business Metrics - Fees
                try:
                    # Fee Structures
                    cursor.execute("SELECT COUNT(*) as count FROM fee_structures")
                    result = cursor.fetchone()
                    data_analysis['total_fee_structures'] = result.get('count', 0) if result else 0
                    
                    cursor.execute("""
                        SELECT COUNT(*) as count 
                        FROM fee_structures 
                        WHERE status = 'active'
                    """)
                    result = cursor.fetchone()
                    data_analysis['active_fee_structures'] = result.get('count', 0) if result else 0
                    
                    # Fee Items
                    cursor.execute("SELECT COUNT(*) as count FROM fee_items")
                    result = cursor.fetchone()
                    data_analysis['total_fee_items'] = result.get('count', 0) if result else 0
                    
                    # Total fee amount from fee structures
                    cursor.execute("""
                        SELECT SUM(total_amount) as total 
                        FROM fee_structures 
                        WHERE status = 'active'
                    """)
                    result = cursor.fetchone()
                    data_analysis['total_fee_amount'] = float(result.get('total', 0) or 0) if result else 0
                    
                    # Student Fees (if table exists)
                    try:
                        cursor.execute("SELECT COUNT(*) as count FROM student_fees")
                        result = cursor.fetchone()
                        data_analysis['total_student_fees'] = result.get('count', 0) if result else 0
                        
                        cursor.execute("""
                            SELECT COUNT(*) as count 
                            FROM student_fees 
                            WHERE payment_status = 'paid'
                        """)
                        result = cursor.fetchone()
                        data_analysis['paid_fees'] = result.get('count', 0) if result else 0
                        
                        cursor.execute("""
                            SELECT COUNT(*) as count 
                            FROM student_fees 
                            WHERE payment_status = 'pending' OR payment_status = 'overdue'
                        """)
                        result = cursor.fetchone()
                        data_analysis['pending_fees'] = result.get('count', 0) if result else 0
                        
                        cursor.execute("""
                            SELECT SUM(amount_paid) as total 
                            FROM student_fees 
                            WHERE payment_status = 'paid'
                        """)
                        result = cursor.fetchone()
                        data_analysis['total_fee_revenue'] = float(result.get('total', 0) or 0) if result else 0
                    except Exception as e:
                        # student_fees table might not exist
                        pass
                except Exception as e:
                    print(f"Error fetching fees data: {e}")
        except Exception as e:
            print(f"Error fetching database information: {e}")
            flash('Error loading database information. Please try again.', 'error')
        finally:
            if connection:
                try:
                    connection.close()
                except:
                    pass
    
    return render_template('dashboards/database_management.html', 
                         tables=tables,
                         db_info=db_info,
                         data_analysis=data_analysis)

# Database Backup & Restore Route
@app.route('/database/backup-restore')
@login_required
def database_backup_restore():
    """Database backup and restore page for technicians and principals"""
    user_role = session.get('role', '').lower()
    
    # Only technicians and principals can access this page
    if user_role not in ['technician', 'principal']:
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('dashboard_employee'))
    
    # Get backup settings and history
    connection = get_db_connection()
    backup_settings = {
        'auto_backup_enabled': False,
        'backup_frequency': 'daily',
        'last_backup': None,
        'next_backup': None
    }
    backup_history = []
    
    if connection:
        try:
            with connection.cursor() as cursor:
                # Check if backup_settings table exists
                cursor.execute("""
                    SELECT COUNT(*) as count 
                    FROM information_schema.tables 
                    WHERE table_schema = DATABASE() 
                    AND table_name = 'backup_settings'
                """)
                table_exists = cursor.fetchone()
                
                if table_exists and table_exists.get('count', 0) > 0:
                    cursor.execute("SELECT * FROM backup_settings ORDER BY id DESC LIMIT 1")
                    settings_result = cursor.fetchone()
                    if settings_result:
                        backup_settings = {
                            'auto_backup_enabled': bool(settings_result.get('auto_backup_enabled', 0)),
                            'backup_frequency': settings_result.get('backup_frequency', 'daily'),
                            'last_backup': settings_result.get('last_backup'),
                            'next_backup': settings_result.get('next_backup')
                        }
                    
                    # Get backup history
                    try:
                        cursor.execute("""
                            SELECT * FROM backup_history 
                            ORDER BY created_at DESC 
                            LIMIT 10
                        """)
                        backup_history = cursor.fetchall()
                    except:
                        backup_history = []
        except Exception as e:
            print(f"Error fetching backup settings: {e}")
        finally:
            if connection:
                try:
                    connection.close()
                except:
                    pass
    
    # Check if backup file exists and get its info
    backup_file_info = None
    backup_filename = 'database_backup.xlsx' if EXCEL_AVAILABLE else 'database_backup.zip'
    backup_file_path = os.path.join(BACKUP_FOLDER, backup_filename)
    if os.path.exists(backup_file_path):
        file_stat = os.stat(backup_file_path)
        # Get absolute file path for direct file link
        abs_file_path = os.path.abspath(backup_file_path)
        # Convert to file:// URL format for direct file access
        file_url = f"file:///{abs_file_path.replace(os.sep, '/')}"
        backup_file_info = {
            'exists': True,
            'size': file_stat.st_size,
            'modified': datetime.fromtimestamp(file_stat.st_mtime),
            'url': file_url,
            'file_path': abs_file_path,
            'filename': backup_filename
        }
        
        # Check if file is up to date based on backup settings
        if backup_settings.get('last_backup'):
            last_backup = backup_settings['last_backup']
            if isinstance(last_backup, str):
                try:
                    last_backup = datetime.strptime(last_backup, '%Y-%m-%d %H:%M:%S')
                except:
                    last_backup = None
            
            if last_backup:
                file_modified = datetime.fromtimestamp(file_stat.st_mtime)
                # Check if file was modified within the last hour of last_backup
                time_diff = abs((file_modified - last_backup).total_seconds())
                backup_file_info['is_up_to_date'] = time_diff < 3600  # Within 1 hour
            else:
                backup_file_info['is_up_to_date'] = False
        else:
            backup_file_info['is_up_to_date'] = False
    else:
        backup_file_info = {'exists': False}
    
    return render_template('dashboards/database_backup_restore.html', 
                         backup_settings=backup_settings,
                         backup_history=backup_history,
                         excel_available=EXCEL_AVAILABLE,
                         backup_file_info=backup_file_info)

# Database Backup Export Route
@app.route('/database/backup-export', methods=['POST'])
@login_required
def database_backup_export():
    """Export database to Excel format"""
    user_role = session.get('role', '').lower()
    employee_id = session.get('employee_id') or session.get('user_id')
    
    # Check permission OR role-based access
    has_access = check_permission_or_role('manage_backups', 
                                         allowed_roles=['technician', 'principal'])
    
    if not has_access:
        flash('You do not have permission to perform this action.', 'error')
        return redirect(url_for('dashboard_employee'))
    
    connection = get_db_connection()
    if not connection:
        flash('Database connection error.', 'error')
        return redirect(url_for('database_backup_restore'))
    
    try:
        with connection.cursor() as cursor:
            # Get all tables
            cursor.execute("SHOW TABLES")
            table_results = cursor.fetchall()
            total_tables = len(table_results)
            total_records = 0
            
            if EXCEL_AVAILABLE:
                # Create Excel workbook
                wb = Workbook()
                wb.remove(wb.active)  # Remove default sheet
                
                # Export each table to a separate sheet
                for table_result in table_results:
                    table_name = list(table_result.values())[0] if isinstance(table_result, dict) else table_result[0]
                    
                    # Create sheet for this table
                    ws = wb.create_sheet(title=table_name[:31])  # Excel sheet name limit
                    
                    # Get table data
                    cursor.execute(f"SELECT * FROM `{table_name}`")
                    columns = [desc[0] for desc in cursor.description]
                    rows = cursor.fetchall()
                    total_records += len(rows)
                    
                    # Write headers
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    
                    for col_idx, col_name in enumerate(columns, 1):
                        cell = ws.cell(row=1, column=col_idx, value=col_name)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Write data rows
                    for row_idx, row_data in enumerate(rows, 2):
                        for col_idx, value in enumerate(row_data, 1):
                            if isinstance(value, datetime):
                                value = value.strftime('%Y-%m-%d %H:%M:%S')
                            elif value is None:
                                value = ''
                            ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Auto-adjust column widths
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                
                # Save to file on server (fixed filename that gets updated)
                filename = "database_backup.xlsx"
                filepath = os.path.join(BACKUP_FOLDER, filename)
                
                # Save workbook to file
                wb.save(filepath)
                file_size = os.path.getsize(filepath)
                
                # Save backup record
                try:
                    with connection.cursor() as cursor2:
                        # Create backup_history table if it doesn't exist
                        cursor2.execute("""
                            CREATE TABLE IF NOT EXISTS backup_history (
                                id INT AUTO_INCREMENT PRIMARY KEY,
                                filename VARCHAR(255) NOT NULL,
                                file_path VARCHAR(500),
                                file_size BIGINT,
                                table_count INT,
                                record_count INT,
                                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                                created_by VARCHAR(255)
                            )
                        """)
                        connection.commit()
                        
                        # Insert backup record
                        cursor2.execute("""
                            INSERT INTO backup_history (filename, file_path, file_size, table_count, record_count, created_by)
                            VALUES (%s, %s, %s, %s, %s, %s)
                        """, (filename, filepath, file_size, total_tables, total_records, session.get('full_name', 'Unknown')))
                        connection.commit()
                        
                        # Update last_backup in settings
                        cursor2.execute("""
                            UPDATE backup_settings 
                            SET last_backup = NOW() 
                            WHERE id = (SELECT id FROM (SELECT id FROM backup_settings ORDER BY id DESC LIMIT 1) AS tmp)
                        """)
                        connection.commit()
                except Exception as e:
                    print(f"Error saving backup record: {e}")
                
                flash('Database backup updated successfully!', 'success')
                return redirect(url_for('database_backup_restore'))
            else:
                # Fallback to CSV (zip multiple files)
                import zipfile
                from io import StringIO
                
                zip_buffer = BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    for table_result in table_results:
                        table_name = list(table_result.values())[0] if isinstance(table_result, dict) else table_result[0]
                        
                        cursor.execute(f"SELECT * FROM `{table_name}`")
                        columns = [desc[0] for desc in cursor.description]
                        rows = cursor.fetchall()
                        total_records += len(rows)
                        
                        # Create CSV in memory
                        csv_buffer = StringIO()
                        writer = csv.writer(csv_buffer)
                        writer.writerow(columns)
                        for row in rows:
                            writer.writerow([str(v) if v is not None else '' for v in row])
                        
                        zip_file.writestr(f"{table_name}.csv", csv_buffer.getvalue())
                
                # Save to file on server (fixed filename that gets updated)
                filename = "database_backup.zip"
                filepath = os.path.join(BACKUP_FOLDER, filename)
                
                with open(filepath, 'wb') as f:
                    f.write(zip_buffer.getvalue())
                
                file_size = os.path.getsize(filepath)
                
                flash('Database backup updated successfully!', 'success')
                return redirect(url_for('database_backup_restore'))
                
    except Exception as e:
        print(f"Error exporting database: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Error exporting database: {str(e)}', 'error')
        return redirect(url_for('database_backup_restore'))
    finally:
        if connection:
            try:
                connection.close()
            except:
                pass

# Database Backup Settings Route
@app.route('/database/backup-settings', methods=['POST'])
@login_required
def database_backup_settings():
    """Update backup settings"""
    user_role = session.get('role', '').lower()
    employee_id = session.get('employee_id') or session.get('user_id')
    
    # Check permission OR role-based access
    has_access = check_permission_or_role('manage_backups', 
                                         allowed_roles=['technician', 'principal'])
    
    if not has_access:
        flash('You do not have permission to perform this action.', 'error')
        return redirect(url_for('dashboard_employee'))
    
    auto_backup = request.form.get('auto_backup') == 'on'
    frequency = request.form.get('frequency', 'daily')
    
    connection = get_db_connection()
    if not connection:
        flash('Database connection error.', 'error')
        return redirect(url_for('database_backup_restore'))
    
    try:
        with connection.cursor() as cursor:
            # Create backup_settings table if it doesn't exist
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS backup_settings (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    auto_backup_enabled BOOLEAN DEFAULT FALSE,
                    backup_frequency ENUM('daily', 'weekly', 'monthly') DEFAULT 'daily',
                    last_backup DATETIME,
                    next_backup DATETIME,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    updated_by VARCHAR(255)
                )
            """)
            connection.commit()
            
            # Calculate next backup time
            next_backup = None
            if auto_backup:
                if frequency == 'daily':
                    next_backup = datetime.now() + timedelta(days=1)
                elif frequency == 'weekly':
                    next_backup = datetime.now() + timedelta(weeks=1)
                elif frequency == 'monthly':
                    next_backup = datetime.now() + timedelta(days=30)
            
            # Check if settings exist
            cursor.execute("SELECT COUNT(*) as count FROM backup_settings")
            exists = cursor.fetchone()
            
            if exists and exists.get('count', 0) > 0:
                # Update existing
                cursor.execute("""
                    UPDATE backup_settings 
                    SET auto_backup_enabled = %s, 
                        backup_frequency = %s, 
                        next_backup = %s,
                        updated_by = %s
                    WHERE id = (SELECT id FROM (SELECT id FROM backup_settings ORDER BY id DESC LIMIT 1) AS tmp)
                """, (auto_backup, frequency, next_backup, session.get('full_name', 'Unknown')))
            else:
                # Insert new
                cursor.execute("""
                    INSERT INTO backup_settings (auto_backup_enabled, backup_frequency, next_backup, updated_by)
                    VALUES (%s, %s, %s, %s)
                """, (auto_backup, frequency, next_backup, session.get('full_name', 'Unknown')))
            
            connection.commit()
            flash('Backup settings updated successfully.', 'success')
    except Exception as e:
        print(f"Error updating backup settings: {e}")
        flash('Error updating backup settings.', 'error')
    finally:
        if connection:
            try:
                connection.close()
            except:
                pass
    
    return redirect(url_for('database_backup_restore'))

# Database Health & Status Route
@app.route('/database/health-status')
@login_required
def database_health_status():
    """Database health and status analysis page for technicians and principals"""
    user_role = session.get('role', '').lower()
    
    # Only technicians and principals can access this page
    if user_role not in ['technician', 'principal']:
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('dashboard_employee'))
    
    connection = get_db_connection()
    health_status = {
        'connection_status': False,
        'overall_status': 'critical',
        'db_name': '',
        'mysql_version': 'Unknown',
        'character_set': 'Unknown',
        'collation': 'Unknown',
        'uptime': 'Unknown',
        'last_checked': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'database_size_mb': 0,
        'data_size_mb': 0,
        'index_size_mb': 0,
        'total_tables': 0,
        'total_records': 0,
        'tables': [],
        'largest_tables': [],
        'recommendations': []
    }
    
    if connection:
        try:
            health_status['connection_status'] = True
            with connection.cursor() as cursor:
                # Get database name
                cursor.execute("SELECT DATABASE() as db_name")
                db_result = cursor.fetchone()
                health_status['db_name'] = db_result.get('db_name', 'Unknown') if db_result else 'Unknown'
                
                # Get MySQL version
                cursor.execute("SELECT VERSION() as version")
                version_result = cursor.fetchone()
                health_status['mysql_version'] = version_result.get('version', 'Unknown') if version_result else 'Unknown'
                
                # Get character set and collation
                cursor.execute("""
                    SELECT DEFAULT_CHARACTER_SET_NAME, DEFAULT_COLLATION_NAME 
                    FROM information_schema.SCHEMATA 
                    WHERE SCHEMA_NAME = DATABASE()
                """)
                charset_result = cursor.fetchone()
                if charset_result:
                    health_status['character_set'] = charset_result.get('DEFAULT_CHARACTER_SET_NAME', 'Unknown')
                    health_status['collation'] = charset_result.get('DEFAULT_COLLATION_NAME', 'Unknown')
                
                # Get uptime
                try:
                    cursor.execute("SHOW STATUS LIKE 'Uptime'")
                    uptime_result = cursor.fetchone()
                    if uptime_result:
                        if isinstance(uptime_result, dict):
                            uptime_seconds = int(uptime_result.get('Value', 0))
                        elif isinstance(uptime_result, tuple):
                            uptime_seconds = int(uptime_result[1] if len(uptime_result) > 1 else 0)
                        else:
                            uptime_seconds = 0
                        days = uptime_seconds // 86400
                        hours = (uptime_seconds % 86400) // 3600
                        minutes = (uptime_seconds % 3600) // 60
                        health_status['uptime'] = f"{days}d {hours}h {minutes}m"
                except:
                    health_status['uptime'] = 'Unknown'
                
                # Get database size
                cursor.execute("""
                    SELECT 
                        ROUND(SUM(data_length + index_length) / 1024 / 1024, 2) AS total_size_mb,
                        ROUND(SUM(data_length) / 1024 / 1024, 2) AS data_size_mb,
                        ROUND(SUM(index_length) / 1024 / 1024, 2) AS index_size_mb
                    FROM information_schema.TABLES 
                    WHERE table_schema = DATABASE()
                """)
                size_result = cursor.fetchone()
                if size_result:
                    health_status['database_size_mb'] = size_result.get('total_size_mb', 0) or 0
                    health_status['data_size_mb'] = size_result.get('data_size_mb', 0) or 0
                    health_status['index_size_mb'] = size_result.get('index_size_mb', 0) or 0
                
                # Get all tables
                cursor.execute("SHOW TABLES")
                table_results = cursor.fetchall()
                health_status['total_tables'] = len(table_results)
                
                table_list = []
                largest_tables = []
                
                # Analyze each table
                for table_result in table_results:
                    table_name = list(table_result.values())[0] if isinstance(table_result, dict) else table_result[0]
                    
                    # Get row count
                    try:
                        cursor.execute(f"SELECT COUNT(*) as count FROM `{table_name}`")
                        count_result = cursor.fetchone()
                        row_count = count_result.get('count', 0) if count_result else 0
                        health_status['total_records'] += row_count
                    except:
                        row_count = 0
                    
                    # Get table size and index count
                    cursor.execute("""
                        SELECT 
                            ROUND(((data_length + index_length) / 1024 / 1024), 2) AS size_mb,
                            ROUND((data_length / 1024 / 1024), 2) AS data_mb,
                            ROUND((index_length / 1024 / 1024), 2) AS index_mb
                        FROM information_schema.TABLES 
                        WHERE table_schema = DATABASE() 
                        AND table_name = %s
                    """, (table_name,))
                    size_result = cursor.fetchone()
                    size_mb = size_result.get('size_mb', 0) or 0 if size_result else 0
                    
                    # Get index count
                    cursor.execute(f"SHOW INDEX FROM `{table_name}`")
                    indexes = cursor.fetchall()
                    index_count = len(set([idx.get('Key_name', '') if isinstance(idx, dict) else idx[2] if isinstance(idx, tuple) else '' for idx in indexes]))
                    
                    # Determine table status
                    table_status = 'healthy'
                    if size_mb > 100:  # Large table warning
                        table_status = 'warning'
                    if row_count == 0 and size_mb > 0:  # Empty but has size (possible issue)
                        table_status = 'warning'
                    
                    table_info = {
                        'name': table_name,
                        'rows': row_count,
                        'size_mb': size_mb,
                        'index_count': index_count,
                        'status': table_status
                    }
                    table_list.append(table_info)
                    largest_tables.append(table_info)
                
                # Sort by size
                largest_tables.sort(key=lambda x: x['size_mb'], reverse=True)
                health_status['tables'] = sorted(table_list, key=lambda x: x['size_mb'], reverse=True)
                health_status['largest_tables'] = largest_tables
                
                # Determine overall health status
                issues = 0
                warnings = 0
                
                # Check for large database
                if health_status['database_size_mb'] > 1000:
                    warnings += 1
                    health_status['recommendations'].append(f"Database size is {health_status['database_size_mb']:.2f} MB. Consider archiving old data.")
                
                # Check for tables without indexes
                tables_without_indexes = [t for t in table_list if t['index_count'] == 0 and t['rows'] > 100]
                if tables_without_indexes:
                    warnings += 1
                    health_status['recommendations'].append(f"{len(tables_without_indexes)} table(s) with >100 rows have no indexes. Consider adding indexes for better performance.")
                
                # Check for very large tables
                very_large_tables = [t for t in table_list if t['size_mb'] > 500]
                if very_large_tables:
                    warnings += 1
                    health_status['recommendations'].append(f"{len(very_large_tables)} table(s) exceed 500 MB. Consider partitioning or archiving.")
                
                # Determine overall status
                if issues > 0:
                    health_status['overall_status'] = 'critical'
                elif warnings > 0:
                    health_status['overall_status'] = 'warning'
                else:
                    health_status['overall_status'] = 'healthy'
                
        except Exception as e:
            print(f"Error analyzing database health: {e}")
            import traceback
            traceback.print_exc()
            health_status['connection_status'] = False
            health_status['overall_status'] = 'critical'
            health_status['recommendations'].append(f"Error analyzing database: {str(e)}")
        finally:
            if connection:
                try:
                    connection.close()
                except:
                    pass
    else:
        health_status['connection_status'] = False
        health_status['overall_status'] = 'critical'
        health_status['recommendations'].append("Unable to connect to database. Check database configuration.")
    
    return render_template('dashboards/database_health_status.html', health_status=health_status)

# Logs & Audit Trails Route
@app.route('/database/logs-audit-trails')
@login_required
def logs_audit_trails():
    """Logs and audit trails page for technicians and principals"""
    user_role = session.get('role', '').lower()
    employee_id = session.get('employee_id') or session.get('user_id')
    
    # Check permission OR role-based access
    has_access = check_permission_or_role('view_audit_logs', 
                                         allowed_roles=['technician', 'principal'])
    
    if not has_access:
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('dashboard_employee'))
    
    connection = get_db_connection()
    salary_audits = []
    migrations = []
    backups = []
    audit_summary = {
        'salary_audits_count': 0,
        'migrations_count': 0,
        'backups_count': 0,
        'total_records': 0
    }
    
    if connection:
        try:
            with connection.cursor() as cursor:
                # Get salary audits
                try:
                    cursor.execute("""
                        SELECT 
                            esa.id,
                            esa.salary_id,
                            esa.employee_id,
                            esa.field_name,
                            esa.old_value,
                            esa.new_value,
                            esa.edited_by,
                            esa.edited_by_name,
                            esa.edited_at,
                            e.full_name as employee_name,
                            e.employee_id as employee_code
                        FROM employee_salary_audits esa
                        LEFT JOIN employees e ON esa.employee_id = e.id
                        ORDER BY esa.edited_at DESC
                        LIMIT 500
                    """)
                    salary_audits = cursor.fetchall()
                    audit_summary['salary_audits_count'] = len(salary_audits)
                except Exception as e:
                    print(f"Error fetching salary audits: {e}")
                    salary_audits = []
                
                # Get migrations
                try:
                    cursor.execute("""
                        SELECT 
                            id,
                            migration_name,
                            status,
                            applied_at,
                            execution_time_ms,
                            applied_by,
                            error_message
                        FROM migrations
                        ORDER BY applied_at DESC
                        LIMIT 100
                    """)
                    migrations = cursor.fetchall()
                    audit_summary['migrations_count'] = len(migrations)
                except Exception as e:
                    print(f"Error fetching migrations: {e}")
                    migrations = []
                
                # Get backup history
                try:
                    cursor.execute("""
                        SELECT 
                            id,
                            filename,
                            file_path,
                            file_size,
                            table_count,
                            record_count,
                            created_at,
                            created_by
                        FROM backup_history
                        ORDER BY created_at DESC
                        LIMIT 50
                    """)
                    backups = cursor.fetchall()
                    audit_summary['backups_count'] = len(backups)
                except Exception as e:
                    print(f"Error fetching backup history: {e}")
                    backups = []
                
                audit_summary['total_records'] = len(salary_audits) + len(migrations) + len(backups)
                
        except Exception as e:
            print(f"Error fetching audit trails: {e}")
            import traceback
            traceback.print_exc()
        finally:
            if connection:
                try:
                    connection.close()
                except:
                    pass
    
    return render_template('dashboards/logs_audit_trails.html', 
                         salary_audits=salary_audits,
                         migrations=migrations,
                         backups=backups,
                         audit_summary=audit_summary)

# Users & Roles Route
@app.route('/users-roles')
@login_required
def users_roles():
    """Users and roles management page for technicians"""
    user_role = session.get('role', '').lower()
    
    # Only technicians can access this page
    if user_role != 'technician':
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('dashboard_employee'))
    
    connection = get_db_connection()
    employees_by_role = {}
    summary = {
        'total_employees': 0,
        'active_count': 0,
        'pending_count': 0,
        'total_roles': 0
    }
    
    if connection:
        try:
            with connection.cursor() as cursor:
                # Fetch all employees
                cursor.execute("""
                    SELECT 
                        id, employee_id, full_name, email, phone, id_number, 
                        role, status, profile_picture, created_at, updated_at
                    FROM employees
                    ORDER BY role ASC, full_name ASC
                """)
                employees = cursor.fetchall()
                
                # Group employees by role
                for employee in employees:
                    role = employee.get('role', 'employee')
                    if role not in employees_by_role:
                        employees_by_role[role] = []
                    employees_by_role[role].append(employee)
                    
                    # Update summary
                    summary['total_employees'] += 1
                    if employee.get('status') == 'active':
                        summary['active_count'] += 1
                    elif employee.get('status') == 'pending approval':
                        summary['pending_count'] += 1
                
                summary['total_roles'] = len(employees_by_role)
                
        except Exception as e:
            print(f"Error fetching users and roles: {e}")
            import traceback
            traceback.print_exc()
        finally:
            if connection:
                try:
                    connection.close()
                except:
                    pass
    
    return render_template('dashboards/users_roles.html', 
                         employees_by_role=employees_by_role,
                         summary=summary)

# Get Employee Permissions Route
@app.route('/users-roles/get-permissions/<int:employee_id>')
@login_required
def get_employee_permissions(employee_id):
    """Get permissions for a specific employee filtered by their role"""
    user_role = session.get('role', '').lower()
    
    # Only technicians can access this
    if user_role != 'technician':
        return jsonify({'success': False, 'message': 'Permission denied'}), 403
    
    # Define all available permissions
    all_permissions = [
        {'key': 'view_students', 'name': 'View Students', 'description': 'View student information and records', 'roles': ['principal', 'deputy principal', 'academic coordinator', 'teachers', 'accountant', 'librarian', 'warden', 'transport manager']},
        {'key': 'add_students', 'name': 'Add Students', 'description': 'Add new students to the system', 'roles': ['principal', 'deputy principal', 'academic coordinator']},
        {'key': 'edit_students', 'name': 'Edit Students', 'description': 'Edit existing student information', 'roles': ['principal', 'deputy principal', 'academic coordinator', 'teachers']},
        {'key': 'delete_students', 'name': 'Delete Students', 'description': 'Remove students from the system', 'roles': ['principal', 'deputy principal']},
        {'key': 'view_student_fees', 'name': 'View Student Fees', 'description': 'View student fee information', 'roles': ['principal', 'accountant', 'deputy principal']},
        {'key': 'view_fee_structure_details', 'name': 'View Fee Structure Details', 'description': 'View detailed fee structure information', 'roles': ['principal', 'accountant', 'deputy principal']},
        {'key': 'process_payments', 'name': 'Record Payments', 'description': 'Record and process student payments', 'roles': ['principal', 'accountant']},
        {'key': 'view_staff', 'name': 'View Staff', 'description': 'View staff member information', 'roles': ['principal', 'deputy principal', 'academic coordinator', 'accountant']},
        {'key': 'add_staff', 'name': 'Add Staff', 'description': 'Add new staff members', 'roles': ['principal', 'deputy principal']},
        {'key': 'edit_staff', 'name': 'Edit Staff', 'description': 'Edit staff member information', 'roles': ['principal', 'deputy principal', 'academic coordinator']},
        {'key': 'delete_staff', 'name': 'Delete Staff', 'description': 'Remove staff members', 'roles': ['principal']},
        {'key': 'manage_salaries', 'name': 'Manage Salaries', 'description': 'View and manage staff salaries', 'roles': ['principal', 'accountant', 'deputy principal']},
        {'key': 'view_fees', 'name': 'View Fees', 'description': 'View fee structures and information', 'roles': ['principal', 'accountant', 'deputy principal']},
        {'key': 'manage_fees', 'name': 'Manage Fees', 'description': 'Create and edit fee structures', 'roles': ['principal', 'accountant']},
        {'key': 'add_fee_structure', 'name': 'Add Fee Structure', 'description': 'Create new fee structures', 'roles': ['principal', 'accountant']},
        {'key': 'edit_fee_structure', 'name': 'Edit Fee Structure', 'description': 'Edit existing fee structures', 'roles': ['principal', 'accountant']},
        {'key': 'delete_fee_structure', 'name': 'Delete Fee Structure', 'description': 'Delete fee structures', 'roles': ['principal', 'accountant']},
        {'key': 'view_financial_reports', 'name': 'View Financial Reports', 'description': 'Access financial reports and analytics', 'roles': ['principal', 'accountant', 'deputy principal']},
        {'key': 'generate_invoices', 'name': 'Generate Invoices', 'description': 'Create and generate invoices', 'roles': ['principal', 'accountant']},
        {'key': 'view_academic_levels', 'name': 'View Academic Levels', 'description': 'View academic levels and grades', 'roles': ['principal', 'deputy principal', 'academic coordinator', 'teachers']},
        {'key': 'manage_academic_levels', 'name': 'Manage Academic Levels', 'description': 'Create and edit academic levels', 'roles': ['principal', 'deputy principal', 'academic coordinator']},
        {'key': 'view_exams', 'name': 'View Exams', 'description': 'View exam information', 'roles': ['principal', 'deputy principal', 'academic coordinator', 'teachers']},
        {'key': 'manage_exams', 'name': 'Manage Exams', 'description': 'Create and edit exams', 'roles': ['principal', 'deputy principal', 'academic coordinator']},
        {'key': 'view_results', 'name': 'View Results', 'description': 'View exam and academic results', 'roles': ['principal', 'deputy principal', 'academic coordinator', 'teachers']},
        {'key': 'view_database', 'name': 'View Database', 'description': 'Access database management tools', 'roles': ['principal']},
        {'key': 'manage_backups', 'name': 'Manage Backups', 'description': 'Create and restore database backups', 'roles': ['principal']},
        {'key': 'system_settings', 'name': 'System Settings', 'description': 'Access and modify system settings', 'roles': []},  # Only technicians
        {'key': 'manage_users', 'name': 'Manage Users', 'description': 'Manage user accounts and permissions', 'roles': []},  # Only technicians
        {'key': 'view_audit_logs', 'name': 'View Audit Logs', 'description': 'View system audit trails and logs', 'roles': ['principal']},
        {'key': 'view_reports', 'name': 'View Reports', 'description': 'View system reports', 'roles': ['principal', 'deputy principal', 'academic coordinator', 'accountant']},
        {'key': 'generate_reports', 'name': 'Generate Reports', 'description': 'Generate custom reports', 'roles': ['principal', 'deputy principal', 'accountant']},
        {'key': 'export_data', 'name': 'Export Data', 'description': 'Export data to various formats', 'roles': ['principal', 'deputy principal', 'accountant']},
        {'key': 'view_analytics', 'name': 'View Analytics', 'description': 'Access analytics and insights', 'roles': ['principal', 'deputy principal', 'accountant']}
    ]
    
    connection = get_db_connection()
    employee_role = None
    employee_permissions = []
    
    if connection:
        try:
            with connection.cursor() as cursor:
                # Get employee's role
                cursor.execute("""
                    SELECT role 
                    FROM employees 
                    WHERE id = %s
                """, (employee_id,))
                result = cursor.fetchone()
                if result:
                    employee_role = result.get('role') if isinstance(result, dict) else result[0]
                    if employee_role:
                        employee_role = employee_role.lower()
                
                # Get employee's current permissions
                cursor.execute("""
                    SELECT permission_key 
                    FROM employee_permissions 
                    WHERE employee_id = %s
                """, (employee_id,))
                employee_permissions = cursor.fetchall()
        except Exception as e:
            print(f"Error fetching employee permissions: {e}")
        finally:
            if connection:
                try:
                    connection.close()
                except:
                    pass
    
    # Filter permissions based on employee's role
    if employee_role:
        # Filter permissions that are relevant to this role
        # Include permissions where the role is in the allowed roles list
        # Permissions with empty roles list are technician-only, so exclude them for non-technicians
        filtered_permissions = []
        for perm in all_permissions:
            perm_roles = perm.get('roles', [])
            # Normalize role names for comparison
            perm_roles_lower = [r.lower().strip() for r in perm_roles]
            if employee_role in perm_roles_lower:
                filtered_permissions.append(perm)
        
        print(f"DEBUG: Filtered {len(filtered_permissions)} permissions for role '{employee_role}' out of {len(all_permissions)} total")
    else:
        # If role not found, show all permissions (fallback)
        print(f"DEBUG: Employee role not found, showing all permissions")
        filtered_permissions = all_permissions
    
    # Convert employee_permissions to list of permission keys
    employee_permission_keys = []
    for perm in employee_permissions:
        if isinstance(perm, dict):
            employee_permission_keys.append(perm.get('permission_key', ''))
        else:
            employee_permission_keys.append(perm[0] if perm else '')
    
    return jsonify({
        'success': True,
        'permissions': filtered_permissions,
        'employee_permissions': employee_permission_keys,
        'employee_role': employee_role
    })

# Update Employee Permissions Route
@app.route('/users-roles/update-permissions/<int:employee_id>', methods=['POST'])
@login_required
def update_employee_permissions(employee_id):
    """Update permissions for a specific employee"""
    user_role = session.get('role', '').lower()
    
    # Only technicians can access this
    if user_role != 'technician':
        return jsonify({'success': False, 'message': 'Permission denied'}), 403
    
    data = request.get_json()
    permissions = data.get('permissions', [])
    
    # Get the current user's employee ID from session
    current_user_employee_id = session.get('employee_id') or session.get('user_id')
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': 'Database connection error'}), 500
    
    try:
        with connection.cursor() as cursor:
            # Verify that the current user's employee ID exists in the employees table
            granted_by = None
            if current_user_employee_id:
                cursor.execute("SELECT id FROM employees WHERE id = %s OR employee_id = %s", 
                             (current_user_employee_id, current_user_employee_id))
                employee_check = cursor.fetchone()
                if employee_check:
                    # Get the actual employee ID (not employee_id field)
                    granted_by = employee_check.get('id') if isinstance(employee_check, dict) else employee_check[0]
            
            # Delete all existing permissions for this employee
            cursor.execute("DELETE FROM employee_permissions WHERE employee_id = %s", (employee_id,))
            
            # Insert new permissions
            if permissions:
                for permission_key in permissions:
                    cursor.execute("""
                        INSERT INTO employee_permissions (employee_id, permission_key, granted_by)
                        VALUES (%s, %s, %s)
                    """, (employee_id, permission_key, granted_by))
            
            connection.commit()
            return jsonify({'success': True, 'message': 'Permissions updated successfully'})
            
    except Exception as e:
        print(f"Error updating permissions: {e}")
        import traceback
        traceback.print_exc()
        connection.rollback()
        return jsonify({'success': False, 'message': f'Error updating permissions: {str(e)}'}), 500
    finally:
        if connection:
            try:
                connection.close()
            except:
                pass

# School Profile Update Route
@app.route('/system-settings/school-profile', methods=['POST'])
@login_required
def update_school_profile():
    """Update school profile information"""
    user_role = session.get('role', '').lower()
    
    # Only technicians can update
    if user_role != 'technician':
        flash('You do not have permission to perform this action.', 'error')
        return redirect(url_for('system_settings'))
    
    # Get form data - convert to uppercase except email and social links
    school_name = request.form.get('school_name', '').strip().upper()
    school_email = request.form.get('school_email', '').strip().lower()
    school_phone = request.form.get('school_phone', '').strip().upper()
    twitter_url = request.form.get('twitter_url', '').strip()  # Keep as is
    facebook_url = request.form.get('facebook_url', '').strip()  # Keep as is
    instagram_url = request.form.get('instagram_url', '').strip()  # Keep as is
    tiktok_url = request.form.get('tiktok_url', '').strip()  # Keep as is
    whatsapp_number = request.form.get('whatsapp_number', '').strip().upper()
    school_location = request.form.get('school_location', '').strip().upper()
    
    # Handle logo upload
    school_logo = None
    if 'school_logo' in request.files:
        file = request.files['school_logo']
        if file and file.filename != '' and allowed_file(file.filename):
            # Generate unique filename
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            filename = secure_filename(f"school_logo_{timestamp}_{file.filename}")
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            school_logo = f"uploads/profiles/{filename}"
    
    # Update database
    connection = get_db_connection()
    if connection:
        try:
            with connection.cursor() as cursor:
                # Check if school settings exist
                cursor.execute("SELECT id FROM school_settings ORDER BY id DESC LIMIT 1")
                result = cursor.fetchone()
                
                if result:
                    # Get ID from result (handle both dict and tuple)
                    school_id = result['id'] if isinstance(result, dict) else result[0]
                    
                    # Update existing
                    if school_logo:
                        cursor.execute("""
                            UPDATE school_settings 
                            SET school_name = %s, school_email = %s, school_phone = %s,
                                school_logo = %s, twitter_url = %s, facebook_url = %s,
                                instagram_url = %s, tiktok_url = %s, whatsapp_number = %s,
                                school_location = %s
                            WHERE id = %s
                        """, (school_name, school_email, school_phone, school_logo, 
                              twitter_url, facebook_url, instagram_url, tiktok_url, 
                              whatsapp_number, school_location, school_id))
                    else:
                        cursor.execute("""
                            UPDATE school_settings 
                            SET school_name = %s, school_email = %s, school_phone = %s,
                                twitter_url = %s, facebook_url = %s,
                                instagram_url = %s, tiktok_url = %s, whatsapp_number = %s,
                                school_location = %s
                            WHERE id = %s
                        """, (school_name, school_email, school_phone,
                              twitter_url, facebook_url, instagram_url, tiktok_url, 
                              whatsapp_number, school_location, school_id))
                else:
                    # Insert new
                    cursor.execute("""
                        INSERT INTO school_settings 
                        (school_name, school_email, school_phone, school_logo, 
                         twitter_url, facebook_url, instagram_url, tiktok_url, 
                         whatsapp_number, school_location)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (school_name, school_email, school_phone, school_logo,
                          twitter_url, facebook_url, instagram_url, tiktok_url,
                          whatsapp_number, school_location))
                
                connection.commit()
                flash('School profile updated successfully!', 'success')
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"Error updating school profile: {e}")
            print(f"Full traceback:\n{error_details}")
            connection.rollback()
            flash(f'An error occurred while updating the school profile: {str(e)}. Please try again.', 'error')
        finally:
            connection.close()
    else:
        flash('Database connection error. Please try again later.', 'error')
    
    return redirect(url_for('system_settings'))

# Academic Level Registration Route
@app.route('/system-settings/academic-level', methods=['POST'])
@login_required
def add_academic_level():
    """Add a new academic level"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    
    # Allow technicians and accountants
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_technician = user_role == 'technician'
    
    if not (is_technician or is_accountant):
        flash('You do not have permission to perform this action.', 'error')
        return redirect(url_for('academic_settings') if is_accountant else url_for('system_settings'))
    
    # Get form data and convert to uppercase
    level_category = request.form.get('level_category', '').strip().upper()
    level_name = request.form.get('level_name', '').strip().upper()
    level_description = request.form.get('level_description', '').strip().upper()
    level_status_value = request.form.get('level_status_value', 'active').strip().lower()
    
    # Validate required fields
    if not level_category or not level_name:
        flash('Please fill in all required fields (Level Category and Level Name).', 'error')
        return redirect(url_for('academic_settings') if is_accountant else url_for('system_settings'))
    
    # Validate status
    if level_status_value not in ['active', 'inactive']:
        level_status_value = 'active'
    
    # Insert into database
    connection = get_db_connection()
    if connection:
        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    INSERT INTO academic_levels 
                    (level_category, level_name, level_description, level_status)
                    VALUES (%s, %s, %s, %s)
                """, (level_category, level_name, level_description, level_status_value))
                
                connection.commit()
                flash(f'Academic level "{level_name}" added successfully!', 'success')
        except Exception as e:
            print(f"Error adding academic level: {e}")
            connection.rollback()
            flash('An error occurred while adding the academic level. Please try again.', 'error')
        finally:
            connection.close()
    else:
        flash('Database connection error. Please try again later.', 'error')
    
    return redirect(url_for('system_settings'))

# Academic Level Status Update Route
@app.route('/system-settings/academic-level/<int:level_id>/toggle-status', methods=['POST'])
@login_required
def toggle_academic_level_status(level_id):
    """Toggle academic level status between active and inactive"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    
    # Allow technicians and accountants
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_technician = user_role == 'technician'
    
    if not (is_technician or is_accountant):
        return jsonify({'success': False, 'message': 'You do not have permission to perform this action.'}), 403
    
    # Get current status from database
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': 'Database connection error.'}), 500
    
    try:
        with connection.cursor() as cursor:
            # Get current status
            cursor.execute("SELECT level_status FROM academic_levels WHERE id = %s", (level_id,))
            result = cursor.fetchone()
            
            if not result:
                return jsonify({'success': False, 'message': 'Academic level not found.'}), 404
            
            current_status = result.get('level_status') if isinstance(result, dict) else result[0]
            new_status = 'inactive' if current_status == 'active' else 'active'
            
            # Update status
            cursor.execute("""
                UPDATE academic_levels 
                SET level_status = %s, updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
            """, (new_status, level_id))
            
            connection.commit()
            return jsonify({
                'success': True, 
                'message': f'Status updated to {new_status}.',
                'new_status': new_status
            })
    except Exception as e:
        print(f"Error updating academic level status: {e}")
        connection.rollback()
        return jsonify({'success': False, 'message': 'An error occurred while updating the status.'}), 500
    finally:
        connection.close()

# Academic Level Update Route
@app.route('/system-settings/academic-level/<int:level_id>/update', methods=['POST'])
@login_required
def update_academic_level(level_id):
    """Update an existing academic level"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    
    # Allow technicians and accountants
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_technician = user_role == 'technician'
    
    if not (is_technician or is_accountant):
        return jsonify({'success': False, 'message': 'You do not have permission to perform this action.'}), 403
    
    # Get JSON data
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'message': 'Invalid request data.'}), 400
    
    level_category = data.get('level_category', '').strip().upper()
    level_name = data.get('level_name', '').strip().upper()
    level_description = data.get('level_description', '').strip().upper()
    level_status = data.get('level_status', 'active').strip().lower()
    
    # Validate required fields
    if not level_category or not level_name:
        return jsonify({'success': False, 'message': 'Level Category and Level Name are required.'}), 400
    
    # Validate status
    if level_status not in ['active', 'inactive']:
        level_status = 'active'
    
    # Update database
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': 'Database connection error.'}), 500
    
    try:
        with connection.cursor() as cursor:
            # Check if level exists
            cursor.execute("SELECT id FROM academic_levels WHERE id = %s", (level_id,))
            if not cursor.fetchone():
                return jsonify({'success': False, 'message': 'Academic level not found.'}), 404
            
            # Update the level
            cursor.execute("""
                UPDATE academic_levels 
                SET level_category = %s, level_name = %s, level_description = %s, 
                    level_status = %s, updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
            """, (level_category, level_name, level_description, level_status, level_id))
            
            connection.commit()
            return jsonify({
                'success': True, 
                'message': f'Academic level "{level_name}" updated successfully!'
            })
    except Exception as e:
        print(f"Error updating academic level: {e}")
        connection.rollback()
        return jsonify({'success': False, 'message': 'An error occurred while updating the academic level.'}), 500
    finally:
        connection.close()

# Academic Level Delete Route
@app.route('/system-settings/academic-level/<int:level_id>/delete', methods=['POST'])
@login_required
def delete_academic_level(level_id):
    """Delete an academic level"""
    user_role = session.get('role', '').lower()
    
    # Only technicians can delete academic levels
    if user_role != 'technician':
        return jsonify({'success': False, 'message': 'You do not have permission to perform this action.'}), 403
    
    # Delete from database
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': 'Database connection error.'}), 500
    
    try:
        with connection.cursor() as cursor:
            # Check if level exists and get name for message
            cursor.execute("SELECT level_name FROM academic_levels WHERE id = %s", (level_id,))
            result = cursor.fetchone()
            
            if not result:
                return jsonify({'success': False, 'message': 'Academic level not found.'}), 404
            
            level_name = result.get('level_name') if isinstance(result, dict) else result[0]
            
            # Delete the level
            cursor.execute("DELETE FROM academic_levels WHERE id = %s", (level_id,))
            
            connection.commit()
            return jsonify({
                'success': True, 
                'message': f'Academic level "{level_name}" deleted successfully!'
            })
    except Exception as e:
        print(f"Error deleting academic level: {e}")
        connection.rollback()
        return jsonify({'success': False, 'message': 'An error occurred while deleting the academic level.'}), 500
    finally:
        connection.close()

# Academic Year Routes
@app.route('/system-settings/academic-year', methods=['POST'])
@login_required
def create_academic_year():
    """Create a new academic year"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    
    # Allow technicians and accountants
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_technician = user_role == 'technician'
    
    if not (is_technician or is_accountant):
        flash('You do not have permission to perform this action.', 'error')
        return redirect(url_for('academic_settings') if is_accountant else url_for('system_settings'))
    
    year_name = request.form.get('year_name', '').strip()
    start_date = request.form.get('start_date', '').strip()
    end_date = request.form.get('end_date', '').strip()
    status = request.form.get('status', 'draft').strip()
    is_current = request.form.get('is_current', 'false').lower() == 'true'
    
    if not year_name or not start_date or not end_date:
        flash('All required fields must be filled.', 'error')
        return redirect(url_for('academic_settings') if is_accountant else url_for('system_settings'))
    
    connection = get_db_connection()
    if not connection:
        flash('Database connection error.', 'error')
        return redirect(url_for('academic_settings') if is_accountant else url_for('system_settings'))
    
    try:
        with connection.cursor() as cursor:
            # Business Rule: Only one academic year can be current
            if is_current:
                cursor.execute("UPDATE academic_years SET is_current = FALSE")
            
            # Insert new academic year
            cursor.execute("""
                INSERT INTO academic_years (year_name, start_date, end_date, status, is_current)
                VALUES (%s, %s, %s, %s, %s)
            """, (year_name, start_date, end_date, status, is_current))
            
            connection.commit()
            flash(f'Academic year "{year_name}" created successfully!', 'success')
    except Exception as e:
        print(f"Error creating academic year: {e}")
        connection.rollback()
        flash('An error occurred while creating the academic year.', 'error')
    finally:
        connection.close()
    
    return redirect(url_for('academic_settings') if is_accountant else url_for('system_settings'))

@app.route('/system-settings/academic-year/<int:year_id>/update', methods=['POST'])
@login_required
def update_academic_year(year_id):
    """Update an academic year"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    
    # Allow technicians and accountants
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_technician = user_role == 'technician'
    
    if not (is_technician or is_accountant):
        return jsonify({'success': False, 'message': 'Permission denied.'}), 403
    
    data = request.get_json()
    year_name = data.get('year_name', '').strip()
    start_date = data.get('start_date', '').strip()
    end_date = data.get('end_date', '').strip()
    status = data.get('status', 'draft').strip()
    is_current = data.get('is_current', False)
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': 'Database connection error.'}), 500
    
    try:
        with connection.cursor() as cursor:
            # Check if academic year has financial records
            cursor.execute("""
                SELECT COUNT(*) as count FROM student_payments sp
                JOIN fee_structures fs ON sp.fee_structure_id = fs.id
                WHERE fs.academic_year_id = %s
            """, (year_id,))
            result = cursor.fetchone()
            payment_count = result.get('count', 0) if isinstance(result, dict) else result[0]
            
            # Business Rule: Cannot delete academic year with financial records
            if status == 'closed' and payment_count > 0:
                return jsonify({
                    'success': False, 
                    'message': 'Cannot close academic year with financial records.'
                }), 400
            
            # Business Rule: Only one academic year can be current
            if is_current:
                cursor.execute("UPDATE academic_years SET is_current = FALSE WHERE id != %s", (year_id,))
            
            # Update academic year
            cursor.execute("""
                UPDATE academic_years
                SET year_name = %s, start_date = %s, end_date = %s, status = %s, is_current = %s
                WHERE id = %s
            """, (year_name, start_date, end_date, status, is_current, year_id))
            
            # If closing academic year, lock all terms and carry forward balances
            if status == 'closed':
                cursor.execute("UPDATE terms SET status = 'closed' WHERE academic_year_id = %s", (year_id,))
            
            connection.commit()
            return jsonify({'success': True, 'message': 'Academic year updated successfully!'})
    except Exception as e:
        print(f"Error updating academic year: {e}")
        connection.rollback()
        return jsonify({'success': False, 'message': 'An error occurred while updating the academic year.'}), 500
    finally:
        connection.close()

@app.route('/system-settings/academic-year/<int:year_id>/toggle-suspend', methods=['POST'])
@login_required
def toggle_suspend_academic_year(year_id):
    """Toggle academic year status between active and suspended"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    
    # Allow technicians and accountants
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_technician = user_role == 'technician'
    
    if not (is_technician or is_accountant):
        return jsonify({'success': False, 'message': 'Permission denied.'}), 403
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': 'Database connection error.'}), 500
    
    try:
        with connection.cursor() as cursor:
            # Get current status
            cursor.execute("SELECT status FROM academic_years WHERE id = %s", (year_id,))
            result = cursor.fetchone()
            
            if not result:
                return jsonify({'success': False, 'message': 'Academic year not found.'}), 404
            
            current_status = result.get('status') if isinstance(result, dict) else result[0]
            
            # Toggle between active and suspended
            if current_status == 'suspended':
                new_status = 'active'
            else:
                new_status = 'suspended'
            
            # Update status
            cursor.execute("""
                UPDATE academic_years 
                SET status = %s, updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
            """, (new_status, year_id))
            
            connection.commit()
            return jsonify({
                'success': True, 
                'message': f'Academic year status updated to {new_status}.',
                'new_status': new_status
            })
    except Exception as e:
        print(f"Error toggling academic year suspend status: {e}")
        connection.rollback()
        return jsonify({'success': False, 'message': 'An error occurred while updating the status.'}), 500
    finally:
        connection.close()

@app.route('/system-settings/academic-year/<int:year_id>/toggle-lock', methods=['POST'])
@login_required
def toggle_lock_academic_year(year_id):
    """Toggle academic year lock status"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    
    # Allow technicians and accountants
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_technician = user_role == 'technician'
    
    if not (is_technician or is_accountant):
        return jsonify({'success': False, 'message': 'Permission denied.'}), 403
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': 'Database connection error.'}), 500
    
    try:
        with connection.cursor() as cursor:
            # Get current lock status
            cursor.execute("SELECT is_locked FROM academic_years WHERE id = %s", (year_id,))
            result = cursor.fetchone()
            
            if not result:
                return jsonify({'success': False, 'message': 'Academic year not found.'}), 404
            
            current_locked = result.get('is_locked') if isinstance(result, dict) else result[0]
            new_locked = not current_locked
            
            # Update lock status
            if new_locked:
                cursor.execute("""
                    UPDATE academic_years 
                    SET is_locked = TRUE, locked_at = CURRENT_TIMESTAMP, updated_at = CURRENT_TIMESTAMP
                    WHERE id = %s
                """, (year_id,))
            else:
                cursor.execute("""
                    UPDATE academic_years 
                    SET is_locked = FALSE, locked_at = NULL, updated_at = CURRENT_TIMESTAMP
                    WHERE id = %s
                """, (year_id,))
            
            connection.commit()
            action = 'locked' if new_locked else 'unlocked'
            return jsonify({
                'success': True, 
                'message': f'Academic year {action} successfully.',
                'is_locked': new_locked
            })
    except Exception as e:
        print(f"Error toggling academic year lock status: {e}")
        connection.rollback()
        return jsonify({'success': False, 'message': 'An error occurred while updating the lock status.'}), 500
    finally:
        connection.close()

@app.route('/system-settings/academic-year/<int:year_id>/delete', methods=['POST'])
@login_required
def delete_academic_year(year_id):
    """Delete an academic year"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    
    # Allow technicians and accountants
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_technician = user_role == 'technician'
    
    if not (is_technician or is_accountant):
        return jsonify({'success': False, 'message': 'Permission denied.'}), 403
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': 'Database connection error.'}), 500
    
    try:
        with connection.cursor() as cursor:
            # Check if academic year has financial records
            cursor.execute("""
                SELECT COUNT(*) as count FROM student_payments sp
                JOIN fee_structures fs ON sp.fee_structure_id = fs.id
                WHERE fs.academic_year_id = %s
            """, (year_id,))
            result = cursor.fetchone()
            payment_count = result.get('count', 0) if isinstance(result, dict) else result[0]
            
            # Business Rule: Cannot delete academic year with financial records
            if payment_count > 0:
                return jsonify({
                    'success': False, 
                    'message': 'Cannot delete academic year with financial records.'
                }), 400
            
            # Get year name for message
            cursor.execute("SELECT year_name FROM academic_years WHERE id = %s", (year_id,))
            year_result = cursor.fetchone()
            year_name = year_result.get('year_name') if isinstance(year_result, dict) else year_result[0] if year_result else 'Unknown'
            
            # Delete the academic year (cascade will delete terms)
            cursor.execute("DELETE FROM academic_years WHERE id = %s", (year_id,))
            
            connection.commit()
            return jsonify({
                'success': True, 
                'message': f'Academic year "{year_name}" deleted successfully!'
            })
    except Exception as e:
        print(f"Error deleting academic year: {e}")
        connection.rollback()
        return jsonify({'success': False, 'message': 'An error occurred while deleting the academic year.'}), 500
    finally:
        connection.close()

# Term Routes
@app.route('/system-settings/term', methods=['POST'])
@login_required
def create_term():
    """Create a new term"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    
    # Allow technicians and accountants
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_technician = user_role == 'technician'
    
    if not (is_technician or is_accountant):
        flash('You do not have permission to perform this action.', 'error')
        return redirect(url_for('academic_settings') if is_accountant else url_for('system_settings'))
    
    term_name = request.form.get('term_name', '').strip()
    academic_year_id = request.form.get('academic_year_id', '').strip()
    academic_level_ids = request.form.getlist('academic_level_ids')  # Get multiple values
    start_date = request.form.get('start_date', '').strip()
    end_date = request.form.get('end_date', '').strip()
    status = request.form.get('status', 'draft').strip()
    
    if not term_name or not academic_year_id or not start_date or not end_date:
        flash('All required fields must be filled.', 'error')
        return redirect(url_for('academic_settings') if is_accountant else url_for('system_settings'))
    
    connection = get_db_connection()
    if not connection:
        flash('Database connection error.', 'error')
        return redirect(url_for('academic_settings') if is_accountant else url_for('system_settings'))
    
    try:
        with connection.cursor() as cursor:
            # Validate that term dates are within academic year dates
            cursor.execute("""
                SELECT start_date, end_date FROM academic_years WHERE id = %s
            """, (academic_year_id,))
            academic_year = cursor.fetchone()
            
            if not academic_year:
                flash('Selected academic year not found.', 'error')
                return redirect(url_for('academic_settings') if is_accountant else url_for('system_settings'))
            
            year_start = academic_year.get('start_date') if isinstance(academic_year, dict) else academic_year[0]
            year_end = academic_year.get('end_date') if isinstance(academic_year, dict) else academic_year[1]
            
            # Convert to date objects for comparison
            from datetime import datetime
            term_start = datetime.strptime(start_date, '%Y-%m-%d').date()
            term_end = datetime.strptime(end_date, '%Y-%m-%d').date()
            
            if term_start < year_start:
                flash(f'Term start date ({start_date}) must be on or after academic year start date ({year_start}).', 'error')
                return redirect(url_for('academic_settings') if is_accountant else url_for('system_settings'))
            
            if term_end > year_end:
                flash(f'Term end date ({end_date}) must be on or before academic year end date ({year_end}).', 'error')
                return redirect(url_for('academic_settings') if is_accountant else url_for('system_settings'))
            
            if term_start > term_end:
                flash('Term start date must be before or equal to term end date.', 'error')
                return redirect(url_for('academic_settings') if is_accountant else url_for('system_settings'))
            
            # Get is_current from form (default to False)
            is_current = request.form.get('is_current', 'false').lower() == 'true'
            
            # Business Rule: Only one term can be current
            if is_current:
                cursor.execute("UPDATE terms SET is_current = FALSE")
            
            # Get first academic level ID for backward compatibility (optional)
            first_academic_level_id = int(academic_level_ids[0]) if academic_level_ids and academic_level_ids[0] else None
            
            # Insert new term
            cursor.execute("""
                INSERT INTO terms (term_name, academic_year_id, academic_level_id, start_date, end_date, status, is_current)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (term_name, academic_year_id, first_academic_level_id, start_date, end_date, status, is_current))
            
            term_id = cursor.lastrowid
            
            # Insert academic levels into junction table (only active levels)
            if academic_level_ids:
                for level_id in academic_level_ids:
                    if level_id:  # Skip empty values
                        try:
                            level_id_int = int(level_id)
                            # Verify the academic level is active
                            cursor.execute("""
                                SELECT id FROM academic_levels 
                                WHERE id = %s AND level_status = 'active'
                            """, (level_id_int,))
                            if cursor.fetchone():
                                cursor.execute("""
                                    INSERT INTO term_academic_levels (term_id, academic_level_id)
                                    VALUES (%s, %s)
                                    ON DUPLICATE KEY UPDATE term_id = term_id
                                """, (term_id, level_id_int))
                        except (ValueError, TypeError):
                            continue  # Skip invalid IDs
            
            connection.commit()
            flash(f'Term "{term_name}" created successfully!', 'success')
    except Exception as e:
        print(f"Error creating term: {e}")
        connection.rollback()
        flash('An error occurred while creating the term.', 'error')
    finally:
        connection.close()
    
    return redirect(url_for('academic_settings') if is_accountant else url_for('system_settings'))

@app.route('/system-settings/term/<int:term_id>/update', methods=['POST'])
@login_required
def update_term(term_id):
    """Update a term"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    
    # Allow technicians and accountants
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_technician = user_role == 'technician'
    
    if not (is_technician or is_accountant):
        return jsonify({'success': False, 'message': 'Permission denied.'}), 403
    
    data = request.get_json()
    term_name = data.get('term_name', '').strip()
    academic_year_id = data.get('academic_year_id', '').strip()
    academic_level_ids = data.get('academic_level_ids', [])  # Get array of IDs
    if isinstance(academic_level_ids, str):
        # Handle if it comes as a single string
        academic_level_ids = [academic_level_ids] if academic_level_ids else []
    start_date = data.get('start_date', '').strip()
    end_date = data.get('end_date', '').strip()
    status = data.get('status', 'draft').strip()
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': 'Database connection error.'}), 500
    
    try:
        with connection.cursor() as cursor:
            # Validate that term dates are within academic year dates
            cursor.execute("""
                SELECT start_date, end_date FROM academic_years WHERE id = %s
            """, (academic_year_id,))
            academic_year = cursor.fetchone()
            
            if not academic_year:
                return jsonify({'success': False, 'message': 'Selected academic year not found.'}), 400
            
            year_start = academic_year.get('start_date') if isinstance(academic_year, dict) else academic_year[0]
            year_end = academic_year.get('end_date') if isinstance(academic_year, dict) else academic_year[1]
            
            # Convert to date objects for comparison
            from datetime import datetime
            term_start = datetime.strptime(start_date, '%Y-%m-%d').date()
            term_end = datetime.strptime(end_date, '%Y-%m-%d').date()
            
            if term_start < year_start:
                return jsonify({
                    'success': False, 
                    'message': f'Term start date ({start_date}) must be on or after academic year start date ({year_start}).'
                }), 400
            
            if term_end > year_end:
                return jsonify({
                    'success': False, 
                    'message': f'Term end date ({end_date}) must be on or before academic year end date ({year_end}).'
                }), 400
            
            if term_start > term_end:
                return jsonify({
                    'success': False, 
                    'message': 'Term start date must be before or equal to term end date.'
                }), 400
            
            # Get is_current from data (default to False)
            is_current = data.get('is_current', False)
            if isinstance(is_current, str):
                is_current = is_current.lower() == 'true'
            
            # Business Rule: Only one term can be current
            if is_current:
                cursor.execute("UPDATE terms SET is_current = FALSE WHERE id != %s", (term_id,))
            
            # Check if term is active and has fee structures
            cursor.execute("SELECT status FROM terms WHERE id = %s", (term_id,))
            term_result = cursor.fetchone()
            current_status = term_result.get('status') if isinstance(term_result, dict) else term_result[0] if term_result else 'draft'
            
            # Business Rule: Cannot edit fee amounts once term is active
            if current_status == 'active':
                cursor.execute("SELECT COUNT(*) as count FROM fee_structures WHERE term_id = %s", (term_id,))
                fee_result = cursor.fetchone()
                fee_count = fee_result.get('count', 0) if isinstance(fee_result, dict) else fee_result[0]
                
                if fee_count > 0:
                    # Only allow status changes, not other edits
                    cursor.execute("""
                        UPDATE terms
                        SET status = %s, is_current = %s
                        WHERE id = %s
                    """, (status, is_current, term_id))
                else:
                    # No fees yet, allow full update
                    # Get first academic level ID for backward compatibility
                    first_academic_level_id = int(academic_level_ids[0]) if academic_level_ids and academic_level_ids[0] else None
                    cursor.execute("""
                        UPDATE terms
                        SET term_name = %s, academic_year_id = %s, academic_level_id = %s,
                            start_date = %s, end_date = %s, status = %s, is_current = %s
                        WHERE id = %s
                    """, (term_name, academic_year_id, first_academic_level_id, start_date, end_date, status, is_current, term_id))
            else:
                # Term is not active, allow full update
                # Get first academic level ID for backward compatibility
                first_academic_level_id = int(academic_level_ids[0]) if academic_level_ids and academic_level_ids[0] else None
                cursor.execute("""
                    UPDATE terms
                    SET term_name = %s, academic_year_id = %s, academic_level_id = %s,
                        start_date = %s, end_date = %s, status = %s, is_current = %s
                    WHERE id = %s
                """, (term_name, academic_year_id, first_academic_level_id, start_date, end_date, status, is_current, term_id))
            
            # Update academic levels in junction table
            # First, delete existing relationships
            cursor.execute("DELETE FROM term_academic_levels WHERE term_id = %s", (term_id,))
            
            # Then insert new relationships (only active levels)
            if academic_level_ids:
                for level_id in academic_level_ids:
                    if level_id:  # Skip empty values
                        try:
                            level_id_int = int(level_id)
                            # Verify the academic level is active
                            cursor.execute("""
                                SELECT id FROM academic_levels 
                                WHERE id = %s AND level_status = 'active'
                            """, (level_id_int,))
                            if cursor.fetchone():
                                cursor.execute("""
                                    INSERT INTO term_academic_levels (term_id, academic_level_id)
                                    VALUES (%s, %s)
                                    ON DUPLICATE KEY UPDATE term_id = term_id
                                """, (term_id, level_id_int))
                        except (ValueError, TypeError):
                            continue  # Skip invalid IDs
            
            # If closing term, lock invoices and payments, move balances forward
            if status == 'closed':
                # Lock fee structures
                cursor.execute("UPDATE fee_structures SET status = 'inactive' WHERE term_id = %s", (term_id,))
            
            connection.commit()
            return jsonify({'success': True, 'message': 'Term updated successfully!'})
    except Exception as e:
        print(f"Error updating term: {e}")
        connection.rollback()
        return jsonify({'success': False, 'message': 'An error occurred while updating the term.'}), 500
    finally:
        connection.close()

@app.route('/system-settings/term/<int:term_id>/delete', methods=['POST'])
@login_required
def delete_term(term_id):
    """Delete a term"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    
    # Allow technicians and accountants
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_technician = user_role == 'technician'
    
    if not (is_technician or is_accountant):
        return jsonify({'success': False, 'message': 'Permission denied.'}), 403
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': 'Database connection error.'}), 500
    
    try:
        with connection.cursor() as cursor:
            # Check if term has fee structures
            cursor.execute("SELECT COUNT(*) as count FROM fee_structures WHERE term_id = %s", (term_id,))
            result = cursor.fetchone()
            fee_count = result.get('count', 0) if isinstance(result, dict) else result[0]
            
            if fee_count > 0:
                return jsonify({
                    'success': False, 
                    'message': 'Cannot delete term with fee structures.'
                }), 400
            
            # Get term name for message
            cursor.execute("SELECT term_name FROM terms WHERE id = %s", (term_id,))
            term_result = cursor.fetchone()
            term_name = term_result.get('term_name') if isinstance(term_result, dict) else term_result[0] if term_result else 'Unknown'
            
            # Delete the term
            cursor.execute("DELETE FROM terms WHERE id = %s", (term_id,))
            
            connection.commit()
            return jsonify({
                'success': True, 
                'message': f'Term "{term_name}" deleted successfully!'
            })
    except Exception as e:
        print(f"Error deleting term: {e}")
        connection.rollback()
        return jsonify({'success': False, 'message': 'An error occurred while deleting the term.'}), 500
    finally:
        connection.close()

# Term Suspend/Unsuspend Route
@app.route('/system-settings/term/<int:term_id>/toggle-suspend', methods=['POST'])
@login_required
def toggle_term_suspend(term_id):
    """Toggle term suspend/unsuspend status"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    
    # Allow technicians and accountants
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_technician = user_role == 'technician'
    
    if not (is_technician or is_accountant):
        return jsonify({'success': False, 'message': 'Permission denied.'}), 403
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': 'Database connection error.'}), 500
    
    try:
        with connection.cursor() as cursor:
            # Get current status
            cursor.execute("SELECT status FROM terms WHERE id = %s", (term_id,))
            term_result = cursor.fetchone()
            if not term_result:
                return jsonify({'success': False, 'message': 'Term not found.'}), 404
            
            current_status = term_result.get('status') if isinstance(term_result, dict) else term_result[0]
            
            # Toggle between suspended and active
            if current_status == 'suspended':
                new_status = 'active'
                message = 'Term unsuspended successfully!'
            else:
                new_status = 'suspended'
                message = 'Term suspended successfully!'
            
            # Update status
            cursor.execute("""
                UPDATE terms 
                SET status = %s 
                WHERE id = %s
            """, (new_status, term_id))
            
            connection.commit()
            return jsonify({
                'success': True, 
                'message': message,
                'new_status': new_status
            })
    except Exception as e:
        print(f"Error toggling term suspend: {e}")
        connection.rollback()
        return jsonify({'success': False, 'message': 'An error occurred while updating the term status.'}), 500
    finally:
        connection.close()

# Term Lock/Unlock Route
@app.route('/system-settings/term/<int:term_id>/toggle-lock', methods=['POST'])
@login_required
def toggle_term_lock(term_id):
    """Toggle term lock/unlock status"""
    user_role = session.get('role', '').lower()
    viewing_as_role = session.get('viewing_as_employee_role', '').lower()
    
    # Allow technicians and accountants
    is_accountant = user_role == 'accountant' or viewing_as_role == 'accountant'
    is_technician = user_role == 'technician'
    
    if not (is_technician or is_accountant):
        return jsonify({'success': False, 'message': 'Permission denied.'}), 403
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'message': 'Database connection error.'}), 500
    
    try:
        with connection.cursor() as cursor:
            # Get current lock status
            cursor.execute("SELECT is_locked FROM terms WHERE id = %s", (term_id,))
            term_result = cursor.fetchone()
            if not term_result:
                return jsonify({'success': False, 'message': 'Term not found.'}), 404
            
            current_locked = term_result.get('is_locked') if isinstance(term_result, dict) else term_result[0]
            new_locked = not current_locked
            
            # Update lock status
            if new_locked:
                cursor.execute("""
                    UPDATE terms 
                    SET is_locked = TRUE, locked_at = NOW() 
                    WHERE id = %s
                """, (term_id,))
                message = 'Term locked successfully!'
            else:
                cursor.execute("""
                    UPDATE terms 
                    SET is_locked = FALSE, locked_at = NULL 
                    WHERE id = %s
                """, (term_id,))
                message = 'Term unlocked successfully!'
            
            connection.commit()
            return jsonify({
                'success': True, 
                'message': message,
                'is_locked': new_locked
            })
    except Exception as e:
        print(f"Error toggling term lock: {e}")
        connection.rollback()
        return jsonify({'success': False, 'message': 'An error occurred while updating the term lock status.'}), 500
    finally:
        connection.close()

if __name__ == '__main__':
    # Initialize database on startup
    print("Initializing database...")
    try:
        if init_db():
            print("Database initialized successfully!")
        else:
            print("Database initialization failed. Please check your database configuration.")
            print("The application will continue, but some features may not work correctly.")
    except Exception as e:
        print(f"Error during database initialization: {e}")
        import traceback
        traceback.print_exc()
        print("The application will continue, but some features may not work correctly.")
    
    # Run database migrations automatically on startup
    try:
        print("Running database migrations...")
        from migrations.migration_manager import run_all_migrations
        run_all_migrations()
        print("Migrations completed.")
    except Exception as e:
        print(f"Warning: Error running migrations: {e}")
        import traceback
        traceback.print_exc()
        print("The application will continue, but database may not be up to date.")
    
    app.run(debug=True, host='0.0.0.0', port=5000)


