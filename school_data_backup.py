"""
Organized school data backup — groups DB tables into user-friendly Excel workbooks.
"""

import json
import os
import re
from datetime import datetime

# Categories shown on the backup page (order matters for README + sheet grouping).
BACKUP_CATEGORIES = [
    {
        'key': 'fees',
        'title': 'Fees & billing',
        'description': 'Fee structures, invoices, student payments, and payment audit history.',
        'icon': 'fa-money-check-alt',
        'color': '#059669',
        'tables': [
            'fee_structures',
            'fee_items',
            'student_payments',
            'student_payment_audit',
        ],
        'sheet_labels': {
            'fee_structures': 'Fee structures',
            'fee_items': 'Fee line items',
            'student_payments': 'Student payments',
            'student_payment_audit': 'Payment audit log',
        },
    },
    {
        'key': 'attendance',
        'title': 'Attendance',
        'description': 'Daily attendance records for students.',
        'icon': 'fa-user-check',
        'color': '#2563eb',
        'tables': ['student_attendance_records'],
        'sheet_labels': {
            'student_attendance_records': 'Attendance records',
        },
    },
    {
        'key': 'exams',
        'title': 'Exams & grades',
        'description': 'Exam sessions, marks, grade registration, and subject combinations.',
        'icon': 'fa-graduation-cap',
        'color': '#7c3aed',
        'tables': [
            'exams',
            'exam_supervisors',
            'student_marks',
            'grade_registrations',
            'subject_grade_overrides',
            'subject_grade_mark_overrides',
            'class_grade_mark_overrides',
            'grade_setting_profiles',
            'grade_setting_profile_bands',
            'class_grade_setting_assignments',
            'subject_exam_combinations',
            'subject_exam_combination_members',
        ],
        'sheet_labels': {
            'exams': 'Exam sessions',
            'exam_supervisors': 'Exam supervisors',
            'student_marks': 'Student marks',
            'grade_registrations': 'Grade registration',
            'subject_grade_overrides': 'Grade overrides',
            'subject_grade_mark_overrides': 'Subject mark overrides',
            'class_grade_mark_overrides': 'Class mark overrides (legacy)',
            'grade_setting_profiles': 'Grading settings',
            'grade_setting_profile_bands': 'Grading setting bands',
            'class_grade_setting_assignments': 'Class grading assignments',
            'subject_exam_combinations': 'Subject combinations',
            'subject_exam_combination_members': 'Combination members',
        },
    },
    {
        'key': 'timetable',
        'title': 'Timetable',
        'description': 'Class timetables and teacher–subject assignments.',
        'icon': 'fa-calendar-week',
        'color': '#d97706',
        'tables': [
            'timetables',
            'teacher_subject_assignments',
            'academic_coordinator_settings',
        ],
        'sheet_labels': {
            'timetables': 'Timetable slots',
            'teacher_subject_assignments': 'Teacher assignments',
            'academic_coordinator_settings': 'Coordinator settings',
        },
    },
    {
        'key': 'accounts',
        'title': 'Accounts & payroll',
        'description': 'Salaries, payroll payments, revenue, and accountant records.',
        'icon': 'fa-calculator',
        'color': '#0d9488',
        'tables': [
            'employee_salaries',
            'employee_salary_payments',
            'employee_salary_audits',
            'accountant_misc_payments',
            'accountant_payment_descriptions',
            'accountant_revenue',
            'store_stock_in_payment_lines',
        ],
        'sheet_labels': {
            'employee_salaries': 'Salary structures',
            'employee_salary_payments': 'Salary payments',
            'employee_salary_audits': 'Salary audit log',
            'accountant_misc_payments': 'Misc payments',
            'accountant_payment_descriptions': 'Payment descriptions',
            'accountant_revenue': 'Revenue records',
            'store_stock_in_payment_lines': 'Store payment lines',
        },
    },
    {
        'key': 'students',
        'title': 'Students & families',
        'description': 'Student profiles, parents, and admissions.',
        'icon': 'fa-users',
        'color': '#4f46e5',
        'tables': ['students', 'parents', 'admissions'],
        'sheet_labels': {
            'students': 'Students',
            'parents': 'Parents / guardians',
            'admissions': 'Admissions',
        },
    },
    {
        'key': 'academic_setup',
        'title': 'Academic setup',
        'description': 'Levels, subjects, years, and terms used across the system.',
        'icon': 'fa-school',
        'color': '#64748b',
        'tables': [
            'academic_levels',
            'subjects',
            'academic_years',
            'terms',
            'term_academic_levels',
            'subject_academic_levels',
        ],
        'sheet_labels': {
            'academic_levels': 'Academic levels',
            'subjects': 'Subjects',
            'academic_years': 'Academic years',
            'terms': 'Terms',
            'term_academic_levels': 'Term levels',
            'subject_academic_levels': 'Subject levels',
        },
    },
]

_CATEGORY_TABLE_MAP = {}
for _cat in BACKUP_CATEGORIES:
    for _t in _cat['tables']:
        _CATEGORY_TABLE_MAP[_t] = _cat['key']

# Excel sheet titles cannot contain: \ / ? * [ ] :
_INVALID_SHEET_TITLE_RE = re.compile(r'[\\/*?:\[\]]')


def sanitize_excel_sheet_title(title):
    """Make a valid, non-empty Excel worksheet name (max 31 chars)."""
    s = _INVALID_SHEET_TITLE_RE.sub('-', (title or 'Sheet').strip())
    s = s.replace('·', '-').strip(' -')
    return (s[:31] if s else 'Sheet')

# Google Drive folder → tables (under School / Year / Term / …)
DRIVE_BACKUP_SLICES = {
    'accounts/fees': {
        'title': 'Fees',
        'tables': ['fee_structures', 'fee_items'],
        'filename': 'fees_backup.xlsx',
    },
    'accounts/payments': {
        'title': 'Payments & payroll',
        'tables': [
            'student_payments',
            'student_payment_audit',
            'employee_salaries',
            'employee_salary_payments',
            'employee_salary_audits',
            'accountant_misc_payments',
            'accountant_payment_descriptions',
            'accountant_revenue',
            'store_stock_in_payment_lines',
        ],
        'filename': 'payments_backup.xlsx',
    },
    'curriculum/exams': {
        'title': 'Exams & grades',
        'tables': [
            'exams',
            'exam_supervisors',
            'student_marks',
            'grade_registrations',
            'subject_grade_overrides',
            'subject_grade_mark_overrides',
            'class_grade_mark_overrides',
            'grade_setting_profiles',
            'grade_setting_profile_bands',
            'class_grade_setting_assignments',
            'subject_exam_combinations',
            'subject_exam_combination_members',
        ],
        'filename': 'exams_backup.xlsx',
    },
    'curriculum/timetable': {
        'title': 'Timetable',
        'tables': ['timetables', 'teacher_subject_assignments', 'academic_coordinator_settings'],
        'filename': 'timetable_backup.xlsx',
    },
    'curriculum/attendance': {
        'title': 'Attendance',
        'tables': ['student_attendance_records'],
        'filename': 'attendance_backup.xlsx',
    },
}


def friendly_sheet_name(category_key, table_name, used_names):
    """Excel sheet name: max 31 chars, unique."""
    cat = next((c for c in BACKUP_CATEGORIES if c['key'] == category_key), None)
    label = table_name
    if cat:
        label = cat.get('sheet_labels', {}).get(table_name) or table_name.replace('_', ' ').title()
    prefix = {'fees': 'Fees', 'attendance': 'Attend', 'exams': 'Exams', 'timetable': 'Time',
              'accounts': 'Acct', 'students': 'People', 'academic_setup': 'Setup'}.get(category_key, 'Data')
    base = sanitize_excel_sheet_title(f'{prefix} - {label}')
    name = base
    n = 2
    while name in used_names:
        suffix = f' {n}'
        name = sanitize_excel_sheet_title(
            (base[: max(1, 31 - len(suffix))] + suffix) if len(base) + len(suffix) > 31 else base + suffix
        )
        n += 1
    used_names.add(name)
    return name


def categorize_tables(all_table_names):
    """Return dict category_key -> [table names], plus uncategorized."""
    grouped = {c['key']: [] for c in BACKUP_CATEGORIES}
    other = []
    known = set()
    for cat in BACKUP_CATEGORIES:
        for t in cat['tables']:
            known.add(t)
    for name in all_table_names:
        key = _CATEGORY_TABLE_MAP.get(name)
        if key and name in grouped:
            if name not in grouped[key]:
                grouped[key].append(name)
        elif key:
            grouped[key].append(name)
        else:
            other.append(name)
    return grouped, sorted(other)


def build_readme_rows(school_name, created_by, category_stats, include_other_count=0):
    """Rows for README sheet."""
    rows = [
        ['School data backup'],
        ['School', school_name or ''],
        ['Created', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['Created by', created_by or 'System'],
        [''],
        ['Category', 'Tables', 'Records', 'Description'],
    ]
    for cat in BACKUP_CATEGORIES:
        stats = category_stats.get(cat['key'], {'tables': 0, 'records': 0})
        rows.append([cat['title'], stats['tables'], stats['records'], cat['description']])
    if include_other_count:
        rows.append(['Other system tables', include_other_count, '', 'Supporting configuration tables'])
    rows.append([''])
    rows.append(['Open each sheet tab to view data. Upload is stored on Google Drive when configured.'])
    return rows


def get_academic_calendar_for_backup(cursor):
    """All academic years with terms for Drive folder creation."""
    calendar = []
    try:
        cursor.execute(
            'SELECT id, year_name FROM academic_years ORDER BY start_date DESC, id DESC'
        )
        years = cursor.fetchall() or []
    except Exception:
        return calendar
    for y in years:
        yid = y.get('id') if isinstance(y, dict) else y[0]
        yname = y.get('year_name') if isinstance(y, dict) else y[1]
        terms = []
        try:
            cursor.execute(
                'SELECT id, term_name FROM terms WHERE academic_year_id = %s ORDER BY start_date, id',
                (yid,),
            )
            for t in cursor.fetchall() or []:
                terms.append({
                    'id': t.get('id') if isinstance(t, dict) else t[0],
                    'name': t.get('term_name') if isinstance(t, dict) else t[1],
                })
        except Exception:
            pass
        calendar.append({'id': yid, 'name': yname, 'terms': terms})
    return calendar


def get_current_year_term_for_backup(cursor):
    """Current academic year and term ids/names for uploads."""
    year_id = year_name = term_id = term_name = None
    try:
        cursor.execute(
            "SELECT id, year_name FROM academic_years WHERE is_current = TRUE AND status = 'active' LIMIT 1"
        )
        row = cursor.fetchone()
        if row:
            year_id = row.get('id')
            year_name = row.get('year_name')
        if not year_id:
            cursor.execute('SELECT id, year_name FROM academic_years ORDER BY start_date DESC LIMIT 1')
            row = cursor.fetchone()
            if row:
                year_id = row.get('id')
                year_name = row.get('year_name')
        if year_id:
            cursor.execute(
                """SELECT id, term_name FROM terms
                   WHERE academic_year_id = %s AND is_current = TRUE AND status = 'active' LIMIT 1""",
                (year_id,),
            )
            trow = cursor.fetchone()
            if trow:
                term_id = trow.get('id')
                term_name = trow.get('term_name')
            if not term_id:
                cursor.execute(
                    'SELECT id, term_name FROM terms WHERE academic_year_id = %s ORDER BY start_date DESC LIMIT 1',
                    (year_id,),
                )
                trow = cursor.fetchone()
                if trow:
                    term_id = trow.get('id')
                    term_name = trow.get('term_name')
    except Exception:
        pass
    if not year_name:
        year_name = 'Current year'
    if not term_name:
        term_name = 'Current term'
    return {
        'year_id': year_id,
        'year_name': year_name,
        'term_id': term_id,
        'term_name': term_name,
    }


def ensure_backup_history_table(cursor, connection):
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS backup_history (
            id INT AUTO_INCREMENT PRIMARY KEY,
            filename VARCHAR(255) NOT NULL,
            file_path VARCHAR(500),
            file_size BIGINT,
            table_count INT,
            record_count INT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            created_by VARCHAR(255)
        )
    """)
    connection.commit()
    ensure_backup_drive_schema(cursor, connection)


def record_backup_history(cursor, connection, filename, filepath, file_size, table_count,
                          record_count, created_by, storage_location='local',
                          drive_file_id=None, drive_url=None, meta=None):
    """Insert backup_history row; returns history id or None."""
    ensure_backup_history_table(cursor, connection)
    meta_json = json.dumps(meta) if meta is not None else None
    try:
        cursor.execute("""
            INSERT INTO backup_history (
                filename, file_path, file_size, table_count, record_count, created_by,
                storage_location, google_drive_file_id, google_drive_url, categories_json
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            filename, filepath, file_size, table_count, record_count, created_by,
            storage_location, drive_file_id, drive_url, meta_json,
        ))
        connection.commit()
        return cursor.lastrowid
    except Exception as e1:
        print(f'backup_history insert (full) failed: {e1}')
        try:
            cursor.execute("""
                INSERT INTO backup_history (
                    filename, file_path, file_size, table_count, record_count, created_by
                )
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (filename, filepath, file_size, table_count, record_count, created_by))
            connection.commit()
            return cursor.lastrowid
        except Exception as e2:
            print(f'backup_history insert (basic) failed: {e2}')
            return None


def fetch_backup_history(cursor, connection, limit=20):
    """Return recent backup_history rows."""
    ensure_backup_history_table(cursor, connection)
    cursor.execute(
        'SELECT * FROM backup_history ORDER BY created_at DESC LIMIT %s',
        (limit,),
    )
    return cursor.fetchall() or []


def ensure_backup_drive_schema(cursor, connection):
    """Add Google Drive columns to backup tables if missing."""
    alters = [
        ("backup_history", "storage_location", "VARCHAR(32) DEFAULT 'local'"),
        ("backup_history", "google_drive_file_id", "VARCHAR(128) NULL"),
        ("backup_history", "google_drive_url", "VARCHAR(500) NULL"),
        ("backup_history", "categories_json", "TEXT NULL"),
        ("backup_settings", "google_drive_enabled", "TINYINT(1) DEFAULT 1"),
        ("backup_settings", "google_drive_folder_id", "VARCHAR(128) NULL"),
        ("backup_settings", "google_drive_oauth_token", "LONGTEXT NULL"),
        ("backup_settings", "google_drive_connected_email", "VARCHAR(255) NULL"),
        ("backup_settings", "google_drive_school_root_id", "VARCHAR(128) NULL"),
        ("backup_settings", "google_drive_folder_map", "LONGTEXT NULL"),
        ("backup_settings", "google_drive_folders_status", "VARCHAR(16) NULL"),
        ("backup_settings", "google_drive_folders_error", "TEXT NULL"),
        ("backup_settings", "google_drive_folders_at", "DATETIME NULL"),
    ]
    for table, col, col_def in alters:
        try:
            cursor.execute(f"SHOW COLUMNS FROM `{table}` LIKE %s", (col,))
            if not cursor.fetchone():
                cursor.execute(f"ALTER TABLE `{table}` ADD COLUMN `{col}` {col_def}")
                connection.commit()
        except Exception:
            pass


def _write_sheet(ws, columns, rows, header_fill, header_font):
    from openpyxl.styles import Alignment

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for row_idx, row_data in enumerate(rows, 2):
        if isinstance(row_data, dict):
            values = [row_data.get(c) for c in columns]
        else:
            values = list(row_data)
        for col_idx, value in enumerate(values, 1):
            if isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d %H:%M:%S')
            elif value is None:
                value = ''
            ws.cell(row=row_idx, column=col_idx, value=value)
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)


def build_organized_workbook(cursor, school_name, created_by):
    """
    Build openpyxl workbook with README + categorized sheets.
    Returns (workbook, total_tables, total_records, category_stats dict).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    cursor.execute('SHOW TABLES')
    table_results = cursor.fetchall()
    all_names = []
    for tr in table_results:
        all_names.append(list(tr.values())[0] if isinstance(tr, dict) else tr[0])

    grouped, other_tables = categorize_tables(all_names)
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    used_sheet_names = set()
    category_stats = {c['key']: {'tables': 0, 'records': 0} for c in BACKUP_CATEGORIES}
    total_tables = 0
    total_records = 0

    def export_table(table_name, category_key):
        nonlocal total_tables, total_records
        try:
            cursor.execute(f'SELECT * FROM `{table_name}`')
        except Exception:
            return
        columns = [desc[0] for desc in cursor.description]
        rows = cursor.fetchall()
        if not columns:
            return
        sheet_title = friendly_sheet_name(category_key, table_name, used_sheet_names)
        ws = wb.create_sheet(title=sheet_title)
        _write_sheet(ws, columns, rows, header_fill, header_font)
        n = len(rows)
        total_tables += 1
        total_records += n
        if category_key in category_stats:
            category_stats[category_key]['tables'] += 1
            category_stats[category_key]['records'] += n

    for cat in BACKUP_CATEGORIES:
        for table_name in cat['tables']:
            if table_name in all_names:
                export_table(table_name, cat['key'])

    if other_tables:
        for table_name in other_tables:
            export_table(table_name, 'academic_setup')

    readme = wb.create_sheet(title='README', index=0)
    for row_idx, row in enumerate(
        build_readme_rows(school_name, created_by, category_stats, len(other_tables)), 1
    ):
        for col_idx, val in enumerate(row, 1):
            cell = readme.cell(row=row_idx, column=col_idx, value=val)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif row_idx == 6:
                cell.font = Font(bold=True)

    return wb, total_tables, total_records, category_stats


def build_slice_workbooks(cursor, school_name, created_by, existing_tables=None):
    """
    Build one workbook per Drive slice (fees, payments, exams, …).
    Returns dict drive_path -> {filepath, tables, records, title}.
    """
    import tempfile
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    if existing_tables is None:
        cursor.execute('SHOW TABLES')
        existing_tables = set()
        for tr in cursor.fetchall():
            existing_tables.add(list(tr.values())[0] if isinstance(tr, dict) else tr[0])

    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    out = {}
    tmp_dir = tempfile.mkdtemp(prefix='school_backup_')

    for drive_path, spec in DRIVE_BACKUP_SLICES.items():
        tables = [t for t in spec['tables'] if t in existing_tables]
        if not tables:
            continue
        wb = Workbook()
        wb.remove(wb.active)
        used = set()
        total_records = 0
        for table_name in tables:
            try:
                cursor.execute(f'SELECT * FROM `{table_name}`')
            except Exception:
                continue
            columns = [d[0] for d in cursor.description]
            rows = cursor.fetchall()
            if not columns:
                continue
            cat_key = _CATEGORY_TABLE_MAP.get(table_name, 'academic_setup')
            sheet_title = friendly_sheet_name(cat_key, table_name, used)
            ws = wb.create_sheet(title=sheet_title)
            _write_sheet(ws, columns, rows, header_fill, header_font)
            total_records += len(rows)

        if not wb.sheetnames:
            continue

        readme = wb.create_sheet(title='README', index=0)
        readme.append([spec['title']])
        readme.append(['School', school_name or ''])
        readme.append(['Created', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        readme.append(['Created by', created_by or 'System'])
        readme.append(['Drive folder', drive_path.replace('/', ' → ')])
        readme.append(['Records', total_records])

        filepath = os.path.join(tmp_dir, spec['filename'])
        wb.save(filepath)
        out[drive_path] = {
            'filepath': filepath,
            'tables': len(tables),
            'records': total_records,
            'title': spec['title'],
            'filename': spec['filename'],
        }
    return out, tmp_dir
